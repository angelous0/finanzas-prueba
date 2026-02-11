from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, Header, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime, date, timedelta, timezone
import asyncpg
import io

from database import init_db, close_db, get_pool
from models import (
    Empresa, EmpresaCreate, EmpresaUpdate,
    Moneda, MonedaCreate,
    Categoria, CategoriaCreate, CategoriaUpdate,
    CentroCosto, CentroCostoCreate,
    LineaNegocio, LineaNegocioCreate,
    CuentaFinanciera, CuentaFinancieraCreate, CuentaFinancieraUpdate,
    Tercero, TerceroCreate, TerceroUpdate,
    EmpleadoDetalle, EmpleadoDetalleCreate,
    ArticuloRef, ArticuloRefCreate,
    OC, OCCreate, OCUpdate, OCLinea,
    FacturaProveedor, FacturaProveedorCreate, FacturaProveedorUpdate, FacturaLinea,
    Pago, PagoCreate, PagoDetalle, PagoAplicacion,
    Letra, LetraCreate, LetraUpdate, GenerarLetrasRequest,
    Gasto, GastoCreate, GastoLinea,
    Planilla, PlanillaCreate, PlanillaDetalle,
    Adelanto, AdelantoCreate,
    VentaPOS, CXC, CXCCreate,
    Presupuesto, PresupuestoCreate,
    Conciliacion, ConciliacionCreate, BancoMovRaw, BancoMov,
    DashboardKPIs,
    CuentaContable, CuentaContableCreate, CuentaContableUpdate, ConfigEmpresaContable,
    Asiento, AsientoCreate, AsientoLinea, GenerarAsientoRequest,
    RetencionDetalle
)
from contabilidad import (
    generar_asiento_fprov, generar_asiento_gasto, generar_asiento_pago,
    check_periodo_cerrado, reporte_mayor, reporte_balance, reporte_pnl
)
from odoo_service import OdooService

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Utility function for safe date handling in PostgreSQL queries
def safe_date_param(fecha_value):
    """
    Convert date to string format for PostgreSQL TO_DATE function.
    Handles date objects, datetime objects, and strings.
    Returns a string in 'YYYY-MM-DD' format.
    """
    if fecha_value is None:
        return None
    
    if isinstance(fecha_value, str):
        # If already a string, try to parse and re-format to ensure consistency
        try:
            if 'T' in fecha_value:  # ISO datetime format
                dt = datetime.fromisoformat(fecha_value.replace('Z', '+00:00'))
                return dt.strftime('%Y-%m-%d')
            else:
                # Assume it's already in YYYY-MM-DD format
                return fecha_value
        except:
            return fecha_value
    
    if isinstance(fecha_value, datetime):
        return fecha_value.strftime('%Y-%m-%d')
    
    if isinstance(fecha_value, date):
        return fecha_value.strftime('%Y-%m-%d')
    
    # If it's something else, convert to string and hope for the best
    return str(fecha_value)

app = FastAPI(title="Finanzas 4.0 API", version="1.0.0")

api_router = APIRouter(prefix="/api")


# =====================
# EMPRESA_ID DEPENDENCY
# =====================
async def get_empresa_id(
    empresa_id: Optional[int] = Query(None),
    x_empresa_id: Optional[str] = Header(None),
) -> int:
    """Extract empresa_id from query param (priority) or X-Empresa-Id header.
    Blocks the request if neither is provided."""
    eid = empresa_id or (int(x_empresa_id) if x_empresa_id else None)
    if not eid:
        raise HTTPException(400, "empresa_id es requerido")
    return eid


async def get_next_correlativo(conn, empresa_id: int, tipo_documento: str, prefijo: str) -> str:
    """Atomically get next correlative number for a document type.
    Uses INSERT ... ON CONFLICT ... UPDATE to guarantee uniqueness."""
    row = await conn.fetchrow("""
        INSERT INTO finanzas2.cont_correlativos (empresa_id, tipo_documento, prefijo, ultimo_numero, updated_at)
        VALUES ($1, $2, $3, 1, NOW())
        ON CONFLICT (empresa_id, tipo_documento, prefijo)
        DO UPDATE SET ultimo_numero = finanzas2.cont_correlativos.ultimo_numero + 1, updated_at = NOW()
        RETURNING ultimo_numero
    """, empresa_id, tipo_documento, prefijo)
    return f"{prefijo}{row['ultimo_numero']:05d}"

# =====================
# STARTUP / SHUTDOWN
# =====================
@app.on_event("startup")
async def startup():
    logger.info("Starting Finanzas 4.0 API...")
    await init_db()
    await seed_data()
    await sync_correlativos()
    logger.info("Finanzas 4.0 API started successfully")

@app.on_event("shutdown")
async def shutdown():
    await close_db()
    logger.info("Finanzas 4.0 API shutdown complete")

async def seed_data():
    """Create initial seed data"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if seed data exists
        empresa_count = await conn.fetchval("SELECT COUNT(*) FROM finanzas2.cont_empresa")
        if empresa_count > 0:
            logger.info("Seed data already exists, skipping...")
            return
        
        logger.info("Creating seed data...")
        
        # Insert empresa
        empresa_id = await conn.fetchval("""
            INSERT INTO finanzas2.cont_empresa (nombre, ruc, direccion, telefono, email)
            VALUES ('Mi Empresa S.A.C.', '20123456789', 'Av. Principal 123, Lima', '01-1234567', 'contacto@miempresa.com')
            RETURNING id
        """)
        
        # Insert moneda PEN
        pen_id = await conn.fetchval("""
            INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal)
            VALUES ('PEN', 'Sol Peruano', 'S/', TRUE)
            RETURNING id
        """)
        
        # Insert moneda USD
        await conn.execute("""
            INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal)
            VALUES ('USD', 'Dólar Americano', '$', FALSE)
        """)
        
        # Insert categorías
        await conn.execute("""
            INSERT INTO finanzas2.cont_categoria (empresa_id, codigo, nombre, tipo) VALUES
            ($1, 'ING-001', 'Ventas', 'ingreso'),
            ($1, 'ING-002', 'Otros Ingresos', 'ingreso'),
            ($1, 'EGR-001', 'Compras Mercadería', 'egreso'),
            ($1, 'EGR-002', 'Servicios', 'egreso'),
            ($1, 'EGR-003', 'Planilla', 'egreso'),
            ($1, 'EGR-004', 'Alquileres', 'egreso'),
            ($1, 'EGR-005', 'Servicios Públicos', 'egreso'),
            ($1, 'EGR-006', 'Otros Gastos', 'egreso')
        """, empresa_id)
        
        # Insert centro de costo
        await conn.execute("""
            INSERT INTO finanzas2.cont_centro_costo (empresa_id, codigo, nombre) VALUES
            ($1, 'CC-001', 'Administración'),
            ($1, 'CC-002', 'Ventas'),
            ($1, 'CC-003', 'Operaciones')
        """, empresa_id)
        
        # Insert línea de negocio
        await conn.execute("""
            INSERT INTO finanzas2.cont_linea_negocio (empresa_id, codigo, nombre) VALUES
            ($1, 'LN-001', 'Línea Principal'),
            ($1, 'LN-002', 'Línea Secundaria')
        """, empresa_id)
        
        # Insert cuenta bancaria
        await conn.execute("""
            INSERT INTO finanzas2.cont_cuenta_financiera (empresa_id, nombre, tipo, banco, numero_cuenta, moneda_id, saldo_actual)
            VALUES ($1, 'Cuenta BCP Soles', 'banco', 'BCP', '191-12345678-0-12', $2, 0)
        """, empresa_id, pen_id)
        
        # Insert caja
        await conn.execute("""
            INSERT INTO finanzas2.cont_cuenta_financiera (empresa_id, nombre, tipo, moneda_id, saldo_actual)
            VALUES ($1, 'Caja Chica', 'caja', $2, 0)
        """, empresa_id, pen_id)
        
        # Insert proveedor
        await conn.execute("""
            INSERT INTO finanzas2.cont_tercero (empresa_id, tipo_documento, numero_documento, nombre, es_proveedor, terminos_pago_dias)
            VALUES ($1, 'RUC', '20987654321', 'Proveedor Demo S.A.C.', TRUE, 30)
        """, empresa_id)
        
        logger.info("Seed data created successfully")


async def sync_correlativos():
    """Sync cont_correlativos with existing document numbers on startup.
    This ensures the sequence counters are >= max existing numbers."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        year = datetime.now().year

        # Map: (table, numero_col, tipo_documento, prefix_template)
        doc_types = [
            ('cont_oc', 'numero', 'oc', f'OC-{year}-'),
            ('cont_factura_proveedor', 'numero', 'factura_proveedor', f'FP-{year}-'),
            ('cont_pago', 'numero', 'pago_ingreso', f'PAG-I-{year}-'),
            ('cont_pago', 'numero', 'pago_egreso', f'PAG-E-{year}-'),
            ('cont_gasto', 'numero', 'gasto', f'GAS-{year}-'),
        ]

        for table, col, tipo_doc, prefix in doc_types:
            # For pago, filter by tipo
            tipo_filter = ""
            if tipo_doc == 'pago_ingreso':
                tipo_filter = "AND tipo = 'ingreso'"
            elif tipo_doc == 'pago_egreso':
                tipo_filter = "AND tipo = 'egreso'"

            rows = await conn.fetch(f"""
                SELECT empresa_id, MAX(
                    CASE WHEN {col} LIKE $1 || '%' 
                    THEN CAST(SPLIT_PART({col}, '-', {len(prefix.split('-'))}) AS INTEGER) 
                    ELSE 0 END
                ) as max_num
                FROM finanzas2.{table}
                WHERE {col} LIKE $1 || '%' {tipo_filter}
                GROUP BY empresa_id
            """, prefix)

            for row in rows:
                if row['max_num'] and row['max_num'] > 0:
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_correlativos (empresa_id, tipo_documento, prefijo, ultimo_numero, updated_at)
                        VALUES ($1, $2, $3, $4, NOW())
                        ON CONFLICT (empresa_id, tipo_documento, prefijo)
                        DO UPDATE SET ultimo_numero = GREATEST(finanzas2.cont_correlativos.ultimo_numero, $4),
                                      updated_at = NOW()
                    """, row['empresa_id'], tipo_doc, prefix, row['max_num'])

        logger.info("Correlatives synced with existing data")

# =====================
# HEALTH CHECK
# =====================
@api_router.get("/")
async def root():
    return {"message": "Finanzas 4.0 API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.fetchval("SELECT 1")
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

# =====================
# DASHBOARD
# =====================
@api_router.get("/dashboard/kpis", response_model=DashboardKPIs)
async def get_dashboard_kpis(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Total CxP pendiente
        total_cxp = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) 
            FROM finanzas2.cont_cxp WHERE estado NOT IN ('pagado', 'anulada') AND empresa_id = $1
        """, empresa_id) or 0
        
        # Total CxC pendiente
        total_cxc = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) 
            FROM finanzas2.cont_cxc WHERE estado NOT IN ('pagado', 'anulada') AND empresa_id = $1
        """, empresa_id) or 0
        
        # Total letras pendientes
        total_letras = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) 
            FROM finanzas2.cont_letra WHERE estado IN ('pendiente', 'parcial') AND empresa_id = $1
        """, empresa_id) or 0
        
        # Saldo bancos
        saldo_bancos = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_actual), 0) 
            FROM finanzas2.cont_cuenta_financiera WHERE activo = TRUE AND empresa_id = $1
        """, empresa_id) or 0
        
        # Ventas del mes
        inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ventas_mes = await conn.fetchval("""
            SELECT COALESCE(SUM(amount_total), 0) 
            FROM finanzas2.cont_venta_pos 
            WHERE date_order >= $1 AND estado_local = 'confirmada' AND empresa_id = $2
        """, inicio_mes, empresa_id) or 0
        
        # Gastos del mes
        gastos_mes = await conn.fetchval("""
            SELECT COALESCE(SUM(total), 0) 
            FROM finanzas2.cont_gasto WHERE fecha >= $1 AND empresa_id = $2
        """, inicio_mes.date(), empresa_id) or 0
        
        # Facturas pendientes
        facturas_pendientes = await conn.fetchval("""
            SELECT COUNT(*) FROM finanzas2.cont_factura_proveedor 
            WHERE estado IN ('pendiente', 'parcial') AND empresa_id = $1
        """, empresa_id) or 0
        
        # Letras por vencer (próximos 7 días)
        fecha_limite = datetime.now().date() + timedelta(days=7)
        letras_por_vencer = await conn.fetchval("""
            SELECT COUNT(*) FROM finanzas2.cont_letra 
            WHERE estado IN ('pendiente', 'parcial') AND fecha_vencimiento <= $1 AND empresa_id = $2
        """, fecha_limite, empresa_id) or 0
        
        return DashboardKPIs(
            total_cxp=float(total_cxp),
            total_cxc=float(total_cxc),
            total_letras_pendientes=float(total_letras),
            saldo_bancos=float(saldo_bancos),
            ventas_mes=float(ventas_mes),
            gastos_mes=float(gastos_mes),
            facturas_pendientes=facturas_pendientes,
            letras_por_vencer=letras_por_vencer
        )

# =====================
# EMPRESAS
# =====================
@api_router.get("/empresas", response_model=List[Empresa])
async def list_empresas():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_empresa ORDER BY nombre")
        return [dict(r) for r in rows]

@api_router.post("/empresas", response_model=Empresa)
async def create_empresa(data: EmpresaCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_empresa (nombre, ruc, direccion, telefono, email, logo_url, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7)
            RETURNING *
        """, data.nombre, data.ruc, data.direccion, data.telefono, data.email, data.logo_url, data.activo)
        
        empresa_id = row['id']
        
        # Create base monedas if none exist
        moneda_count = await conn.fetchval("SELECT COUNT(*) FROM finanzas2.cont_moneda")
        if moneda_count == 0:
            await conn.execute("INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal) VALUES ('PEN', 'Sol Peruano', 'S/', TRUE)")
            await conn.execute("INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal) VALUES ('USD', 'Dólar Americano', '$', FALSE)")
        
        # Create default categorias for new empresa
        cat_count = await conn.fetchval("SELECT COUNT(*) FROM finanzas2.cont_categoria WHERE empresa_id = $1", empresa_id)
        if cat_count == 0:
            await conn.execute("""
                INSERT INTO finanzas2.cont_categoria (empresa_id, codigo, nombre, tipo) VALUES
                ($1, 'ING-001', 'Ventas', 'ingreso'),
                ($1, 'ING-002', 'Otros Ingresos', 'ingreso'),
                ($1, 'EGR-001', 'Compras Mercadería', 'egreso'),
                ($1, 'EGR-002', 'Servicios', 'egreso'),
                ($1, 'EGR-003', 'Planilla', 'egreso'),
                ($1, 'EGR-004', 'Alquileres', 'egreso'),
                ($1, 'EGR-005', 'Servicios Públicos', 'egreso'),
                ($1, 'EGR-006', 'Otros Gastos', 'egreso')
            """, empresa_id)
        
        return dict(row)

@api_router.put("/empresas/{id}", response_model=Empresa)
async def update_empresa(id: int, data: EmpresaUpdate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Build dynamic update
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        
        if not updates:
            raise HTTPException(400, "No fields to update")
        
        values.append(id)
        query = f"UPDATE finanzas2.cont_empresa SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Empresa not found")
        return dict(row)

@api_router.delete("/empresas/{id}")
async def delete_empresa(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_empresa WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Empresa not found")
        return {"message": "Empresa deleted"}

# =====================
# MONEDAS
# =====================
@api_router.get("/monedas", response_model=List[Moneda])
async def list_monedas():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_moneda ORDER BY codigo")
        return [dict(r) for r in rows]

@api_router.post("/monedas", response_model=Moneda)
async def create_moneda(data: MonedaCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_moneda (codigo, nombre, simbolo, es_principal, activo)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING *
        """, data.codigo, data.nombre, data.simbolo, data.es_principal, data.activo)
        return dict(row)

@api_router.delete("/monedas/{id}")
async def delete_moneda(id: int):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_moneda WHERE id = $1", id)
        if result == "DELETE 0":
            raise HTTPException(404, "Moneda not found")
        return {"message": "Moneda deleted"}

# =====================
# CATEGORIAS
# =====================
@api_router.get("/categorias")
async def list_categorias(empresa_id: int = Depends(get_empresa_id), tipo: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        base_query = """
            SELECT c.*, cp.nombre as padre_nombre
            FROM finanzas2.cont_categoria c
            LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
            WHERE c.empresa_id = $1
        """
        if tipo:
            rows = await conn.fetch(
                base_query + " AND c.tipo = $2 ORDER BY c.nombre", empresa_id, tipo
            )
        else:
            rows = await conn.fetch(base_query + " ORDER BY c.tipo, c.nombre", empresa_id)
        result = []
        for r in rows:
            item = dict(r)
            padre_nombre = item.pop("padre_nombre", None)
            item["nombre_completo"] = f"{padre_nombre} > {item['nombre']}" if padre_nombre else item["nombre"]
            result.append(item)
        return result

@api_router.post("/categorias", response_model=Categoria)
async def create_categoria(data: CategoriaCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_categoria (empresa_id, codigo, nombre, tipo, padre_id, descripcion, cuenta_gasto_id, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            RETURNING *
        """, empresa_id, data.codigo, data.nombre, data.tipo, data.padre_id, data.descripcion, data.cuenta_gasto_id, data.activo)
        return dict(row)

@api_router.put("/categorias/{id}", response_model=Categoria)
async def update_categoria(id: int, data: CategoriaUpdate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        if not updates:
            raise HTTPException(400, "No fields to update")
        values.append(empresa_id)
        values.append(id)
        query = f"UPDATE finanzas2.cont_categoria SET {', '.join(updates)}, updated_at = NOW() WHERE empresa_id = ${idx} AND id = ${idx+1} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Categoria not found")
        return dict(row)

@api_router.delete("/categorias/{id}")
async def delete_categoria(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_categoria WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if result == "DELETE 0":
            raise HTTPException(404, "Categoria not found")
        return {"message": "Categoria deleted"}

# =====================
# CENTROS DE COSTO
# =====================
@api_router.get("/centros-costo", response_model=List[CentroCosto])
async def list_centros_costo(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_centro_costo WHERE empresa_id = $1 ORDER BY nombre", empresa_id)
        return [dict(r) for r in rows]

@api_router.post("/centros-costo", response_model=CentroCosto)
async def create_centro_costo(data: CentroCostoCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_centro_costo (empresa_id, codigo, nombre, descripcion, activo)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING *
        """, empresa_id, data.codigo, data.nombre, data.descripcion, data.activo)
        return dict(row)

@api_router.delete("/centros-costo/{id}")
async def delete_centro_costo(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_centro_costo WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if result == "DELETE 0":
            raise HTTPException(404, "Centro de costo not found")
        return {"message": "Centro de costo deleted"}

@api_router.put("/centros-costo/{id}", response_model=CentroCosto)
async def update_centro_costo(id: int, data: CentroCostoCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            UPDATE finanzas2.cont_centro_costo SET codigo = $1, nombre = $2, descripcion = $3
            WHERE id = $4 AND empresa_id = $5 RETURNING *
        """, data.codigo, data.nombre, data.descripcion, id, empresa_id)
        if not row:
            raise HTTPException(404, "Centro de costo not found")
        return dict(row)

# =====================
# LINEAS DE NEGOCIO
# =====================
@api_router.get("/lineas-negocio", response_model=List[LineaNegocio])
async def list_lineas_negocio(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("SELECT * FROM finanzas2.cont_linea_negocio WHERE empresa_id = $1 ORDER BY nombre", empresa_id)
        return [dict(r) for r in rows]

@api_router.post("/lineas-negocio", response_model=LineaNegocio)
async def create_linea_negocio(data: LineaNegocioCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_linea_negocio (empresa_id, codigo, nombre, descripcion, activo)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING *
        """, empresa_id, data.codigo, data.nombre, data.descripcion, data.activo)
        return dict(row)

@api_router.delete("/lineas-negocio/{id}")
async def delete_linea_negocio(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_linea_negocio WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if result == "DELETE 0":
            raise HTTPException(404, "Línea de negocio not found")
        return {"message": "Línea de negocio deleted"}

@api_router.put("/lineas-negocio/{id}", response_model=LineaNegocio)
async def update_linea_negocio(id: int, data: LineaNegocioCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            UPDATE finanzas2.cont_linea_negocio SET codigo = $1, nombre = $2, descripcion = $3
            WHERE id = $4 AND empresa_id = $5 RETURNING *
        """, data.codigo, data.nombre, data.descripcion, id, empresa_id)
        if not row:
            raise HTTPException(404, "Línea de negocio not found")
        return dict(row)

# =====================
# CUENTAS FINANCIERAS
# =====================
@api_router.get("/cuentas-financieras", response_model=List[CuentaFinanciera])
async def list_cuentas_financieras(empresa_id: int = Depends(get_empresa_id), tipo: Optional[str] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        query = """
            SELECT cf.*, m.codigo as moneda_codigo
            FROM finanzas2.cont_cuenta_financiera cf
            LEFT JOIN finanzas2.cont_moneda m ON cf.moneda_id = m.id
            WHERE cf.empresa_id = $1
        """
        if tipo:
            query += " AND cf.tipo = $2"
            rows = await conn.fetch(query + " ORDER BY cf.nombre", empresa_id, tipo)
        else:
            rows = await conn.fetch(query + " ORDER BY cf.nombre", empresa_id)
        return [dict(r) for r in rows]

@api_router.post("/cuentas-financieras", response_model=CuentaFinanciera)
async def create_cuenta_financiera(data: CuentaFinancieraCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_cuenta_financiera 
            (empresa_id, nombre, tipo, banco, numero_cuenta, cci, moneda_id, saldo_actual, saldo_inicial, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $8, $9)
            RETURNING *
        """, empresa_id, data.nombre, data.tipo, data.banco, data.numero_cuenta, data.cci, 
            data.moneda_id, data.saldo_actual, data.activo)
        return dict(row)

@api_router.put("/cuentas-financieras/{id}", response_model=CuentaFinanciera)
async def update_cuenta_financiera(id: int, data: CuentaFinancieraUpdate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        updates = []
        values = []
        idx = 1
        saldo_inicial_changed = False
        new_saldo_inicial = None
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            if field == 'saldo_inicial':
                saldo_inicial_changed = True
                new_saldo_inicial = value
            idx += 1
        if not updates:
            raise HTTPException(400, "No fields to update")
        values.append(empresa_id)
        values.append(id)
        query = f"UPDATE finanzas2.cont_cuenta_financiera SET {', '.join(updates)}, updated_at = NOW() WHERE empresa_id = ${idx} AND id = ${idx+1} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Cuenta financiera not found")
        
        # If saldo_inicial changed, recalculate saldo_actual
        if saldo_inicial_changed and new_saldo_inicial is not None:
            ingresos = await conn.fetchval("""
                SELECT COALESCE(SUM(pd.monto), 0) FROM finanzas2.cont_pago_detalle pd
                JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
                WHERE pd.cuenta_financiera_id = $1 AND p.tipo = 'ingreso'
            """, id)
            egresos = await conn.fetchval("""
                SELECT COALESCE(SUM(pd.monto), 0) FROM finanzas2.cont_pago_detalle pd
                JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
                WHERE pd.cuenta_financiera_id = $1 AND p.tipo = 'egreso'
            """, id)
            nuevo_saldo = float(new_saldo_inicial) + float(ingresos) - float(egresos)
            row = await conn.fetchrow(
                "UPDATE finanzas2.cont_cuenta_financiera SET saldo_actual = $1 WHERE id = $2 RETURNING *",
                nuevo_saldo, id
            )
        
        return dict(row)

@api_router.delete("/cuentas-financieras/{id}")
async def delete_cuenta_financiera(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("DELETE FROM finanzas2.cont_cuenta_financiera WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if result == "DELETE 0":
            raise HTTPException(404, "Cuenta financiera not found")
        return {"message": "Cuenta financiera deleted"}

@api_router.get("/cuentas-financieras/{id}/kardex")
async def get_kardex_cuenta(
    id: int,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    """Kardex bancario: historial de movimientos de una cuenta con saldo acumulado."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        cuenta = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_cuenta_financiera WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not cuenta:
            raise HTTPException(404, "Cuenta no encontrada")
        
        conditions = ["pd.cuenta_financiera_id = $1", "pd.empresa_id = $2"]
        params = [id, empresa_id]
        idx = 3
        
        if fecha_desde:
            conditions.append(f"p.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"p.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        rows = await conn.fetch(f"""
            SELECT p.id as pago_id, p.numero, p.tipo, p.fecha, p.notas,
                   pd.medio_pago, pd.monto, pd.referencia
            FROM finanzas2.cont_pago_detalle pd
            JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
            WHERE {' AND '.join(conditions)}
            ORDER BY p.fecha ASC, pd.id ASC
        """, *params)
        
        saldo_base = float(cuenta.get('saldo_inicial') or 0)
        # If filtering by date, compute saldo before the period
        if fecha_desde:
            pre_ingresos = await conn.fetchval("""
                SELECT COALESCE(SUM(pd.monto), 0) FROM finanzas2.cont_pago_detalle pd
                JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
                WHERE pd.cuenta_financiera_id = $1 AND pd.empresa_id = $2 AND p.tipo = 'ingreso' AND p.fecha < $3
            """, id, empresa_id, fecha_desde) or 0
            pre_egresos = await conn.fetchval("""
                SELECT COALESCE(SUM(pd.monto), 0) FROM finanzas2.cont_pago_detalle pd
                JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
                WHERE pd.cuenta_financiera_id = $1 AND pd.empresa_id = $2 AND p.tipo = 'egreso' AND p.fecha < $3
            """, id, empresa_id, fecha_desde) or 0
            saldo_periodo = saldo_base + float(pre_ingresos) - float(pre_egresos)
        else:
            saldo_periodo = saldo_base
        
        movimientos = []
        saldo = saldo_periodo
        for r in rows:
            monto = float(r['monto'])
            if r['tipo'] == 'ingreso':
                saldo += monto
            else:
                saldo -= monto
            movimientos.append({
                "fecha": str(r['fecha']),
                "numero": r['numero'],
                "tipo": r['tipo'],
                "concepto": r['notas'] or r['medio_pago'],
                "medio_pago": r['medio_pago'],
                "referencia": r['referencia'],
                "ingreso": monto if r['tipo'] == 'ingreso' else 0,
                "egreso": monto if r['tipo'] == 'egreso' else 0,
                "saldo": round(saldo, 2)
            })
        
        # Compute real balance
        total_ingresos = sum(m['ingreso'] for m in movimientos)
        total_egresos = sum(m['egreso'] for m in movimientos)
        
        return {
            "cuenta": dict(cuenta),
            "saldo_inicial": round(saldo_periodo, 2),
            "total_ingresos": round(total_ingresos, 2),
            "total_egresos": round(total_egresos, 2),
            "saldo_final": round(saldo, 2),
            "movimientos": movimientos
        }

@api_router.post("/cuentas-financieras/recalcular-saldos")
async def recalcular_saldos(empresa_id: int = Depends(get_empresa_id)):
    """Recalculate real balance for all accounts based on pago_detalle movements."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        cuentas = await conn.fetch(
            "SELECT id, saldo_actual FROM finanzas2.cont_cuenta_financiera WHERE empresa_id = $1", empresa_id)
        
        results = []
        for cuenta in cuentas:
            cid = cuenta['id']
            saldo_anterior = float(cuenta['saldo_actual'])
            
            ingresos = await conn.fetchval("""
                SELECT COALESCE(SUM(pd.monto), 0) FROM finanzas2.cont_pago_detalle pd
                JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
                WHERE pd.cuenta_financiera_id = $1 AND pd.empresa_id = $2 AND p.tipo = 'ingreso'
            """, cid, empresa_id) or 0
            
            egresos = await conn.fetchval("""
                SELECT COALESCE(SUM(pd.monto), 0) FROM finanzas2.cont_pago_detalle pd
                JOIN finanzas2.cont_pago p ON pd.pago_id = p.id
                WHERE pd.cuenta_financiera_id = $1 AND pd.empresa_id = $2 AND p.tipo = 'egreso'
            """, cid, empresa_id) or 0
            
            nuevo_saldo = float(ingresos) - float(egresos)
            
            # Get saldo_inicial (the base balance set when creating the account)
            saldo_base = await conn.fetchval("""
                SELECT COALESCE(saldo_inicial, 0) FROM finanzas2.cont_cuenta_financiera
                WHERE id = $1 AND empresa_id = $2
            """, cid, empresa_id) or 0
            
            nuevo_saldo = float(saldo_base) + float(ingresos) - float(egresos)
            
            await conn.execute("""
                UPDATE finanzas2.cont_cuenta_financiera SET saldo_actual = $1, updated_at = NOW()
                WHERE id = $2 AND empresa_id = $3
            """, nuevo_saldo, cid, empresa_id)
            
            results.append({"cuenta_id": cid, "saldo_anterior": saldo_anterior, "saldo_nuevo": round(nuevo_saldo, 2)})
        
        return {"message": f"Saldos recalculados para {len(results)} cuentas", "cuentas": results}


# =====================
# TERCEROS (Proveedores, Clientes, Personal)
# =====================
@api_router.get("/terceros", response_model=List[Tercero])
async def list_terceros(
    empresa_id: int = Depends(get_empresa_id),
    es_cliente: Optional[bool] = None,
    es_proveedor: Optional[bool] = None,
    es_personal: Optional[bool] = None,
    search: Optional[str] = None
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["activo = TRUE", "empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if es_cliente is not None:
            conditions.append(f"es_cliente = ${idx}")
            params.append(es_cliente)
            idx += 1
        if es_proveedor is not None:
            conditions.append(f"es_proveedor = ${idx}")
            params.append(es_proveedor)
            idx += 1
        if es_personal is not None:
            conditions.append(f"es_personal = ${idx}")
            params.append(es_personal)
            idx += 1
        if search:
            conditions.append(f"(nombre ILIKE ${idx} OR numero_documento ILIKE ${idx})")
            params.append(f"%{search}%")
            idx += 1
        
        query = f"SELECT * FROM finanzas2.cont_tercero WHERE {' AND '.join(conditions)} ORDER BY nombre"
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.get("/terceros/{id}", response_model=Tercero)
async def get_tercero(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("SELECT * FROM finanzas2.cont_tercero WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not row:
            raise HTTPException(404, "Tercero not found")
        return dict(row)

@api_router.post("/terceros", response_model=Tercero)
async def create_tercero(data: TerceroCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_tercero 
            (empresa_id, tipo_documento, numero_documento, nombre, nombre_comercial, direccion, telefono, email,
             es_cliente, es_proveedor, es_personal, terminos_pago_dias, limite_credito, notas, activo)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15)
            RETURNING *
        """, empresa_id, data.tipo_documento, data.numero_documento, data.nombre, data.nombre_comercial,
            data.direccion, data.telefono, data.email, data.es_cliente, data.es_proveedor,
            data.es_personal, data.terminos_pago_dias, data.limite_credito, data.notas, data.activo)
        return dict(row)

@api_router.put("/terceros/{id}", response_model=Tercero)
async def update_tercero(id: int, data: TerceroUpdate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        if not updates:
            raise HTTPException(400, "No fields to update")
        values.append(empresa_id)
        values.append(id)
        query = f"UPDATE finanzas2.cont_tercero SET {', '.join(updates)}, updated_at = NOW() WHERE empresa_id = ${idx} AND id = ${idx+1} RETURNING *"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Tercero not found")
        return dict(row)

@api_router.delete("/terceros/{id}")
async def delete_tercero(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        result = await conn.execute("UPDATE finanzas2.cont_tercero SET activo = FALSE WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if result == "UPDATE 0":
            raise HTTPException(404, "Tercero not found")
        return {"message": "Tercero deactivated"}

# Aliases for separated views
@api_router.get("/proveedores", response_model=List[Tercero])
async def list_proveedores(empresa_id: int = Depends(get_empresa_id), search: Optional[str] = None):
    return await list_terceros(empresa_id=empresa_id, es_proveedor=True, search=search)

@api_router.get("/clientes", response_model=List[Tercero])
async def list_clientes(empresa_id: int = Depends(get_empresa_id), search: Optional[str] = None):
    return await list_terceros(empresa_id=empresa_id, es_cliente=True, search=search)

@api_router.get("/empleados")
async def list_empleados(empresa_id: int = Depends(get_empresa_id), search: Optional[str] = None):
    """Get empleados with their salary details"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["t.activo = TRUE", "t.es_personal = TRUE", "t.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if search:
            conditions.append(f"(t.nombre ILIKE ${idx} OR t.numero_documento ILIKE ${idx})")
            params.append(f"%{search}%")
            idx += 1
        
        query = f"""
            SELECT t.*, 
                   ed.salario_base,
                   ed.cargo,
                   ed.fecha_ingreso,
                   ed.cuenta_bancaria,
                   ed.banco,
                   ed.centro_costo_id,
                   ed.linea_negocio_id,
                   cc.nombre as centro_costo_nombre,
                   ln.nombre as linea_negocio_nombre
            FROM finanzas2.cont_tercero t
            LEFT JOIN finanzas2.cont_empleado_detalle ed ON t.id = ed.tercero_id
            LEFT JOIN finanzas2.cont_centro_costo cc ON ed.centro_costo_id = cc.id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON ed.linea_negocio_id = ln.id
            WHERE {' AND '.join(conditions)}
            ORDER BY t.nombre
        """
        rows = await conn.fetch(query, *params)
        
        # Convert Decimal to float for salario_base
        result = []
        for r in rows:
            emp = dict(r)
            if emp.get('salario_base') is not None:
                emp['salario_base'] = float(emp['salario_base'])
            result.append(emp)
        
        return result

# Empleado detalle endpoints
@api_router.post("/empleados/{tercero_id}/detalle", response_model=EmpleadoDetalle)
async def create_or_update_empleado_detalle(tercero_id: int, data: EmpleadoDetalleCreate, empresa_id: int = Depends(get_empresa_id)):
    """Create or update empleado detalle (salary, position, etc.)"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if empleado exists
        tercero = await conn.fetchrow("SELECT * FROM finanzas2.cont_tercero WHERE id = $1 AND es_personal = TRUE", tercero_id)
        if not tercero:
            raise HTTPException(404, "Empleado no encontrado")
        
        # Check if detalle already exists
        existing = await conn.fetchrow("SELECT * FROM finanzas2.cont_empleado_detalle WHERE tercero_id = $1", tercero_id)
        
        if existing:
            # Update existing
            row = await conn.fetchrow("""
                UPDATE finanzas2.cont_empleado_detalle 
                SET fecha_ingreso = $1, cargo = $2, salario_base = $3, 
                    cuenta_bancaria = $4, banco = $5, activo = $6,
                    centro_costo_id = $7, linea_negocio_id = $8
                WHERE tercero_id = $9
                RETURNING *
            """, data.fecha_ingreso, data.cargo, data.salario_base,
                data.cuenta_bancaria, data.banco, data.activo,
                data.centro_costo_id, data.linea_negocio_id, tercero_id)
        else:
            # Create new
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_empleado_detalle 
                (tercero_id, fecha_ingreso, cargo, salario_base, cuenta_bancaria, banco, activo, 
                 centro_costo_id, linea_negocio_id, empresa_id)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                RETURNING *
            """, tercero_id, data.fecha_ingreso, data.cargo, data.salario_base,
                data.cuenta_bancaria, data.banco, data.activo,
                data.centro_costo_id, data.linea_negocio_id, empresa_id)
        
        return dict(row)

@api_router.get("/empleados/{tercero_id}/detalle", response_model=EmpleadoDetalle)
async def get_empleado_detalle(tercero_id: int, empresa_id: int = Depends(get_empresa_id)):
    """Get empleado detalle"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("SELECT * FROM finanzas2.cont_empleado_detalle WHERE tercero_id = $1", tercero_id)
        if not row:
            raise HTTPException(404, "Detalle de empleado no encontrado")
        return dict(row)

# =====================
# ARTICULOS REF
# =====================
@api_router.get("/articulos", response_model=List[ArticuloRef])
async def list_articulos(search: Optional[str] = None, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # First try to get from produccion.prod_inventario if exists
        try:
            inv_rows = await conn.fetch("""
                SELECT id, id as prod_inventario_id, 
                       COALESCE(codigo, '') as codigo,
                       COALESCE(nombre, descripcion, 'Sin nombre') as nombre,
                       descripcion,
                       COALESCE(precio_ref, 0) as precio_referencia,
                       TRUE as activo,
                       NOW() as created_at
                FROM produccion.prod_inventario
                WHERE ($1::text IS NULL OR nombre ILIKE $1 OR codigo ILIKE $1)
                LIMIT 100
            """, f"%{search}%" if search else None)
            if inv_rows:
                return [dict(r) for r in inv_rows]
        except Exception as e:
            logger.warning(f"Could not fetch from prod_inventario: {e}")
        
        # Fallback to local articulo_ref
        query = "SELECT * FROM finanzas2.cont_articulo_ref WHERE activo = TRUE"
        if search:
            query += " AND (nombre ILIKE $1 OR codigo ILIKE $1)"
            rows = await conn.fetch(query + " ORDER BY nombre LIMIT 100", f"%{search}%")
        else:
            rows = await conn.fetch(query + " ORDER BY nombre LIMIT 100")
        return [dict(r) for r in rows]

# =====================
# INVENTARIO (produccion.prod_inventario)
# =====================
@api_router.get("/inventario")
async def list_inventario(search: Optional[str] = None, empresa_id: int = Depends(get_empresa_id)):
    """Get items from produccion.prod_inventario"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            search_param = f"%{search}%" if search else None
            rows = await conn.fetch("""
                SELECT id, codigo, nombre, descripcion, categoria, unidad_medida,
                       COALESCE(stock_actual, 0) as stock_actual,
                       COALESCE(stock_minimo, 0) as stock_minimo,
                       COALESCE(precio_ref, 0) as precio_ref,
                       COALESCE(costo_compra, 0) as costo_compra,
                       modelo, marca, activo
                FROM produccion.prod_inventario
                WHERE ($1::text IS NULL OR nombre ILIKE $1 OR codigo ILIKE $1 OR descripcion ILIKE $1)
                ORDER BY nombre
                LIMIT 200
            """, search_param)
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Error fetching inventario: {e}")
            return []

# =====================
# MODELOS/CORTES (produccion.prod_registros + prod_modelos)
# =====================
@api_router.get("/modelos-cortes")
async def list_modelos_cortes(search: Optional[str] = None, empresa_id: int = Depends(get_empresa_id)):
    """Get modelos/cortes from produccion.prod_registros joining with prod_modelos"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            search_param = f"%{search}%" if search else None
            rows = await conn.fetch("""
                SELECT r.id, r.n_corte, r.modelo_id, r.estado,
                       m.nombre as modelo_nombre,
                       CONCAT(m.nombre, ' - Corte ', r.n_corte) as display_name
                FROM produccion.prod_registros r
                LEFT JOIN produccion.prod_modelos m ON r.modelo_id = m.id
                WHERE ($1::text IS NULL OR m.nombre ILIKE $1 OR r.n_corte ILIKE $1)
                ORDER BY r.fecha_creacion DESC
                LIMIT 200
            """, search_param)
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Error fetching modelos/cortes: {e}")
            return []

@api_router.get("/modelos")
async def list_modelos(search: Optional[str] = None, empresa_id: int = Depends(get_empresa_id)):
    """Get modelos from produccion.prod_modelos"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            search_param = f"%{search}%" if search else None
            rows = await conn.fetch("""
                SELECT id, nombre
                FROM produccion.prod_modelos
                WHERE ($1::text IS NULL OR nombre ILIKE $1)
                ORDER BY nombre
                LIMIT 100
            """, search_param)
            return [dict(r) for r in rows]
        except Exception as e:
            logger.error(f"Error fetching modelos: {e}")
            return []

@api_router.post("/articulos", response_model=ArticuloRef)
async def create_articulo(data: ArticuloRefCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_articulo_ref 
            (prod_inventario_id, codigo, nombre, descripcion, precio_referencia, activo, empresa_id)
            VALUES ($1, $2, $3, $4, $5, $6, $7)
            RETURNING *
        """, data.prod_inventario_id, data.codigo, data.nombre, data.descripcion, 
            data.precio_referencia, data.activo, empresa_id)
        return dict(row)

# =====================
# ORDENES DE COMPRA
# =====================
async def generate_oc_number(conn, empresa_id: int) -> str:
    """Generate next OC number using secure correlatives"""
    year = datetime.now().year
    prefijo = f"OC-{year}-"
    return await get_next_correlativo(conn, empresa_id, 'oc', prefijo)

@api_router.get("/ordenes-compra", response_model=List[OC])
async def list_ordenes_compra(
    estado: Optional[str] = None,
    proveedor_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["oc.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if estado:
            conditions.append(f"oc.estado = ${idx}")
            params.append(estado)
            idx += 1
        if proveedor_id:
            conditions.append(f"oc.proveedor_id = ${idx}")
            params.append(proveedor_id)
            idx += 1
        if fecha_desde:
            conditions.append(f"oc.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"oc.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT oc.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_oc oc
            LEFT JOIN finanzas2.cont_tercero t ON oc.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON oc.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY oc.fecha DESC, oc.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            oc_dict = dict(row)
            # Get lines
            lineas = await conn.fetch("""
                SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1 ORDER BY id
            """, row['id'])
            oc_dict['lineas'] = [dict(l) for l in lineas]
            result.append(oc_dict)
        
        return result

@api_router.get("/ordenes-compra/{id}", response_model=OC)
async def get_orden_compra(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT oc.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_oc oc
            LEFT JOIN finanzas2.cont_tercero t ON oc.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON oc.moneda_id = m.id
            WHERE oc.id = $1 AND oc.empresa_id = $2
        """, id, empresa_id)
        
        if not row:
            raise HTTPException(404, "Orden de compra not found")
        
        oc_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1 ORDER BY id
        """, id)
        oc_dict['lineas'] = [dict(l) for l in lineas]
        
        return oc_dict

@api_router.post("/ordenes-compra", response_model=OC)
async def create_orden_compra(data: OCCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            numero = await generate_oc_number(conn, empresa_id)
            
            # Calculate totals
            subtotal = 0
            igv = 0
            for linea in data.lineas:
                if data.igv_incluido and linea.igv_aplica:
                    # Price includes IGV: extract base and tax
                    base = linea.cantidad * linea.precio_unitario / 1.18
                    linea_igv = linea.cantidad * linea.precio_unitario - base
                    subtotal += base
                    igv += linea_igv
                else:
                    linea_subtotal = linea.cantidad * linea.precio_unitario
                    subtotal += linea_subtotal
                    if linea.igv_aplica:
                        igv += linea_subtotal * 0.18
            total = subtotal + igv
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_oc 
                (empresa_id, numero, fecha, proveedor_id, moneda_id, estado, subtotal, igv, total, notas)
                VALUES ($1, $2, $3, $4, $5, 'borrador', $6, $7, $8, $9)
                RETURNING *
            """, empresa_id, numero, data.fecha, data.proveedor_id, data.moneda_id, subtotal, igv, total, data.notas)
            
            oc_id = row['id']
            
            # Insert lines
            for linea in data.lineas:
                if data.igv_incluido and linea.igv_aplica:
                    linea_subtotal = linea.cantidad * linea.precio_unitario / 1.18
                else:
                    linea_subtotal = linea.cantidad * linea.precio_unitario
                
                # Handle articulo_id: if it's a UUID (from prod_inventario), set to None
                # and store the UUID in notas field temporarily
                articulo_id_value = None
                if linea.articulo_id:
                    # Try to parse as integer, if fails it's a UUID
                    try:
                        articulo_id_value = int(linea.articulo_id)
                    except (ValueError, TypeError):
                        # It's a UUID, store in descripcion
                        if not linea.descripcion:
                            linea.descripcion = f"Artículo UUID: {linea.articulo_id}"
                
                await conn.execute("""
                    INSERT INTO finanzas2.cont_oc_linea 
                    (empresa_id, oc_id, articulo_id, descripcion, cantidad, precio_unitario, igv_aplica, subtotal)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                """, empresa_id, oc_id, articulo_id_value, linea.descripcion, linea.cantidad, 
                    linea.precio_unitario, linea.igv_aplica, linea_subtotal)
            
            # Get full OC with joins within transaction
            oc_row = await conn.fetchrow("""
                SELECT oc.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
                FROM finanzas2.cont_oc oc
                LEFT JOIN finanzas2.cont_tercero t ON oc.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON oc.moneda_id = m.id
                WHERE oc.id = $1
            """, oc_id)
            
            oc_dict = dict(oc_row)
            lineas_rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1 ORDER BY id
            """, oc_id)
            oc_dict['lineas'] = [dict(l) for l in lineas_rows]
            
            return oc_dict

@api_router.put("/ordenes-compra/{id}", response_model=OC)
async def update_orden_compra(id: int, data: OCUpdate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        
        if not updates:
            raise HTTPException(400, "No fields to update")
        
        values.append(id)
        query = f"UPDATE finanzas2.cont_oc SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx} RETURNING id"
        row = await conn.fetchrow(query, *values)
        if not row:
            raise HTTPException(404, "Orden de compra not found")
        
        return await get_orden_compra(id)

@api_router.delete("/ordenes-compra/{id}")
async def delete_orden_compra(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if OC has been converted to factura
        oc = await conn.fetchrow("SELECT * FROM finanzas2.cont_oc WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not oc:
            raise HTTPException(404, "Orden de compra not found")
        if oc['factura_generada_id']:
            raise HTTPException(400, "Cannot delete OC that has generated a factura")
        
        await conn.execute("DELETE FROM finanzas2.cont_oc WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        return {"message": "Orden de compra deleted"}

@api_router.post("/ordenes-compra/{id}/generar-factura", response_model=FacturaProveedor)
async def generar_factura_desde_oc(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Generate a factura proveedor from an OC"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get OC
            oc = await conn.fetchrow("""
                SELECT * FROM finanzas2.cont_oc WHERE id = $1
            """, id)
            if not oc:
                raise HTTPException(404, "Orden de compra not found")
            if oc['factura_generada_id']:
                raise HTTPException(400, "Esta OC ya generó una factura")
            
            # Generate factura number
            year = datetime.now().year
            prefix = f"FP-{year}-"
            last = await conn.fetchval(f"""
                SELECT numero FROM finanzas2.cont_factura_proveedor 
                WHERE numero LIKE '{prefix}%' 
                ORDER BY id DESC LIMIT 1
            """)
            if last:
                num = int(last.split('-')[-1]) + 1
            else:
                num = 1
            numero = f"{prefix}{num:05d}"
            
            # Create factura
            factura = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_factura_proveedor 
                (empresa_id, numero, proveedor_id, moneda_id, fecha_factura, fecha_vencimiento, 
                 terminos_dias, tipo_documento, estado, subtotal, igv, total, saldo_pendiente, 
                 notas, oc_origen_id)
                VALUES ($1, $2, $3, $4, TO_DATE($5, 'YYYY-MM-DD'), TO_DATE($6, 'YYYY-MM-DD'), $7, 'factura', 'pendiente', $8, $9, $10, $10, $11, $12)
                RETURNING *
            """, empresa_id, numero, oc['proveedor_id'], oc['moneda_id'], safe_date_param(datetime.now().date()),
                safe_date_param(datetime.now().date() + timedelta(days=30)), 30,
                oc['subtotal'], oc['igv'], oc['total'], oc['notas'], id)
            
            factura_id = factura['id']
            
            # Copy lines
            oc_lineas = await conn.fetch("SELECT * FROM finanzas2.cont_oc_linea WHERE oc_id = $1", id)
            for linea in oc_lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_factura_proveedor_linea 
                    (empresa_id, factura_id, articulo_id, descripcion, importe, igv_aplica)
                    VALUES ($1, $2, $3, $4, $5, $6)
                """, empresa_id, factura_id, linea['articulo_id'], linea['descripcion'], 
                    linea['subtotal'], linea['igv_aplica'])
            
            # Create CxP
            await conn.execute("""
                INSERT INTO finanzas2.cont_cxp 
                (empresa_id, factura_id, proveedor_id, monto_original, saldo_pendiente, fecha_vencimiento, estado)
                VALUES ($1, $2, $3, $4, $4, $5, 'pendiente')
            """, factura_id, oc['proveedor_id'], oc['total'], 
                datetime.now().date() + timedelta(days=30), empresa_id)
            
            # Update OC
            await conn.execute("""
                UPDATE finanzas2.cont_oc SET estado = 'facturada', factura_generada_id = $1 WHERE id = $2 AND empresa_id = $3
            """, factura_id, id, empresa_id)
            
            # Get full factura within transaction
            f_row = await conn.fetchrow("""
                SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
                FROM finanzas2.cont_factura_proveedor fp
                LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
                WHERE fp.id = $1
            """, factura_id)
            
            factura_dict = dict(f_row)
            f_lineas = await conn.fetch("""
                SELECT fpl.*, c.nombre as categoria_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
                WHERE fpl.factura_id = $1 ORDER BY fpl.id
            """, factura_id)
            factura_dict['lineas'] = [dict(l) for l in f_lineas]
            
            return factura_dict

# =====================
# FACTURAS PROVEEDOR
# =====================
async def generate_factura_number(conn, empresa_id: int) -> str:
    year = datetime.now().year
    prefijo = f"FP-{year}-"
    return await get_next_correlativo(conn, empresa_id, 'factura_proveedor', prefijo)

@api_router.get("/facturas-proveedor", response_model=List[FacturaProveedor])
async def list_facturas_proveedor(
    estado: Optional[str] = None,
    proveedor_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["fp.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if estado:
            conditions.append(f"fp.estado = ${idx}")
            params.append(estado)
            idx += 1
        if proveedor_id:
            conditions.append(f"fp.proveedor_id = ${idx}")
            params.append(proveedor_id)
            idx += 1
        if fecha_desde:
            conditions.append(f"fp.fecha_factura >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"fp.fecha_factura <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
            FROM finanzas2.cont_factura_proveedor fp
            LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY fp.fecha_factura DESC, fp.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            fp_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT fpl.*, c.nombre as categoria_nombre, c.padre_id as categoria_padre_id,
                       cp.nombre as categoria_padre_nombre,
                       ln.nombre as linea_negocio_nombre, cc.nombre as centro_costo_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
                LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
                LEFT JOIN finanzas2.cont_linea_negocio ln ON fpl.linea_negocio_id = ln.id
                LEFT JOIN finanzas2.cont_centro_costo cc ON fpl.centro_costo_id = cc.id
                WHERE fpl.factura_id = $1 ORDER BY fpl.id
            """, row['id'])
            fp_dict['lineas'] = [dict(l) for l in lineas]
            result.append(fp_dict)
        
        return result

async def get_factura_proveedor(id: int, empresa_id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
            FROM finanzas2.cont_factura_proveedor fp
            LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE fp.id = $1 AND fp.empresa_id = $2
        """, id, empresa_id)
        
        if not row:
            raise HTTPException(404, "Factura not found")
        
        fp_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT fpl.*, c.nombre as categoria_nombre, c.padre_id as categoria_padre_id,
                   cp.nombre as categoria_padre_nombre,
                   ln.nombre as linea_negocio_nombre, cc.nombre as centro_costo_nombre
            FROM finanzas2.cont_factura_proveedor_linea fpl
            LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
            LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON fpl.linea_negocio_id = ln.id
            LEFT JOIN finanzas2.cont_centro_costo cc ON fpl.centro_costo_id = cc.id
            WHERE fpl.factura_id = $1 ORDER BY fpl.id
        """, id)
        fp_dict['lineas'] = [dict(l) for l in lineas]
        
        return fp_dict

@api_router.get("/facturas-proveedor/{id}", response_model=FacturaProveedor)
async def get_factura_proveedor_endpoint(id: int, empresa_id: int = Depends(get_empresa_id)):
    return await get_factura_proveedor(id, empresa_id)

@api_router.post("/facturas-proveedor", response_model=FacturaProveedor)
async def create_factura_proveedor(data: FacturaProveedorCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            numero = data.numero or await generate_factura_number(conn, empresa_id)
            
            # Calculate totals
            subtotal = sum(l.importe for l in data.lineas)
            if data.impuestos_incluidos:
                subtotal = subtotal / 1.18
                igv = subtotal * 0.18
            else:
                igv = sum(l.importe * 0.18 for l in data.lineas if l.igv_aplica)
            total = subtotal + igv
            
            # Calculate SUNAT tax breakdown from lines
            base_gravada = 0.0
            igv_sunat = 0.0
            base_no_gravada = 0.0
            for linea in data.lineas:
                imp = linea.importe
                if linea.igv_aplica:
                    if data.impuestos_incluidos:
                        base = imp / 1.18
                        base_gravada += base
                        igv_sunat += imp - base
                    else:
                        base_gravada += imp
                        igv_sunat += imp * 0.18
                else:
                    base_no_gravada += imp
            base_gravada = round(base_gravada, 2)
            igv_sunat = round(igv_sunat, 2)
            base_no_gravada = round(base_no_gravada, 2)
            isc_val = data.isc or 0.0
            
            # Calculate fecha_vencimiento
            fecha_vencimiento = data.fecha_vencimiento
            if not fecha_vencimiento and data.terminos_dias:
                fecha_vencimiento = data.fecha_factura + timedelta(days=data.terminos_dias)
            
            # Default fecha_contable = fecha_factura
            fecha_contable = data.fecha_contable or data.fecha_factura
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_factura_proveedor 
                (empresa_id, numero, proveedor_id, beneficiario_nombre, moneda_id, fecha_factura, fecha_contable, fecha_vencimiento,
                 terminos_dias, tipo_documento, estado, subtotal, igv, total, saldo_pendiente, 
                 impuestos_incluidos, tipo_comprobante_sunat, base_gravada, igv_sunat, base_no_gravada, isc, tipo_cambio, notas)
                VALUES ($1, $2, $3, $4, $5, TO_DATE($6, 'YYYY-MM-DD'), TO_DATE($7, 'YYYY-MM-DD'), TO_DATE($8, 'YYYY-MM-DD'), $9, $10, 'pendiente', $11, $12, $13, $13, $14, $15, $16, $17, $18, $19, $20, $21)
                RETURNING id
            """, empresa_id, numero, data.proveedor_id, data.beneficiario_nombre, data.moneda_id,
                safe_date_param(data.fecha_factura), safe_date_param(fecha_contable), safe_date_param(fecha_vencimiento), data.terminos_dias, data.tipo_documento,
                subtotal, igv, total, data.impuestos_incluidos, data.tipo_comprobante_sunat, base_gravada, igv_sunat, base_no_gravada, isc_val, data.tipo_cambio, data.notas)
            
            factura_id = row['id']
            
            # Insert lines
            for linea in data.lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_factura_proveedor_linea 
                    (empresa_id, factura_id, categoria_id, articulo_id, descripcion, linea_negocio_id, 
                     centro_costo_id, importe, igv_aplica)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                """, empresa_id, factura_id, linea.categoria_id, linea.articulo_id, linea.descripcion,
                    linea.linea_negocio_id, linea.centro_costo_id, linea.importe, linea.igv_aplica)
            
            # Create CxP
            await conn.execute("""
                INSERT INTO finanzas2.cont_cxp 
                (empresa_id, factura_id, proveedor_id, monto_original, saldo_pendiente, fecha_vencimiento, estado)
                VALUES ($1, $2, $3, $4, $4, $5, 'pendiente')
            """, empresa_id, factura_id, data.proveedor_id, total, fecha_vencimiento)
            
            # Get the created factura with relations within the same transaction
            row = await conn.fetchrow("""
                SELECT fp.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
                FROM finanzas2.cont_factura_proveedor fp
                LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
                WHERE fp.id = $1
            """, factura_id)
            
            if not row:
                raise HTTPException(404, "Factura not found after creation")
            
            fp_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT fpl.*, c.nombre as categoria_nombre, c.padre_id as categoria_padre_id,
                       cp.nombre as categoria_padre_nombre,
                       ln.nombre as linea_negocio_nombre, cc.nombre as centro_costo_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria c ON fpl.categoria_id = c.id
                LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
                LEFT JOIN finanzas2.cont_linea_negocio ln ON fpl.linea_negocio_id = ln.id
                LEFT JOIN finanzas2.cont_centro_costo cc ON fpl.centro_costo_id = cc.id
                WHERE fpl.factura_id = $1 ORDER BY fpl.id
            """, factura_id)
            fp_dict['lineas'] = [dict(l) for l in lineas]
            
            return fp_dict

@api_router.put("/facturas-proveedor/{id}", response_model=FacturaProveedor)
async def update_factura_proveedor(id: int, data: FacturaProveedorUpdate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if factura can be edited
        factura = await conn.fetchrow("SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not factura:
            raise HTTPException(404, "Factura not found")
        if factura['estado'] in ('pagado', 'anulada'):
            raise HTTPException(400, "Cannot edit paid or cancelled factura")
        
        updates = []
        values = []
        idx = 1
        for field, value in data.model_dump(exclude_unset=True).items():
            updates.append(f"{field} = ${idx}")
            values.append(value)
            idx += 1
        
        if not updates:
            raise HTTPException(400, "No fields to update")
        
        values.append(id)
        query = f"UPDATE finanzas2.cont_factura_proveedor SET {', '.join(updates)}, updated_at = NOW() WHERE id = ${idx}"
        await conn.execute(query, *values)
        
        return await get_factura_proveedor(id, empresa_id)

@api_router.delete("/facturas-proveedor/{id}")
async def delete_factura_proveedor(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            factura = await conn.fetchrow("SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not factura:
                raise HTTPException(404, "Factura not found")
            
            # Check if has payments or letras
            pagos = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_pago_aplicacion 
                WHERE tipo_documento = 'factura' AND documento_id = $1
            """, id)
            if pagos > 0:
                raise HTTPException(400, "Cannot delete factura with payments. Reverse payments first.")
            
            letras = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_letra WHERE factura_id = $1
            """, id)
            if letras > 0:
                raise HTTPException(400, "Cannot delete factura with letras. Delete letras first.")
            
            # Delete CxP and factura
            await conn.execute("DELETE FROM finanzas2.cont_cxp WHERE factura_id = $1", id)
            await conn.execute("DELETE FROM finanzas2.cont_factura_proveedor WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            
            return {"message": "Factura deleted"}

@api_router.get("/facturas-proveedor/{id}/pagos")
async def get_pagos_de_factura(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Get all payments applied to a specific factura"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Get pagos through pago_aplicacion
        rows = await conn.fetch("""
            SELECT p.*, pa.monto_aplicado, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo, m.simbolo as moneda_simbolo
            FROM finanzas2.cont_pago p
            JOIN finanzas2.cont_pago_aplicacion pa ON pa.pago_id = p.id
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
            WHERE pa.tipo_documento = 'factura' AND pa.documento_id = $1
            ORDER BY p.fecha DESC
        """, id)
        
        return [dict(r) for r in rows]

@api_router.get("/facturas-proveedor/{id}/letras")
async def get_letras_de_factura(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Get all letras generated from a specific factura"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        rows = await conn.fetch("""
            SELECT l.*, 
                   fp.moneda_id,
                   m.codigo as moneda_codigo, 
                   m.simbolo as moneda_simbolo
            FROM finanzas2.cont_letra l
            LEFT JOIN finanzas2.cont_factura_proveedor fp ON l.factura_id = fp.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE l.factura_id = $1
            ORDER BY l.fecha_vencimiento ASC
        """, id)
        
        return [dict(r) for r in rows]

@api_router.post("/facturas-proveedor/{id}/deshacer-canje")
async def deshacer_canje_letras(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Reverse the canje of letras - delete all letras and set factura back to pendiente"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            factura = await conn.fetchrow("SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not factura:
                raise HTTPException(404, "Factura not found")
            
            if factura['estado'] != 'canjeado':
                raise HTTPException(400, "Factura is not in canjeado state")
            
            # Check if any letra has payments
            pagos_letras = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_pago_aplicacion pa
                JOIN finanzas2.cont_letra l ON pa.tipo_documento = 'letra' AND pa.documento_id = l.id
                WHERE l.factura_id = $1
            """, id)
            
            if pagos_letras > 0:
                raise HTTPException(400, "Cannot undo canje - some letras have payments. Delete payments first.")
            
            # Delete all letras
            await conn.execute("DELETE FROM finanzas2.cont_letra WHERE factura_id = $1", id)
            
            # Update factura back to pendiente
            await conn.execute("""
                UPDATE finanzas2.cont_factura_proveedor 
                SET estado = 'pendiente', updated_at = NOW() 
                WHERE id = $1
            """, id)
            
            return {"message": "Canje reversed successfully"}

# =====================
# PAGOS
# =====================
async def generate_pago_number(conn, tipo: str, empresa_id: int) -> str:
    year = datetime.now().year
    prefijo = f"PAG-{tipo[0].upper()}-{year}-"
    return await get_next_correlativo(conn, empresa_id, f'pago_{tipo}', prefijo)

@api_router.get("/pagos", response_model=List[Pago])
async def list_pagos(
    tipo: Optional[str] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    cuenta_financiera_id: Optional[int] = None,
    conciliado: Optional[bool] = None,
    centro_costo_id: Optional[int] = None,
    linea_negocio_id: Optional[int] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["p.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if tipo:
            conditions.append(f"p.tipo = ${idx}")
            params.append(tipo)
            idx += 1
        if fecha_desde:
            conditions.append(f"p.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"p.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        if cuenta_financiera_id:
            conditions.append(f"p.cuenta_financiera_id = ${idx}")
            params.append(cuenta_financiera_id)
            idx += 1
        if conciliado is not None:
            conditions.append(f"COALESCE(p.conciliado, false) = ${idx}")
            params.append(conciliado)
            idx += 1
        if centro_costo_id:
            conditions.append(f"p.centro_costo_id = ${idx}")
            params.append(centro_costo_id)
            idx += 1
        if linea_negocio_id:
            conditions.append(f"p.linea_negocio_id = ${idx}")
            params.append(linea_negocio_id)
            idx += 1
        
        query = f"""
            SELECT p.*, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo,
                   cc.nombre as centro_costo_nombre, ln.nombre as linea_negocio_nombre
            FROM finanzas2.cont_pago p
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
            LEFT JOIN finanzas2.cont_centro_costo cc ON p.centro_costo_id = cc.id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON p.linea_negocio_id = ln.id
            WHERE {' AND '.join(conditions)}
            ORDER BY p.fecha DESC, p.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            pago_dict = dict(row)
            
            # Get detalles
            detalles = await conn.fetch("""
                SELECT pd.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_pago_detalle pd
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON pd.cuenta_financiera_id = cf.id
                WHERE pd.pago_id = $1
            """, row['id'])
            pago_dict['detalles'] = [dict(d) for d in detalles]
            
            # Get aplicaciones
            aplicaciones = await conn.fetch("""
                SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1
            """, row['id'])
            pago_dict['aplicaciones'] = [dict(a) for a in aplicaciones]
            
            result.append(pago_dict)
        
        return result

@api_router.post("/pagos", response_model=Pago)
async def create_pago(data: PagoCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Validate that payment amount doesn't exceed document balance
            for aplicacion in data.aplicaciones:
                if aplicacion.tipo_documento == 'factura':
                    doc = await conn.fetchrow("""
                        SELECT saldo_pendiente, total, estado FROM finanzas2.cont_factura_proveedor WHERE id = $1
                    """, aplicacion.documento_id)
                    if not doc:
                        raise HTTPException(404, f"Factura {aplicacion.documento_id} not found")
                    if doc['estado'] == 'canjeado':
                        raise HTTPException(400, "No se puede pagar una factura canjeada. Debe pagar las letras.")
                    if aplicacion.monto_aplicado > float(doc['saldo_pendiente']):
                        raise HTTPException(400, f"El monto ({aplicacion.monto_aplicado:.2f}) excede el saldo pendiente ({doc['saldo_pendiente']:.2f})")
                elif aplicacion.tipo_documento == 'letra':
                    doc = await conn.fetchrow("""
                        SELECT saldo_pendiente, monto, estado FROM finanzas2.cont_letra WHERE id = $1
                    """, aplicacion.documento_id)
                    if not doc:
                        raise HTTPException(404, f"Letra {aplicacion.documento_id} not found")
                    if aplicacion.monto_aplicado > float(doc['saldo_pendiente']):
                        raise HTTPException(400, f"El monto ({aplicacion.monto_aplicado:.2f}) excede el saldo pendiente ({doc['saldo_pendiente']:.2f})")
            
            numero = await generate_pago_number(conn, data.tipo, empresa_id)
            
            # Create pago
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (empresa_id, numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, referencia, notas)
                VALUES ($1, $2, $3, TO_DATE($4, 'YYYY-MM-DD'), $5, $6, $7, $8, $9)
                RETURNING *
            """, empresa_id, numero, data.tipo, safe_date_param(data.fecha), data.cuenta_financiera_id, data.moneda_id,
                data.monto_total, data.referencia, data.notas)
            
            pago_id = pago['id']
            
            # Insert detalles (multi-medio)
            for detalle in data.detalles:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_detalle 
                    (empresa_id, pago_id, cuenta_financiera_id, medio_pago, monto, referencia)
                    VALUES ($1, $2, $3, $4, $5, $6)
                """, empresa_id, pago_id, detalle.cuenta_financiera_id, detalle.medio_pago, 
                    detalle.monto, detalle.referencia)
                
                # Update cuenta financiera saldo
                if data.tipo == 'egreso':
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual - $1 WHERE id = $2
                    """, detalle.monto, detalle.cuenta_financiera_id)
                else:
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual + $1 WHERE id = $2
                    """, detalle.monto, detalle.cuenta_financiera_id)
            
            # Insert aplicaciones and update documents
            for aplicacion in data.aplicaciones:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_aplicacion 
                    (empresa_id, pago_id, tipo_documento, documento_id, monto_aplicado)
                    VALUES ($1, $2, $3, $4, $5)
                """, empresa_id, pago_id, aplicacion.tipo_documento, aplicacion.documento_id, 
                    aplicacion.monto_aplicado)
                
                # Update document saldo
                if aplicacion.tipo_documento == 'factura':
                    await conn.execute("""
                        UPDATE finanzas2.cont_factura_proveedor 
                        SET saldo_pendiente = saldo_pendiente - $1 WHERE id = $2
                    """, aplicacion.monto_aplicado, aplicacion.documento_id)
                    
                    # Update estado
                    fp = await conn.fetchrow("""
                        SELECT total, saldo_pendiente FROM finanzas2.cont_factura_proveedor WHERE id = $1
                    """, aplicacion.documento_id)
                    if fp['saldo_pendiente'] <= 0:
                        await conn.execute("""
                            UPDATE finanzas2.cont_factura_proveedor SET estado = 'pagado' WHERE id = $1
                        """, aplicacion.documento_id)
                        await conn.execute("""
                            UPDATE finanzas2.cont_cxp SET estado = 'pagado', saldo_pendiente = 0 WHERE factura_id = $1
                        """, aplicacion.documento_id)
                    else:
                        await conn.execute("""
                            UPDATE finanzas2.cont_factura_proveedor SET estado = 'parcial' WHERE id = $1
                        """, aplicacion.documento_id)
                        await conn.execute("""
                            UPDATE finanzas2.cont_cxp SET estado = 'parcial', saldo_pendiente = $2 WHERE factura_id = $1
                        """, aplicacion.documento_id, fp['saldo_pendiente'])
                
                elif aplicacion.tipo_documento == 'letra':
                    await conn.execute("""
                        UPDATE finanzas2.cont_letra 
                        SET saldo_pendiente = saldo_pendiente - $1 WHERE id = $2
                    """, aplicacion.monto_aplicado, aplicacion.documento_id)
                    
                    letra = await conn.fetchrow("""
                        SELECT monto, saldo_pendiente, factura_id FROM finanzas2.cont_letra WHERE id = $1
                    """, aplicacion.documento_id)
                    if letra['saldo_pendiente'] <= 0:
                        await conn.execute("""
                            UPDATE finanzas2.cont_letra SET estado = 'pagada' WHERE id = $1
                        """, aplicacion.documento_id)
                    else:
                        await conn.execute("""
                            UPDATE finanzas2.cont_letra SET estado = 'parcial' WHERE id = $1
                        """, aplicacion.documento_id)
                    
                    # Update parent factura CxP saldo based on remaining letras
                    if letra['factura_id']:
                        total_letras_pendiente = await conn.fetchval("""
                            SELECT COALESCE(SUM(saldo_pendiente), 0) 
                            FROM finanzas2.cont_letra WHERE factura_id = $1
                        """, letra['factura_id'])
                        nuevo_saldo = float(total_letras_pendiente)
                        nuevo_estado = 'pagado' if nuevo_saldo <= 0 else 'parcial'
                        await conn.execute("""
                            UPDATE finanzas2.cont_cxp 
                            SET saldo_pendiente = $2, estado = $3
                            WHERE factura_id = $1
                        """, letra['factura_id'], nuevo_saldo, nuevo_estado)
                        # Also update factura saldo_pendiente for consistent display
                        await conn.execute("""
                            UPDATE finanzas2.cont_factura_proveedor 
                            SET saldo_pendiente = $2
                            WHERE id = $1
                        """, letra['factura_id'], nuevo_saldo)
            
            # Get full pago with relations within the same transaction
            row = await conn.fetchrow("""
                SELECT p.*, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo
                FROM finanzas2.cont_pago p
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
                LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
                WHERE p.id = $1
            """, pago_id)
            
            if not row:
                raise HTTPException(404, "Pago not found after creation")
            
            pago_dict = dict(row)
            
            detalles = await conn.fetch("""
                SELECT pd.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_pago_detalle pd
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON pd.cuenta_financiera_id = cf.id
                WHERE pd.pago_id = $1
            """, pago_id)
            pago_dict['detalles'] = [dict(d) for d in detalles]
            
            aplicaciones = await conn.fetch("""
                SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1
            """, pago_id)
            pago_dict['aplicaciones'] = [dict(a) for a in aplicaciones]
            
            return pago_dict

async def get_pago(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT p.*, cf.nombre as cuenta_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_pago p
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            LEFT JOIN finanzas2.cont_moneda m ON p.moneda_id = m.id
            WHERE p.id = $1
        """, id)
        
        if not row:
            raise HTTPException(404, "Pago not found")
        
        pago_dict = dict(row)
        
        detalles = await conn.fetch("""
            SELECT pd.*, cf.nombre as cuenta_nombre
            FROM finanzas2.cont_pago_detalle pd
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON pd.cuenta_financiera_id = cf.id
            WHERE pd.pago_id = $1
        """, id)
        pago_dict['detalles'] = [dict(d) for d in detalles]
        
        aplicaciones = await conn.fetch("""
            SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1
        """, id)
        pago_dict['aplicaciones'] = [dict(a) for a in aplicaciones]
        
        return pago_dict

@api_router.get("/pagos/{id}", response_model=Pago)
async def get_pago_endpoint(id: int, empresa_id: int = Depends(get_empresa_id)):
    return await get_pago(id)


@api_router.put("/pagos/{id}")
async def update_pago(id: int, data: dict, empresa_id: int = Depends(get_empresa_id)):
    """Update pago - only allow editing referencia if conciliado"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if pago exists and if it's conciliado
        pago = await conn.fetchrow("SELECT * FROM finanzas2.cont_pago WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not pago:
            raise HTTPException(404, "Pago no encontrado")
        
        # If conciliado, only allow updating referencia
        if pago['conciliado']:
            if 'referencia' in data:
                await conn.execute("""
                    UPDATE finanzas2.cont_pago 
                    SET referencia = $1, updated_at = NOW()
                    WHERE id = $2
                """, data.get('referencia'), id)
            return {"message": "Referencia actualizada (pago conciliado)"}
        
        # If not conciliado, allow updating multiple fields
        update_fields = []
        values = []
        param_count = 1
        
        if 'fecha' in data:
            update_fields.append(f"fecha = TO_DATE(${param_count}, 'YYYY-MM-DD')")
            values.append(safe_date_param(data['fecha']))
            param_count += 1
        
        if 'referencia' in data:
            update_fields.append(f"referencia = ${param_count}")
            values.append(data['referencia'])
            param_count += 1
        
        if 'notas' in data:
            update_fields.append(f"notas = ${param_count}")
            values.append(data['notas'])
            param_count += 1
        
        update_fields.append(f"updated_at = NOW()")
        
        if update_fields:
            values.append(id)
            query = f"""
                UPDATE finanzas2.cont_pago 
                SET {', '.join(update_fields)}
                WHERE id = ${param_count}
            """
            await conn.execute(query, *values)
        
        return {"message": "Pago actualizado exitosamente"}


@api_router.delete("/pagos/{id}")
async def delete_pago(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Delete a payment and reverse all its effects"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            pago = await conn.fetchrow("SELECT * FROM finanzas2.cont_pago WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not pago:
                raise HTTPException(404, "Pago not found")
            
            # Reverse cuenta financiera updates
            detalles = await conn.fetch("SELECT * FROM finanzas2.cont_pago_detalle WHERE pago_id = $1", id)
            for detalle in detalles:
                if pago['tipo'] == 'egreso':
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual + $1 WHERE id = $2
                    """, detalle['monto'], detalle['cuenta_financiera_id'])
                else:
                    await conn.execute("""
                        UPDATE finanzas2.cont_cuenta_financiera 
                        SET saldo_actual = saldo_actual - $1 WHERE id = $2
                    """, detalle['monto'], detalle['cuenta_financiera_id'])
            
            # Reverse aplicaciones
            aplicaciones = await conn.fetch("SELECT * FROM finanzas2.cont_pago_aplicacion WHERE pago_id = $1", id)
            for aplicacion in aplicaciones:
                if aplicacion['tipo_documento'] == 'factura':
                    await conn.execute("""
                        UPDATE finanzas2.cont_factura_proveedor 
                        SET saldo_pendiente = saldo_pendiente + $1, estado = 'pendiente' WHERE id = $2
                    """, aplicacion['monto_aplicado'], aplicacion['documento_id'])
                    await conn.execute("""
                        UPDATE finanzas2.cont_cxp 
                        SET saldo_pendiente = saldo_pendiente + $1, estado = 'pendiente' WHERE factura_id = $2
                    """, aplicacion['monto_aplicado'], aplicacion['documento_id'])
                elif aplicacion['tipo_documento'] == 'letra':
                    await conn.execute("""
                        UPDATE finanzas2.cont_letra 
                        SET saldo_pendiente = saldo_pendiente + $1, estado = 'pendiente' WHERE id = $2
                    """, aplicacion['monto_aplicado'], aplicacion['documento_id'])
            
            # Delete pago
            await conn.execute("DELETE FROM finanzas2.cont_pago WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            
            return {"message": "Pago deleted and reversed"}

# =====================
# LETRAS
# =====================
@api_router.get("/letras", response_model=List[Letra])
async def list_letras(
    estado: Optional[str] = None,
    proveedor_id: Optional[int] = None,
    factura_id: Optional[int] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["l.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if estado:
            conditions.append(f"l.estado = ${idx}")
            params.append(estado)
            idx += 1
        if proveedor_id:
            conditions.append(f"l.proveedor_id = ${idx}")
            params.append(proveedor_id)
            idx += 1
        if factura_id:
            conditions.append(f"l.factura_id = ${idx}")
            params.append(factura_id)
            idx += 1
        
        query = f"""
            SELECT l.*, t.nombre as proveedor_nombre, fp.numero as factura_numero
            FROM finanzas2.cont_letra l
            LEFT JOIN finanzas2.cont_tercero t ON l.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_factura_proveedor fp ON l.factura_id = fp.id
            WHERE {' AND '.join(conditions)}
            ORDER BY l.fecha_vencimiento ASC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.post("/letras/generar", response_model=List[Letra])
async def generar_letras(data: GenerarLetrasRequest, empresa_id: int = Depends(get_empresa_id)):
    """Generate letras from a factura proveedor"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get factura
            factura = await conn.fetchrow("""
                SELECT * FROM finanzas2.cont_factura_proveedor WHERE id = $1
            """, data.factura_id)
            if not factura:
                raise HTTPException(404, "Factura not found")
            if factura['estado'] in ('pagado', 'anulada', 'canjeado'):
                raise HTTPException(400, "Cannot generate letras for this factura")
            
            # Check if already has letras
            existing = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_letra WHERE factura_id = $1
            """, data.factura_id)
            if existing > 0:
                raise HTTPException(400, "Factura already has letras")
            
            letras = []
            
            # Check if using custom letras
            if data.letras_personalizadas and len(data.letras_personalizadas) > 0:
                # Validate total matches factura total
                total_letras = sum(l.monto for l in data.letras_personalizadas)
                if abs(total_letras - float(factura['total'])) > 0.01:
                    raise HTTPException(400, f"El total de las letras ({total_letras:.2f}) debe ser igual al total de la factura ({factura['total']:.2f})")
                
                for i, letra_data in enumerate(data.letras_personalizadas):
                    numero = f"L-{factura['numero']}-{i+1:02d}"
                    
                    letra = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_letra 
                        (empresa_id, numero, factura_id, proveedor_id, monto, fecha_emision, fecha_vencimiento, 
                         estado, saldo_pendiente)
                        VALUES ($1, $2, $3, $4, $5, TO_DATE($6, 'YYYY-MM-DD'), TO_DATE($7, 'YYYY-MM-DD'), 'pendiente', $5)
                        RETURNING *
                    """, empresa_id, numero, data.factura_id, factura['proveedor_id'], letra_data.monto,
                        safe_date_param(datetime.now().date()), safe_date_param(letra_data.fecha_vencimiento))
                    letras.append(dict(letra))
            else:
                # Use automatic calculation
                monto_por_letra = data.monto_por_letra or (factura['total'] / data.cantidad_letras)
                fecha_base = factura['fecha_vencimiento'] or datetime.now().date()
                
                for i in range(data.cantidad_letras):
                    fecha_vencimiento = fecha_base + timedelta(days=data.dias_entre_letras * i)
                    numero = f"L-{factura['numero']}-{i+1:02d}"
                    
                    letra = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_letra 
                        (empresa_id, numero, factura_id, proveedor_id, monto, fecha_emision, fecha_vencimiento, 
                         estado, saldo_pendiente)
                        VALUES ($1, $2, $3, $4, $5, TO_DATE($6, 'YYYY-MM-DD'), TO_DATE($7, 'YYYY-MM-DD'), 'pendiente', $5)
                        RETURNING *
                    """, empresa_id, numero, data.factura_id, factura['proveedor_id'], monto_por_letra,
                        safe_date_param(datetime.now().date()), safe_date_param(fecha_vencimiento))
                    letras.append(dict(letra))
            
            # Update factura status to 'canjeado'
            await conn.execute("""
                UPDATE finanzas2.cont_factura_proveedor SET estado = 'canjeado' WHERE id = $1
            """, data.factura_id)
            await conn.execute("""
                UPDATE finanzas2.cont_cxp SET estado = 'canjeado' WHERE factura_id = $1
            """, data.factura_id)
            
            return letras

@api_router.delete("/letras/{id}")
async def delete_letra(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            letra = await conn.fetchrow("SELECT * FROM finanzas2.cont_letra WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not letra:
                raise HTTPException(404, "Letra not found")
            
            # Check if has payments
            pagos = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_pago_aplicacion 
                WHERE tipo_documento = 'letra' AND documento_id = $1
            """, id)
            if pagos > 0:
                raise HTTPException(400, "Cannot delete letra with payments. Reverse payments first.")
            
            factura_id = letra['factura_id']
            
            await conn.execute("DELETE FROM finanzas2.cont_letra WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            
            # Check if factura has remaining letras
            remaining = await conn.fetchval("""
                SELECT COUNT(*) FROM finanzas2.cont_letra WHERE factura_id = $1
            """, factura_id)
            
            if remaining == 0:
                # Revert factura to pendiente
                await conn.execute("""
                    UPDATE finanzas2.cont_factura_proveedor SET estado = 'pendiente' WHERE id = $1
                """, factura_id)
                await conn.execute("""
                    UPDATE finanzas2.cont_cxp SET estado = 'pendiente' WHERE factura_id = $1
                """, factura_id)
            
            return {"message": "Letra deleted"}

# =====================
# GASTOS
# =====================
async def generate_gasto_number(conn, empresa_id: int) -> str:
    year = datetime.now().year
    prefijo = f"GAS-{year}-"
    return await get_next_correlativo(conn, empresa_id, 'gasto', prefijo)

@api_router.get("/gastos", response_model=List[Gasto])
async def list_gastos(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["g.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if fecha_desde:
            conditions.append(f"g.fecha >= ${idx}")
            params.append(fecha_desde)
            idx += 1
        if fecha_hasta:
            conditions.append(f"g.fecha <= ${idx}")
            params.append(fecha_hasta)
            idx += 1
        
        query = f"""
            SELECT g.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_gasto g
            LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY g.fecha DESC, g.id DESC
        """
        rows = await conn.fetch(query, *params)
        
        result = []
        for row in rows:
            gasto_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT gl.*, c.nombre as categoria_nombre, c.padre_id as categoria_padre_id,
                       cp.nombre as categoria_padre_nombre
                FROM finanzas2.cont_gasto_linea gl
                LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
                LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
                WHERE gl.gasto_id = $1 ORDER BY gl.id
            """, row['id'])
            gasto_dict['lineas'] = [dict(l) for l in lineas]
            result.append(gasto_dict)
        
        return result

@api_router.post("/gastos", response_model=Gasto)
async def create_gasto(data: GastoCreate, empresa_id: int = Depends(get_empresa_id)):
    """Create a gasto with mandatory payment(s) that must cover total"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            numero = await generate_gasto_number(conn, empresa_id)
            
            # Calculate totals
            subtotal = sum(l.importe for l in data.lineas)
            igv = sum(l.importe * 0.18 for l in data.lineas if l.igv_aplica)
            total = subtotal + igv
            
            # Calculate SUNAT tax breakdown from lines
            base_gravada = 0.0
            igv_sunat = 0.0
            base_no_gravada = 0.0
            for linea in data.lineas:
                imp = linea.importe
                if linea.igv_aplica:
                    base_gravada += imp
                    igv_sunat += imp * 0.18
                else:
                    base_no_gravada += imp
            base_gravada = round(base_gravada, 2)
            igv_sunat = round(igv_sunat, 2)
            base_no_gravada = round(base_no_gravada, 2)
            isc_val = data.isc or 0.0
            
            # Validate payments sum to total
            if not data.pagos:
                raise HTTPException(400, "El gasto debe tener al menos un pago")
            
            total_pagos = sum(p.monto for p in data.pagos)
            if abs(total_pagos - total) > 0.01:
                raise HTTPException(400, f"El total de pagos ({total_pagos:.2f}) debe ser igual al total del gasto ({total:.2f})")
            
            # Create gasto first
            fecha_contable = data.fecha_contable or data.fecha
            gasto = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_gasto 
                (empresa_id, numero, fecha, fecha_contable, proveedor_id, beneficiario_nombre, moneda_id, subtotal, igv, total,
                 tipo_documento, numero_documento, tipo_comprobante_sunat, base_gravada, igv_sunat, base_no_gravada, isc, tipo_cambio, notas)
                VALUES ($1, $2, TO_DATE($3, 'YYYY-MM-DD'), TO_DATE($4, 'YYYY-MM-DD'), $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19)
                RETURNING id
            """, empresa_id, numero, safe_date_param(data.fecha), safe_date_param(fecha_contable), data.proveedor_id, data.beneficiario_nombre, data.moneda_id,
                subtotal, igv, total, data.tipo_documento, data.numero_documento, data.tipo_comprobante_sunat, base_gravada, igv_sunat, base_no_gravada, isc_val, data.tipo_cambio, data.notas)
            
            gasto_id = gasto['id']
            
            # Insert lineas
            for linea in data.lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_gasto_linea 
                    (empresa_id, gasto_id, categoria_id, descripcion, linea_negocio_id, centro_costo_id, importe, igv_aplica)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                """, empresa_id, gasto_id, linea.categoria_id, linea.descripcion, linea.linea_negocio_id,
                    linea.centro_costo_id, linea.importe, linea.igv_aplica)
            
            # Create pago(s)
            pago_numero = await generate_pago_number(conn, 'egreso', empresa_id)
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (empresa_id, numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, notas)
                VALUES ($1, $2, 'egreso', TO_DATE($3, 'YYYY-MM-DD'), $4, $5, $6, $7)
                RETURNING id
            """, empresa_id, pago_numero, safe_date_param(data.fecha), data.pagos[0].cuenta_financiera_id, data.moneda_id,
                total, f"Pago de gasto {numero}")
            
            pago_id = pago['id']
            
            # Insert pago detalles (multiple methods)
            for pago_det in data.pagos:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_detalle 
                    (empresa_id, pago_id, cuenta_financiera_id, medio_pago, monto, referencia)
                    VALUES ($1, $2, $3, $4, $5, $6)
                """, empresa_id, pago_id, pago_det.cuenta_financiera_id, pago_det.medio_pago, 
                    pago_det.monto, pago_det.referencia)
                
                # Update cuenta financiera
                await conn.execute("""
                    UPDATE finanzas2.cont_cuenta_financiera 
                    SET saldo_actual = saldo_actual - $1 WHERE id = $2
                """, pago_det.monto, pago_det.cuenta_financiera_id)
            
            # Insert pago aplicacion
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                VALUES ($1, 'gasto', $2, $3, $4)
            """, pago_id, gasto_id, total, empresa_id)
            
            # Update gasto with pago_id
            await conn.execute("""
                UPDATE finanzas2.cont_gasto SET pago_id = $1 WHERE id = $2
            """, pago_id, gasto_id)
            
            # Get full gasto data within transaction
            row = await conn.fetchrow("""
                SELECT g.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
                FROM finanzas2.cont_gasto g
                LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
                LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
                WHERE g.id = $1
            """, gasto_id)
            
            gasto_dict = dict(row)
            lineas_rows = await conn.fetch("""
                SELECT gl.*, c.nombre as categoria_nombre, c.padre_id as categoria_padre_id,
                       cp.nombre as categoria_padre_nombre
                FROM finanzas2.cont_gasto_linea gl
                LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
                LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
                WHERE gl.gasto_id = $1 ORDER BY gl.id
            """, gasto_id)
            gasto_dict['lineas'] = [dict(l) for l in lineas_rows]
            
            return gasto_dict

async def get_gasto(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            SELECT g.*, t.nombre as proveedor_nombre, m.codigo as moneda_codigo
            FROM finanzas2.cont_gasto g
            LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
            WHERE g.id = $1
        """, id)
        
        if not row:
            raise HTTPException(404, "Gasto not found")
        
        gasto_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT gl.*, c.nombre as categoria_nombre, c.padre_id as categoria_padre_id,
                   cp.nombre as categoria_padre_nombre
            FROM finanzas2.cont_gasto_linea gl
            LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
            LEFT JOIN finanzas2.cont_categoria cp ON c.padre_id = cp.id
            WHERE gl.gasto_id = $1 ORDER BY gl.id
        """, id)
        gasto_dict['lineas'] = [dict(l) for l in lineas]
        
        return gasto_dict

@api_router.get("/gastos/{id}", response_model=Gasto)
async def get_gasto_endpoint(id: int, empresa_id: int = Depends(get_empresa_id)):
    return await get_gasto(id)

@api_router.delete("/gastos/{id}")
async def delete_gasto(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Delete a gasto and its associated lines and payments"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if gasto exists
        gasto = await conn.fetchrow("SELECT * FROM finanzas2.cont_gasto WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not gasto:
            raise HTTPException(status_code=404, detail="Gasto no encontrado")
        
        # Delete associated lines
        await conn.execute("DELETE FROM finanzas2.cont_gasto_linea WHERE gasto_id = $1", id)
        
        # Delete the gasto (pagos should cascade or be handled separately)
        await conn.execute("DELETE FROM finanzas2.cont_gasto WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        
        return {"message": "Gasto eliminado exitosamente"}

# =====================
# ADELANTOS
# =====================
@api_router.get("/adelantos", response_model=List[Adelanto])
async def list_adelantos(
    empleado_id: Optional[int] = None,
    pagado: Optional[bool] = None,
    descontado: Optional[bool] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["a.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if empleado_id:
            conditions.append(f"a.empleado_id = ${idx}")
            params.append(empleado_id)
            idx += 1
        if pagado is not None:
            conditions.append(f"a.pagado = ${idx}")
            params.append(pagado)
            idx += 1
        if descontado is not None:
            conditions.append(f"a.descontado = ${idx}")
            params.append(descontado)
            idx += 1
        
        query = f"""
            SELECT a.*, t.nombre as empleado_nombre
            FROM finanzas2.cont_adelanto_empleado a
            LEFT JOIN finanzas2.cont_tercero t ON a.empleado_id = t.id
            WHERE {' AND '.join(conditions)}
            ORDER BY a.fecha DESC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.post("/adelantos", response_model=Adelanto)
async def create_adelanto(data: AdelantoCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            pago_id = None
            
            if data.pagar and data.cuenta_financiera_id:
                # Get centro_costo and linea_negocio from employee
                emp_info = await conn.fetchrow("""
                    SELECT centro_costo_id, linea_negocio_id
                    FROM finanzas2.cont_empleado_detalle WHERE tercero_id = $1
                """, data.empleado_id)
                cc_id = emp_info['centro_costo_id'] if emp_info else None
                ln_id = emp_info['linea_negocio_id'] if emp_info else None
                
                # Create pago
                pago_numero = await generate_pago_number(conn, 'egreso', empresa_id)
                pago = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_pago 
                    (numero, tipo, fecha, cuenta_financiera_id, monto_total, notas, centro_costo_id, linea_negocio_id, empresa_id)
                    VALUES ($1, 'egreso', TO_DATE($2, 'YYYY-MM-DD'), $3, $4, $5, $6, $7, $8)
                    RETURNING id
                """, pago_numero, safe_date_param(data.fecha), data.cuenta_financiera_id, data.monto, 
                    "Adelanto a empleado", cc_id, ln_id, empresa_id)
                pago_id = pago['id']
                
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_detalle 
                    (pago_id, cuenta_financiera_id, medio_pago, monto, empresa_id)
                    VALUES ($1, $2, $3, $4, $5)
                """, pago_id, data.cuenta_financiera_id, data.medio_pago, data.monto, empresa_id)
                
                await conn.execute("""
                    UPDATE finanzas2.cont_cuenta_financiera 
                    SET saldo_actual = saldo_actual - $1 WHERE id = $2
                """, data.monto, data.cuenta_financiera_id)
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_adelanto_empleado 
                (empleado_id, fecha, monto, motivo, pagado, pago_id, empresa_id)
                VALUES ($1, TO_DATE($2, 'YYYY-MM-DD'), $3, $4, $5, $6, $7)
                RETURNING *
            """, data.empleado_id, safe_date_param(data.fecha), data.monto, data.motivo, 
                data.pagar, pago_id, empresa_id)
            
            if pago_id:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_pago_aplicacion 
                    (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                    VALUES ($1, 'adelanto', $2, $3, $4)
                """, pago_id, row['id'], data.monto, empresa_id)
            
            # Get empleado nombre
            emp = await conn.fetchrow("SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1 AND empresa_id = $2", data.empleado_id, empresa_id)
            result = dict(row)
            result['empleado_nombre'] = emp['nombre'] if emp else None
            
            return result

@api_router.post("/adelantos/{id}/pagar", response_model=Adelanto)
async def pagar_adelanto(
    id: int, 
    cuenta_financiera_id: int = Query(...),
    medio_pago: str = Query(default="efectivo"),
    empresa_id: int = Depends(get_empresa_id),
):
    """Register payment for an existing unpaid advance"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get the adelanto
            adelanto = await conn.fetchrow(
                "SELECT * FROM finanzas2.cont_adelanto_empleado WHERE id = $1 AND empresa_id = $2", id
            , empresa_id)
            if not adelanto:
                raise HTTPException(404, "Adelanto no encontrado")
            if adelanto['pagado']:
                raise HTTPException(400, "Este adelanto ya fue pagado")
            
            # Get centro_costo and linea_negocio from employee
            emp_info = await conn.fetchrow("""
                SELECT centro_costo_id, linea_negocio_id
                FROM finanzas2.cont_empleado_detalle WHERE tercero_id = $1
            """, adelanto['empleado_id'])
            cc_id = emp_info['centro_costo_id'] if emp_info else None
            ln_id = emp_info['linea_negocio_id'] if emp_info else None
            
            # Create pago
            pago_numero = await generate_pago_number(conn, 'egreso', empresa_id)
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, monto_total, notas, centro_costo_id, linea_negocio_id, empresa_id)
                VALUES ($1, 'egreso', CURRENT_DATE, $2, $3, $4, $5, $6, $7)
                RETURNING id
            """, pago_numero, cuenta_financiera_id, adelanto['monto'], 
                "Pago de adelanto a empleado", cc_id, ln_id, empresa_id)
            pago_id = pago['id']
            
            # Create pago detalle
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_detalle 
                (pago_id, cuenta_financiera_id, medio_pago, monto, empresa_id)
                VALUES ($1, $2, $3, $4, $5)
            """, pago_id, cuenta_financiera_id, medio_pago, adelanto['monto'], empresa_id)
            
            # Update cuenta saldo
            await conn.execute("""
                UPDATE finanzas2.cont_cuenta_financiera 
                SET saldo_actual = saldo_actual - $1 WHERE id = $2
            """, adelanto['monto'], cuenta_financiera_id)
            
            # Update adelanto
            row = await conn.fetchrow("""
                UPDATE finanzas2.cont_adelanto_empleado 
                SET pagado = TRUE, pago_id = $1
                WHERE id = $2
                RETURNING *
            """, pago_id, id)
            
            # Create pago aplicacion
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                VALUES ($1, 'adelanto', $2, $3, $4)
            """, pago_id, id, adelanto['monto'], empresa_id)
            
            # Get empleado nombre
            emp = await conn.fetchrow(
                "SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1 AND empresa_id = $2", 
                row['empleado_id']
            , empresa_id)
            result = dict(row)
            result['empleado_nombre'] = emp['nombre'] if emp else None
            
            return result

@api_router.put("/adelantos/{id}", response_model=Adelanto)
async def update_adelanto(id: int, data: AdelantoCreate, empresa_id: int = Depends(get_empresa_id)):
    """Update an existing advance"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if adelanto exists and is not already paid/discounted
        existing = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_adelanto_empleado WHERE id = $1 AND empresa_id = $2", id
        , empresa_id)
        if not existing:
            raise HTTPException(404, "Adelanto no encontrado")
        if existing['pagado'] or existing['descontado']:
            raise HTTPException(400, "No se puede editar un adelanto pagado o descontado")
        
        row = await conn.fetchrow("""
            UPDATE finanzas2.cont_adelanto_empleado 
            SET empleado_id = $1, fecha = $2, monto = $3, motivo = $4
            WHERE id = $5
            RETURNING *
        """, data.empleado_id, data.fecha, data.monto, data.motivo, id)
        
        # Get empleado nombre
        emp = await conn.fetchrow(
            "SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1 AND empresa_id = $2", 
            row['empleado_id']
        , empresa_id)
        result = dict(row)
        result['empleado_nombre'] = emp['nombre'] if emp else None
        
        return result

@api_router.delete("/adelantos/{id}")
async def delete_adelanto(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Delete an advance (only if not paid or discounted)"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if adelanto exists and can be deleted
        existing = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_adelanto_empleado WHERE id = $1 AND empresa_id = $2", id
        , empresa_id)
        if not existing:
            raise HTTPException(404, "Adelanto no encontrado")
        if existing['pagado']:
            raise HTTPException(400, "No se puede eliminar un adelanto pagado. Primero anule el pago.")
        if existing['descontado']:
            raise HTTPException(400, "No se puede eliminar un adelanto ya descontado en planilla")
        
        await conn.execute(
            "DELETE FROM finanzas2.cont_adelanto_empleado WHERE id = $1 AND empresa_id = $2", id
        , empresa_id)
        return {"message": "Adelanto eliminado"}

# =====================
# PLANILLA
# =====================
@api_router.get("/planillas", response_model=List[Planilla])
async def list_planillas(empresa_id: Optional[int] = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        if empresa_id is not None:
            rows = await conn.fetch("""
                SELECT DISTINCT p.* FROM finanzas2.cont_planilla p
                JOIN finanzas2.cont_planilla_detalle pd ON pd.planilla_id = p.id
                JOIN finanzas2.cont_tercero t ON pd.empleado_id = t.id
                WHERE t.empresa_id = $1
                ORDER BY p.periodo DESC
            """, empresa_id)
        else:
            rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_planilla ORDER BY periodo DESC
            """)
        
        result = []
        for row in rows:
            planilla_dict = dict(row)
            detalles = await conn.fetch("""
                SELECT pd.*, t.nombre as empleado_nombre
                FROM finanzas2.cont_planilla_detalle pd
                LEFT JOIN finanzas2.cont_tercero t ON pd.empleado_id = t.id
                WHERE pd.planilla_id = $1
            """, row['id'])
            planilla_dict['detalles'] = [dict(d) for d in detalles]
            result.append(planilla_dict)
        
        return result

@api_router.post("/planillas", response_model=Planilla)
async def create_planilla(data: PlanillaCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Calculate totals
            total_bruto = sum(d.salario_base + d.bonificaciones for d in data.detalles)
            total_adelantos = sum(d.adelantos for d in data.detalles)
            total_descuentos = sum(d.otros_descuentos for d in data.detalles)
            total_neto = total_bruto - total_adelantos - total_descuentos
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_planilla 
                (periodo, fecha_inicio, fecha_fin, total_bruto, total_adelantos, 
                 total_descuentos, total_neto, estado, empresa_id)
                VALUES ($1, TO_DATE($2, 'YYYY-MM-DD'), TO_DATE($3, 'YYYY-MM-DD'), $4, $5, $6, $7, 'borrador', $8)
                RETURNING *
            """, data.periodo, safe_date_param(data.fecha_inicio), safe_date_param(data.fecha_fin), total_bruto,
                total_adelantos, total_descuentos, total_neto, empresa_id)
            
            planilla_id = row['id']
            planilla_dict = dict(row)
            detalles_list = []
            
            for detalle in data.detalles:
                neto_pagar = detalle.salario_base + detalle.bonificaciones - detalle.adelantos - detalle.otros_descuentos
                detalle_row = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_planilla_detalle 
                    (planilla_id, empleado_id, salario_base, bonificaciones, adelantos, 
                     otros_descuentos, neto_pagar, empresa_id)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                    RETURNING *
                """, planilla_id, detalle.empleado_id, detalle.salario_base, detalle.bonificaciones,
                    detalle.adelantos, detalle.otros_descuentos, neto_pagar, empresa_id)
                
                # Get employee name
                emp = await conn.fetchrow("SELECT nombre FROM finanzas2.cont_tercero WHERE id = $1 AND empresa_id = $2", detalle.empleado_id, empresa_id)
                detalle_dict = dict(detalle_row)
                detalle_dict['empleado_nombre'] = emp['nombre'] if emp else None
                detalles_list.append(detalle_dict)
            
            planilla_dict['detalles'] = detalles_list
            return planilla_dict

async def get_planilla(id: int, empresa_id: int = None) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        if empresa_id:
            row = await conn.fetchrow("SELECT * FROM finanzas2.cont_planilla WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        else:
            row = await conn.fetchrow("SELECT * FROM finanzas2.cont_planilla WHERE id = $1", id)
        if not row:
            raise HTTPException(404, "Planilla not found")
        
        planilla_dict = dict(row)
        detalles = await conn.fetch("""
            SELECT pd.*, t.nombre as empleado_nombre
            FROM finanzas2.cont_planilla_detalle pd
            LEFT JOIN finanzas2.cont_tercero t ON pd.empleado_id = t.id
            WHERE pd.planilla_id = $1
        """, id)
        planilla_dict['detalles'] = [dict(d) for d in detalles]
        
        return planilla_dict

@api_router.get("/planillas/{id}", response_model=Planilla)
async def get_planilla_endpoint(id: int, empresa_id: int = Depends(get_empresa_id)):
    return await get_planilla(id, empresa_id)

@api_router.post("/planillas/{id}/pagar", response_model=Planilla)
async def pagar_planilla(id: int, cuenta_financiera_id: int = Query(...), empresa_id: int = Depends(get_empresa_id)):
    """Pay the entire planilla"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            planilla = await conn.fetchrow("SELECT * FROM finanzas2.cont_planilla WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not planilla:
                raise HTTPException(404, "Planilla not found")
            if planilla['estado'] == 'pagada':
                raise HTTPException(400, "Planilla already paid")
            
            # Get centro_costo and linea_negocio from first employee with details
            emp_info = await conn.fetchrow("""
                SELECT ed.centro_costo_id, ed.linea_negocio_id
                FROM finanzas2.cont_planilla_detalle pd
                JOIN finanzas2.cont_empleado_detalle ed ON pd.empleado_id = ed.tercero_id
                WHERE pd.planilla_id = $1 AND (ed.centro_costo_id IS NOT NULL OR ed.linea_negocio_id IS NOT NULL)
                LIMIT 1
            """, id)
            
            cc_id = emp_info['centro_costo_id'] if emp_info else None
            ln_id = emp_info['linea_negocio_id'] if emp_info else None
            
            # Create pago
            pago_numero = await generate_pago_number(conn, 'egreso', empresa_id)
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, monto_total, notas, centro_costo_id, linea_negocio_id, empresa_id)
                VALUES ($1, 'egreso', $2, $3, $4, $5, $6, $7, $8)
                RETURNING id
            """, pago_numero, datetime.now().date(), cuenta_financiera_id, 
                planilla['total_neto'], f"Pago planilla {planilla['periodo']}", cc_id, ln_id, empresa_id)
            
            pago_id = pago['id']
            
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_detalle 
                (pago_id, cuenta_financiera_id, medio_pago, monto, empresa_id)
                VALUES ($1, $2, 'transferencia', $3, $4)
            """, pago_id, cuenta_financiera_id, planilla['total_neto'], empresa_id)
            
            await conn.execute("""
                UPDATE finanzas2.cont_cuenta_financiera 
                SET saldo_actual = saldo_actual - $1 WHERE id = $2
            """, planilla['total_neto'], cuenta_financiera_id)
            
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                VALUES ($1, 'planilla', $2, $3, $4)
            """, pago_id, id, planilla['total_neto'], empresa_id)
            
            # Mark adelantos as descontado
            await conn.execute("""
                UPDATE finanzas2.cont_adelanto_empleado 
                SET descontado = TRUE, planilla_id = $1
                WHERE empleado_id IN (SELECT empleado_id FROM finanzas2.cont_planilla_detalle WHERE planilla_id = $1)
                AND pagado = TRUE AND descontado = FALSE
            """, id)
            
            # Update planilla status
            await conn.execute("""
                UPDATE finanzas2.cont_planilla SET estado = 'pagada', pago_id = $1 WHERE id = $2
            """, pago_id, id)
            
            return await get_planilla(id, empresa_id)

@api_router.delete("/planillas/{id}")
async def delete_planilla(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Delete a planilla (only if in draft status)"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Check if planilla exists and can be deleted
        existing = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_planilla WHERE id = $1 AND empresa_id = $2", id
        , empresa_id)
        if not existing:
            raise HTTPException(404, "Planilla no encontrada")
        if existing['estado'] == 'pagada':
            raise HTTPException(400, "No se puede eliminar una planilla pagada")
        
        # Delete detalles first (due to foreign key)
        await conn.execute(
            "DELETE FROM finanzas2.cont_planilla_detalle WHERE planilla_id = $1", id
        )
        # Delete planilla
        await conn.execute(
            "DELETE FROM finanzas2.cont_planilla WHERE id = $1 AND empresa_id = $2", id
        , empresa_id)
        return {"message": "Planilla eliminada"}

# =====================
# VENTAS POS (Odoo)
# =====================
@api_router.get("/ventas-pos", response_model=List[VentaPOS])
async def list_ventas_pos(
    estado: Optional[str] = None,
    company_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    search: Optional[str] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["v.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if estado:
            conditions.append(f"estado_local = ${idx}")
            params.append(estado)
            idx += 1
        if company_id:
            conditions.append(f"company_id = ${idx}")
            params.append(company_id)
            idx += 1
        if fecha_desde:
            # date_order is in UTC, Lima is UTC-5, so add 5 hours to filter
            conditions.append(f"date_order >= ${idx}")
            params.append(datetime.combine(fecha_desde, datetime.min.time()) + timedelta(hours=5))
            idx += 1
        if fecha_hasta:
            conditions.append(f"date_order <= ${idx}")
            params.append(datetime.combine(fecha_hasta, datetime.max.time()) + timedelta(hours=5))
            idx += 1
        
        # Advanced search: num_comp, partner_name, name (order reference)
        if search:
            search_pattern = f"%{search}%"
            conditions.append(f"(num_comp ILIKE ${idx} OR partner_name ILIKE ${idx} OR name ILIKE ${idx})")
            params.append(search_pattern)
            idx += 1
        
        query = f"""
            SELECT v.*, 
                   COALESCE((SELECT SUM(p.monto) FROM finanzas2.cont_venta_pos_pago p WHERE p.venta_pos_id = v.id), 0) as pagos_asignados,
                   COALESCE((SELECT COUNT(*) FROM finanzas2.cont_venta_pos_pago p WHERE p.venta_pos_id = v.id), 0) as num_pagos,
                   COALESCE((SELECT SUM(pa.monto_aplicado) 
                            FROM finanzas2.cont_pago_aplicacion pa 
                            WHERE pa.tipo_documento = 'venta_pos' AND pa.documento_id = v.id), 0) as pagos_oficiales,
                   COALESCE((SELECT COUNT(*) 
                            FROM finanzas2.cont_pago_aplicacion pa 
                            WHERE pa.tipo_documento = 'venta_pos' AND pa.documento_id = v.id), 0) as num_pagos_oficiales
            FROM finanzas2.cont_venta_pos v
            WHERE {' AND '.join(conditions)}
            ORDER BY v.date_order DESC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.post("/ventas-pos/sync")
async def sync_ventas_pos(company: str = "ambission", days_back: int = 30, empresa_id: int = Depends(get_empresa_id)):
    """Sync POS orders from Odoo"""
    try:
        odoo = OdooService(company=company)
        if not odoo.authenticate():
            raise HTTPException(401, f"Could not authenticate with Odoo ({company})")
        
        orders = odoo.get_pos_orders(days_back=days_back)
        
        if not orders:
            return {"message": "No orders found", "synced": 0}
        
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.execute("SET search_path TO finanzas2, public")
            
            synced = 0
            for order in orders:
                try:
                    # Extract values with proper handling of Odoo's tuple format
                    odoo_id = order.get('id')
                    
                    # Parse date_order (comes as string from Odoo)
                    date_order_str = order.get('date_order')
                    if date_order_str and isinstance(date_order_str, str):
                        date_order = datetime.strptime(date_order_str, '%Y-%m-%d %H:%M:%S')
                    else:
                        date_order = date_order_str
                    
                    name = order.get('name')
                    tipo_comp = order.get('tipo_comp')
                    num_comp = order.get('num_comp')
                    
                    # partner_id comes as [id, name] tuple from Odoo
                    partner_id = order['partner_id'][0] if isinstance(order.get('partner_id'), list) else order.get('partner_id')
                    partner_name = order['partner_id'][1] if isinstance(order.get('partner_id'), list) else None
                    
                    # x_tienda comes as [id, name] tuple
                    tienda_id = order['x_tienda'][0] if isinstance(order.get('x_tienda'), list) else None
                    tienda_name = order['x_tienda'][1] if isinstance(order.get('x_tienda'), list) else order.get('x_tienda')
                    
                    # vendedor_id comes as [id, name] tuple
                    vendedor_id = order['vendedor_id'][0] if isinstance(order.get('vendedor_id'), list) else order.get('vendedor_id')
                    vendedor_name = order['vendedor_id'][1] if isinstance(order.get('vendedor_id'), list) else None
                    
                    # company_id comes as [id, name] tuple
                    company_id = order['company_id'][0] if isinstance(order.get('company_id'), list) else order.get('company_id')
                    company_name = order['company_id'][1] if isinstance(order.get('company_id'), list) else None
                    
                    # x_pagos: convert False to None or empty string
                    x_pagos = order.get('x_pagos')
                    if x_pagos == False or x_pagos == 'False':
                        x_pagos = None
                    
                    quantity_total = order.get('quantity_pos_order')
                    amount_total = order.get('amount_total')
                    state = order.get('state')
                    
                    reserva_pendiente = order.get('x_reserva_pendiente', 0)
                    reserva_facturada = order.get('x_reserva_facturada', 0)
                    is_cancel = order.get('is_cancel', False)
                    
                    # order_cancel: can be False, string, or [id, name] list
                    order_cancel_raw = order.get('order_cancel')
                    if order_cancel_raw == False or order_cancel_raw == 'False':
                        order_cancel = None
                    elif isinstance(order_cancel_raw, list) and len(order_cancel_raw) > 1:
                        order_cancel = order_cancel_raw[1]  # Use the name part
                    else:
                        order_cancel = order_cancel_raw
                    
                    reserva = order.get('reserva', False)
                    is_credit = order.get('is_credit', False)
                    
                    # reserva_use_id: can be False, int, or [id, name] list
                    reserva_use_id_raw = order.get('reserva_use_id')
                    if reserva_use_id_raw == False or reserva_use_id_raw == 'False':
                        reserva_use_id = None
                    elif isinstance(reserva_use_id_raw, list) and len(reserva_use_id_raw) > 0:
                        reserva_use_id = reserva_use_id_raw[0]  # Use the ID part
                    else:
                        reserva_use_id = reserva_use_id_raw
                    
                    # Check if already exists
                    existing = await conn.fetchrow("""
                        SELECT id, estado_local FROM finanzas2.cont_venta_pos 
                        WHERE odoo_id = $1
                    """, odoo_id)
                    
                    # If exists and already processed (confirmada/credito/descartada), skip update
                    if existing:
                        existing_estado = existing['estado_local']
                        if existing_estado in ['confirmada', 'credito', 'descartada']:
                            # Skip this order - already processed, don't re-import
                            continue
                    
                    if existing:
                        # Update existing record
                        await conn.execute("""
                            UPDATE finanzas2.cont_venta_pos SET
                                date_order = $2,
                                name = $3,
                                tipo_comp = $4,
                                num_comp = $5,
                                partner_id = $6,
                                partner_name = $7,
                                tienda_id = $8,
                                tienda_name = $9,
                                vendedor_id = $10,
                                vendedor_name = $11,
                                company_id = $12,
                                company_name = $13,
                                x_pagos = $14,
                                quantity_total = $15,
                                amount_total = $16,
                                state = $17,
                                reserva_pendiente = $18,
                                reserva_facturada = $19,
                                is_cancel = $20,
                                order_cancel = $21,
                                reserva = $22,
                                is_credit = $23,
                                reserva_use_id = $24,
                                synced_at = NOW()
                            WHERE odoo_id = $1
                        """, odoo_id, date_order, name, tipo_comp, num_comp,
                            partner_id, partner_name, tienda_id, tienda_name,
                            vendedor_id, vendedor_name, company_id, company_name,
                            x_pagos, quantity_total, amount_total, state,
                            reserva_pendiente, reserva_facturada, is_cancel,
                            order_cancel, reserva, is_credit, reserva_use_id)
                    else:
                        # Insert new record
                        await conn.execute("""
                            INSERT INTO finanzas2.cont_venta_pos 
                            (odoo_id, date_order, name, tipo_comp, num_comp,
                             partner_id, partner_name, tienda_id, tienda_name,
                             vendedor_id, vendedor_name, company_id, company_name,
                             x_pagos, quantity_total, amount_total, state,
                             reserva_pendiente, reserva_facturada, is_cancel,
                             order_cancel, reserva, is_credit, reserva_use_id,
                             synced_at, empresa_id)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10,
                                    $11, $12, $13, $14, $15, $16, $17, $18, $19,
                                    $20, $21, $22, $23, $24, NOW(), $25)
                        """, odoo_id, date_order, name, tipo_comp, num_comp,
                            partner_id, partner_name, tienda_id, tienda_name,
                            vendedor_id, vendedor_name, company_id, company_name,
                            x_pagos, quantity_total, amount_total, state,
                            reserva_pendiente, reserva_facturada, is_cancel,
                            order_cancel, reserva, is_credit, reserva_use_id, empresa_id)
                    
                    # Sync product lines for this order
                    try:
                        # Get the local venta_pos id
                        local_venta = await conn.fetchrow("""
                            SELECT id FROM finanzas2.cont_venta_pos WHERE odoo_id = $1
                        """, odoo_id)
                        
                        if local_venta:
                            venta_pos_id = local_venta['id']
                            
                            # Delete existing lines (to avoid duplicates on re-sync)
                            await conn.execute("""
                                DELETE FROM finanzas2.cont_venta_pos_linea WHERE venta_pos_id = $1
                            """, venta_pos_id)
                            
                            # Get lines from Odoo with marca and tipo
                            lines = odoo.get_order_lines(odoo_id)
                            
                            for line in lines:
                                product_name = line['product_id'][1] if isinstance(line.get('product_id'), list) else 'Producto'
                                product_id_val = line['product_id'][0] if isinstance(line.get('product_id'), list) else line.get('product_id')
                                
                                await conn.execute("""
                                    INSERT INTO finanzas2.cont_venta_pos_linea
                                    (venta_pos_id, odoo_line_id, product_id, product_name, product_code,
                                     qty, price_unit, price_subtotal, price_subtotal_incl, discount,
                                     marca, tipo, empresa_id)
                                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13)
                                """, venta_pos_id, line.get('id'), product_id_val, product_name,
                                    line.get('product_code', ''), line.get('qty', 0),
                                    line.get('price_unit', 0), line.get('price_subtotal', 0),
                                    line.get('price_subtotal_incl', 0), line.get('discount', 0),
                                    line.get('marca', ''), line.get('tipo', ''), empresa_id)
                    except Exception as line_error:
                        logger.error(f"Error syncing lines for order {odoo_id}: {line_error}")
                        # Continue even if lines fail - main order is synced
                    
                    synced += 1
                    
                except Exception as row_error:
                    logger.error(f"Error processing order {order.get('id')}: {row_error}")
                    continue
        
        return {"message": f"Synced {synced} orders from {company}", "synced": synced}
        
    except Exception as e:
        logger.error(f"Error syncing from Odoo: {e}")
        raise HTTPException(500, f"Error syncing: {str(e)}")

@api_router.post("/ventas-pos/{id}/confirmar")
async def confirmar_venta_pos(id: int, empresa_id: int = Depends(get_empresa_id)):
    """
    Confirm a POS sale. 
    NOTE: Should have assigned payments before confirming, but this is not enforced yet.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # TODO: Validate that venta has assigned payments
        # For now, just check if it exists
        venta = await conn.fetchrow("SELECT * FROM finanzas2.cont_venta_pos WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not venta:
            raise HTTPException(404, "Venta not found")
        
        # Check if already processed
        if venta['estado_local'] in ['confirmada', 'credito', 'descartada']:
            raise HTTPException(400, f"Venta already {venta['estado_local']}")
        
        # TODO: Check if has payments assigned
        # For now, just confirm
        await conn.execute("""
            UPDATE finanzas2.cont_venta_pos SET estado_local = 'confirmada' WHERE id = $1
        """, id)
        
        return {"message": "Venta confirmada"}

@api_router.post("/ventas-pos/{id}/credito")
async def marcar_credito_venta_pos(id: int, fecha_vencimiento: Optional[date] = None, empresa_id: int = Depends(get_empresa_id)):
    """Mark sale as credit and create CxC"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            venta = await conn.fetchrow("SELECT * FROM finanzas2.cont_venta_pos WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not venta:
                raise HTTPException(404, "Venta not found")
            
            # Create CxC
            cxc = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_cxc 
                (venta_pos_id, monto_original, saldo_pendiente, fecha_vencimiento, estado, empresa_id)
                VALUES ($1, $2, $2, TO_DATE($3, 'YYYY-MM-DD'), 'pendiente', $4)
                RETURNING id
            """, id, venta['amount_total'], safe_date_param(fecha_vencimiento or (datetime.now().date() + timedelta(days=30))), empresa_id)
            
            await conn.execute("""
                UPDATE finanzas2.cont_venta_pos SET estado_local = 'credito', cxc_id = $1, is_credit = TRUE WHERE id = $2
            """, cxc['id'], id)
            
            return {"message": "Venta marcada como crédito", "cxc_id": cxc['id']}

@api_router.post("/ventas-pos/{id}/descartar")
async def descartar_venta_pos(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        await conn.execute("""
            UPDATE finanzas2.cont_venta_pos SET estado_local = 'descartada', is_cancel = TRUE WHERE id = $1
        """, id)
        return {"message": "Venta descartada"}

@api_router.post("/ventas-pos/{id}/desconfirmar")
async def desconfirmar_venta_pos(id: int, empresa_id: int = Depends(get_empresa_id)):
    """
    Desconfirmar una venta POS confirmada.
    - Elimina los pagos oficiales (cont_pago, cont_pago_detalle, cont_pago_aplicacion)
    - Restaura los pagos temporales en cont_venta_pos_pago
    - Cambia el estado de la venta a 'pendiente'
    
    Este patrón se puede reutilizar para facturas y otros documentos.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Verificar que la venta existe y está confirmada
            venta = await conn.fetchrow("SELECT * FROM finanzas2.cont_venta_pos WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not venta:
                raise HTTPException(404, "Venta not found")
            
            if venta['estado_local'] != 'confirmada':
                raise HTTPException(400, f"La venta debe estar confirmada para desconfirmarla. Estado actual: {venta['estado_local']}")
            
            # Obtener los pagos oficiales vinculados a esta venta
            pagos_oficiales = await conn.fetch("""
                SELECT p.id as pago_id, pd.medio_pago, pd.monto, pd.referencia, 
                       p.fecha, pd.cuenta_financiera_id, p.notas
                FROM finanzas2.cont_pago_aplicacion pa
                JOIN finanzas2.cont_pago p ON p.id = pa.pago_id
                LEFT JOIN finanzas2.cont_pago_detalle pd ON pd.pago_id = p.id
                WHERE pa.tipo_documento = 'venta_pos' AND pa.documento_id = $1
            """, id)
            
            # IMPORTANTE: Eliminar pagos temporales existentes (evita duplicados)
            await conn.execute("""
                DELETE FROM finanzas2.cont_venta_pos_pago WHERE venta_pos_id = $1
            """, id)
            
            # Restaurar los pagos en cont_venta_pos_pago (tabla temporal) si existen
            if pagos_oficiales:
                for pago in pagos_oficiales:
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_venta_pos_pago 
                        (venta_pos_id, forma_pago, cuenta_financiera_id, monto, referencia, fecha_pago, observaciones, empresa_id)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                    """, id, pago['medio_pago'], pago['cuenta_financiera_id'], 
                        pago['monto'], pago['referencia'], pago['fecha'], 
                        pago['notas'] or 'Pago restaurado desde confirmación', empresa_id)
                
                # Eliminar los pagos oficiales (CASCADE eliminará cont_pago_detalle y cont_pago_aplicacion)
                pago_ids = list(set(p['pago_id'] for p in pagos_oficiales))
                for pago_id in pago_ids:
                    await conn.execute("""
                        DELETE FROM finanzas2.cont_pago WHERE id = $1
                    """, pago_id)
            
            # Cambiar estado de la venta a pendiente
            await conn.execute("""
                UPDATE finanzas2.cont_venta_pos SET estado_local = 'pendiente' WHERE id = $1
            """, id)
            
            return {
                "message": "Venta desconfirmada exitosamente",
                "pagos_restaurados": len(pagos_oficiales),
                "nuevo_estado": "pendiente"
            }

# Endpoints for payment management
@api_router.get("/ventas-pos/{id}/pagos")
async def get_pagos_venta_pos(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Get all payments assigned to a POS sale"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        pagos = await conn.fetch("""
            SELECT id, venta_pos_id, forma_pago, monto, referencia, 
                   fecha_pago, observaciones, created_at
            FROM finanzas2.cont_venta_pos_pago
            WHERE venta_pos_id = $1
            ORDER BY created_at DESC
        """, id)
        
        return [dict(p) for p in pagos]

@api_router.get("/ventas-pos/{id}/pagos-oficiales")
async def get_pagos_oficiales_venta_pos(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Get official payments from cont_pago for a confirmed POS sale"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        pagos = await conn.fetch("""
            SELECT p.id, p.numero, p.fecha, pd.medio_pago as forma_pago, pd.monto,
                   pd.referencia, p.notas as observaciones, cf.nombre as cuenta_nombre
            FROM finanzas2.cont_pago_aplicacion pa
            JOIN finanzas2.cont_pago p ON p.id = pa.pago_id
            LEFT JOIN finanzas2.cont_pago_detalle pd ON pd.pago_id = p.id
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON cf.id = pd.cuenta_financiera_id
            WHERE pa.tipo_documento = 'venta_pos' AND pa.documento_id = $1
            ORDER BY p.fecha DESC, p.id DESC
        """, id)
        
        return [dict(p) for p in pagos]

@api_router.get("/ventas-pos/{id}/lineas")
async def get_lineas_venta_pos(id: int, empresa_id: int = Depends(get_empresa_id)):
    """Get product lines for a POS sale with marca and tipo for business line analysis"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        lineas = await conn.fetch("""
            SELECT 
                id,
                product_name,
                product_code,
                qty,
                price_unit,
                price_subtotal,
                price_subtotal_incl,
                discount,
                marca,
                tipo
            FROM finanzas2.cont_venta_pos_linea
            WHERE venta_pos_id = $1
            ORDER BY id ASC
        """, id)
        
        return [dict(l) for l in lineas]

@api_router.post("/ventas-pos/{id}/pagos")
async def add_pago_venta_pos(id: int, pago: dict, empresa_id: int = Depends(get_empresa_id)):
    """Add a payment to a POS sale. Auto-confirms if total matches."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get venta
            venta = await conn.fetchrow("SELECT * FROM finanzas2.cont_venta_pos WHERE id = $1 AND empresa_id = $2", id, empresa_id)
            if not venta:
                raise HTTPException(404, "Venta not found")
            
            if venta['estado_local'] != 'pendiente':
                raise HTTPException(400, f"Venta already {venta['estado_local']}")
            
            # Insert payment
            await conn.execute("""
                INSERT INTO finanzas2.cont_venta_pos_pago 
                (venta_pos_id, forma_pago, cuenta_financiera_id, monto, referencia, fecha_pago, observaciones, empresa_id)
                VALUES ($1, $2, $3, $4, $5, TO_DATE($6, 'YYYY-MM-DD'), $7, $8)
            """, id, pago.get('forma_pago'), int(pago.get('cuenta_financiera_id')), pago.get('monto'), 
                pago.get('referencia'), pago.get('fecha_pago'), pago.get('observaciones'), empresa_id)
            
            # Calculate total payments
            total_pagos = await conn.fetchval("""
                SELECT COALESCE(SUM(monto), 0) FROM finanzas2.cont_venta_pos_pago
                WHERE venta_pos_id = $1
            """, id)
            
            # Auto-confirm if total matches
            amount_total = float(venta['amount_total'])
            if abs(float(total_pagos) - amount_total) < 0.01:  # Match with 0.01 tolerance
                await conn.execute("""
                    UPDATE finanzas2.cont_venta_pos SET estado_local = 'confirmada' WHERE id = $1
                """, id)
                
                # ✅ IMPORTANTE: Crear pagos en el módulo de Pagos (cont_pago)
                # Obtener todos los pagos asignados de esta venta
                pagos_venta = await conn.fetch("""
                    SELECT * FROM finanzas2.cont_venta_pos_pago WHERE venta_pos_id = $1
                """, id)
                
                # Crear un pago en cont_pago por cada pago asignado
                for pago_item in pagos_venta:
                    # Generar número de pago INGRESO (PAG-I-YYYY-XXXXX) porque son ventas
                    last_pago = await conn.fetchval("""
                        SELECT numero FROM finanzas2.cont_pago 
                        WHERE tipo = 'ingreso' 
                        ORDER BY id DESC LIMIT 1
                    """)
                    
                    if last_pago and '-' in last_pago:
                        parts = last_pago.split('-')
                        if len(parts) >= 3:
                            num = int(parts[-1]) + 1
                        else:
                            num = 1
                    else:
                        num = 1
                    
                    numero_pago = f"PAG-I-{datetime.now().year}-{num:05d}"
                    
                    # Insertar en cont_pago como INGRESO (ventas = dinero que entra)
                    pago_result = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_pago 
                        (numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, referencia, notas, empresa_id)
                        VALUES ($1, 'ingreso', $2::date, $3, 1, $4, $5, $6, $7)
                        RETURNING id
                    """, numero_pago, pago_item['fecha_pago'], pago_item['cuenta_financiera_id'],
                        pago_item['monto'], pago_item['referencia'],
                        f"Pago de venta POS {venta['name']} - {pago_item['observaciones'] or ''}", empresa_id)
                    
                    pago_id = pago_result['id']
                    
                    # Insertar detalle del pago
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_pago_detalle 
                        (pago_id, cuenta_financiera_id, medio_pago, monto, referencia, empresa_id)
                        VALUES ($1, $2, $3, $4, $5, $6)
                    """, pago_id, pago_item['cuenta_financiera_id'], pago_item['forma_pago'], 
                        pago_item['monto'], pago_item['referencia'], empresa_id)
                    
                    # Vincular el pago con la venta POS en cont_pago_aplicacion
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_pago_aplicacion
                        (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                        VALUES ($1, 'venta_pos', $2, $3, $4)
                    """, pago_id, id, pago_item['monto'], empresa_id)
                
                return {
                    "message": "Pago agregado y venta confirmada automáticamente", 
                    "total_pagos": float(total_pagos),
                    "auto_confirmed": True,
                    "pagos_registrados": len(pagos_venta)
                }
            
            return {
                "message": "Pago agregado", 
                "total_pagos": float(total_pagos),
                "faltante": amount_total - float(total_pagos),
                "auto_confirmed": False
            }

@api_router.put("/ventas-pos/{venta_id}/pagos/{pago_id}")
async def update_pago_venta_pos(venta_id: int, pago_id: int, pago: dict, empresa_id: int = Depends(get_empresa_id)):
    """Update a payment assigned to a POS sale"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        try:
            # Update the payment
            await conn.execute("""
                UPDATE finanzas2.cont_venta_pos_pago 
                SET forma_pago = $1,
                    cuenta_financiera_id = $2,
                    monto = $3,
                    referencia = $4,
                    fecha_pago = TO_DATE($5, 'YYYY-MM-DD'),
                    observaciones = $6
                WHERE id = $7 AND venta_pos_id = $8
            """, pago['forma_pago'], pago.get('cuenta_financiera_id'), 
                pago['monto'], pago.get('referencia'), 
                pago.get('fecha_pago'), pago.get('observaciones'),
                pago_id, venta_id)
            
            return {"message": "Pago actualizado correctamente"}
        except Exception as e:
            logger.error(f"Error updating payment: {e}")
            raise HTTPException(500, f"Error al actualizar pago: {str(e)}")

@api_router.delete("/ventas-pos/{venta_id}/pagos/{pago_id}")
async def delete_pago_venta_pos(venta_id: int, pago_id: int, empresa_id: int = Depends(get_empresa_id)):
    """Delete a payment from a POS sale"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        await conn.execute("""
            DELETE FROM finanzas2.cont_venta_pos_pago 
            WHERE id = $1 AND venta_pos_id = $2
        """, pago_id, venta_id)
        return {"message": "Pago eliminado"}


# =====================
# CXC (Cuentas por Cobrar)
# =====================
@api_router.get("/cxc", response_model=List[CXC])
async def list_cxc(estado: Optional[str] = None, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["cxc.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if estado:
            conditions.append(f"cxc.estado = ${idx}")
            params.append(estado)
            idx += 1
        
        query = f"""
            SELECT cxc.*, t.nombre as cliente_nombre
            FROM finanzas2.cont_cxc cxc
            LEFT JOIN finanzas2.cont_tercero t ON cxc.cliente_id = t.id
            WHERE {' AND '.join(conditions)}
            ORDER BY cxc.fecha_vencimiento ASC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

# =====================
# CXP (Cuentas por Pagar)
# =====================
@api_router.get("/cxp")
async def list_cxp(estado: Optional[str] = None, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["cxp.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if estado:
            conditions.append(f"cxp.estado = ${idx}")
            params.append(estado)
            idx += 1
        
        query = f"""
            SELECT cxp.*, t.nombre as proveedor_nombre, fp.numero as factura_numero
            FROM finanzas2.cont_cxp cxp
            LEFT JOIN finanzas2.cont_tercero t ON cxp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_factura_proveedor fp ON cxp.factura_id = fp.id
            WHERE {' AND '.join(conditions)}
            ORDER BY cxp.fecha_vencimiento ASC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

# =====================
# PRESUPUESTOS
# =====================
@api_router.get("/presupuestos", response_model=List[Presupuesto])
async def list_presupuestos(anio: Optional[int] = None, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        if anio:
            rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_presupuesto WHERE anio = $1 ORDER BY version DESC
            """, anio)
        else:
            rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_presupuesto ORDER BY anio DESC, version DESC
            """)
        
        result = []
        for row in rows:
            pres_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT pl.*, c.nombre as categoria_nombre
                FROM finanzas2.cont_presupuesto_linea pl
                LEFT JOIN finanzas2.cont_categoria c ON pl.categoria_id = c.id
                WHERE pl.presupuesto_id = $1
            """, row['id'])
            pres_dict['lineas'] = [dict(l) for l in lineas]
            result.append(pres_dict)
        
        return result

@api_router.post("/presupuestos", response_model=Presupuesto)
async def create_presupuesto(data: PresupuestoCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get next version for this year
            version = await conn.fetchval("""
                SELECT COALESCE(MAX(version), 0) + 1 FROM finanzas2.cont_presupuesto WHERE anio = $1
            """, data.anio) or 1
            
            row = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_presupuesto 
                (nombre, anio, version, estado, notas, empresa_id)
                VALUES ($1, $2, $3, 'borrador', $4, $5)
                RETURNING *
            """, data.nombre, data.anio, version, data.notas, empresa_id)
            
            presupuesto_id = row['id']
            
            for linea in data.lineas:
                await conn.execute("""
                    INSERT INTO finanzas2.cont_presupuesto_linea 
                    (presupuesto_id, categoria_id, centro_costo_id, linea_negocio_id, mes, monto_presupuestado, empresa_id)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                """, presupuesto_id, linea.categoria_id, linea.centro_costo_id,
                    linea.linea_negocio_id, linea.mes, linea.monto_presupuestado, empresa_id)
            
            return await get_presupuesto(presupuesto_id)

async def get_presupuesto(id: int) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("SELECT * FROM finanzas2.cont_presupuesto WHERE id = $1 AND empresa_id = $2", id, empresa_id)
        if not row:
            raise HTTPException(404, "Presupuesto not found")
        
        pres_dict = dict(row)
        lineas = await conn.fetch("""
            SELECT pl.*, c.nombre as categoria_nombre
            FROM finanzas2.cont_presupuesto_linea pl
            LEFT JOIN finanzas2.cont_categoria c ON pl.categoria_id = c.id
            WHERE pl.presupuesto_id = $1
        """, id)
        pres_dict['lineas'] = [dict(l) for l in lineas]
        
        return pres_dict

@api_router.get("/presupuestos/{id}", response_model=Presupuesto)
async def get_presupuesto_endpoint(id: int, empresa_id: int = Depends(get_empresa_id)):
    return await get_presupuesto(id)

# =====================
# CONCILIACION BANCARIA
# =====================
@api_router.post("/conciliacion/previsualizar-excel")
async def previsualizar_excel_banco(
    file: UploadFile = File(...),
    banco: str = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    """Preview bank movements from Excel before importing"""
    import io
    from datetime import datetime as dt
    
    try:
        content = await file.read()
        
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Find header row
        header_row = 1
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(row):
                row_str = ' '.join([str(c or '') for c in row]).lower()
                if 'fecha' in row_str or 'f. valor' in row_str or 'f. operación' in row_str:
                    header_row = idx
                    break
        
        preview_data = []
        
        for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 51, values_only=True):
            if not row or not any(row):
                continue
            
            fecha = None
            descripcion = None
            referencia = None
            monto = None
            
            try:
                if banco == 'BCP':
                    # BCP: Nº, Fecha, Fecha valuta, Descripción operación, Monto, Saldo, Sucursal, Operación-Número
                    fecha = row[1] if len(row) > 1 else None
                    descripcion = row[3] if len(row) > 3 else None
                    monto_val = row[4] if len(row) > 4 else None
                    referencia = row[7] if len(row) > 7 else None
                    
                    if monto_val:
                        monto = float(monto_val) if not isinstance(monto_val, str) else float(str(monto_val).replace(',', ''))
                            
                elif banco == 'BBVA':
                    # BBVA: N°, F. Operación, F. Valor, Código, Nº. Doc., Concepto, Importe, Oficina
                    # Skip "Saldo Final" rows
                    concepto = row[5] if len(row) > 5 else None
                    if concepto and 'saldo final' in str(concepto).lower():
                        continue
                    
                    fecha = row[1] if len(row) > 1 else None
                    referencia = row[4] if len(row) > 4 else None
                    descripcion = row[5] if len(row) > 5 else None
                    importe = row[6] if len(row) > 6 else None
                    
                    if importe:
                        monto = float(importe) if not isinstance(importe, str) else float(str(importe).replace(',', ''))
                            
                elif banco == 'IBK':
                    # IBK format: Nº, Fecha de operación, Fecha de proceso, Nro. de operación, 
                    #             Movimiento, Descripción, Canal, Cargo, Abono, Saldo contable
                    # Skip metadata rows (first 13 rows usually) and rows without valid number
                    if len(row) < 10:
                        continue
                    
                    # Check if it's a data row (should have a number in first column)
                    try:
                        nro = int(row[0]) if row[0] and str(row[0]).strip() else None
                        if not nro:
                            continue
                    except (ValueError, TypeError):
                        continue
                    
                    fecha = row[1] if len(row) > 1 else None  # Fecha de operación
                    referencia = row[3] if len(row) > 3 else None  # Nro. de operación
                    descripcion = row[5] if len(row) > 5 else None  # Descripción
                    cargo = row[7] if len(row) > 7 else None  # Cargo
                    abono = row[8] if len(row) > 8 else None  # Abono
                    
                    # Calculate monto (negative for cargo, positive for abono)
                    if cargo and cargo != '' and str(cargo).strip() != 'nan':
                        try:
                            cargo_val = float(cargo) if not isinstance(cargo, str) else float(str(cargo).replace(',', ''))
                            monto = -abs(cargo_val)  # Ensure it's negative
                        except (ValueError, TypeError):
                            monto = None
                    elif abono and abono != '' and str(abono).strip() != 'nan':
                        try:
                            abono_val = float(abono) if not isinstance(abono, str) else float(str(abono).replace(',', ''))
                            monto = abs(abono_val)  # Ensure it's positive
                        except (ValueError, TypeError):
                            monto = None
                    else:
                        monto = None
                        
                else:
                    # PERSONALIZADO
                    fecha = row[0] if len(row) > 0 else None
                    descripcion = row[1] if len(row) > 1 else None
                    referencia = row[2] if len(row) > 2 else None
                    monto = row[3] if len(row) > 3 else None
                
                # Parse date
                if fecha:
                    if isinstance(fecha, dt):
                        fecha = fecha.date().isoformat()
                    elif hasattr(fecha, 'date'):
                        fecha = fecha.date().isoformat()
                    elif isinstance(fecha, str):
                        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                            try:
                                fecha = dt.strptime(fecha.strip(), fmt).date().isoformat()
                                break
                            except:
                                continue
                
                if not fecha or monto is None:
                    continue
                
                preview_data.append({
                    "fecha": fecha,
                    "banco": banco,  # Use selected bank name
                    "referencia": str(referencia)[:200] if referencia else "",
                    "descripcion": str(descripcion)[:500] if descripcion else "",
                    "monto": float(monto) if monto else 0.0
                })
                
                if len(preview_data) >= 50:
                    break
                    
            except Exception as row_error:
                logger.warning(f"Error parsing row: {row_error}")
                continue
        
        return {
            "preview": preview_data,
            "total_rows": len(preview_data)
        }
        
    except Exception as e:
        logger.error(f"Error previewing Excel: {e}")
        raise HTTPException(500, f"Error al previsualizar: {str(e)}")

@api_router.post("/conciliacion/importar-excel")
async def importar_excel_banco(
    file: UploadFile = File(...),
    cuenta_financiera_id: int = Query(...),
    banco: str = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    """Import bank movements from Excel - UPSERT based on banco + referencia"""
    import io
    from datetime import datetime as dt
    
    pool = await get_pool()
    
    try:
        content = await file.read()
        
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        async with pool.acquire() as conn:
            await conn.execute("SET search_path TO finanzas2, public")
            
            imported = 0
            updated = 0
            skipped = 0
            
            # Find header row
            header_row = 1
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(row):
                    row_str = ' '.join([str(c or '') for c in row]).lower()
                    if 'fecha' in row_str or 'f. valor' in row_str or 'f. operación' in row_str:
                        header_row = idx
                        break
            
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or not any(row):
                    continue
                
                fecha = None
                descripcion = None
                referencia = None
                monto = None
                
                try:
                    if banco == 'BCP':
                        # BCP: Nº, Fecha, Fecha valuta, Descripción operación, Monto, Saldo, Sucursal, Operación-Número
                        fecha = row[1] if len(row) > 1 else None
                        descripcion = row[3] if len(row) > 3 else None
                        monto_val = row[4] if len(row) > 4 else None
                        referencia = row[7] if len(row) > 7 else None
                        
                        if monto_val:
                            monto = float(monto_val) if not isinstance(monto_val, str) else float(str(monto_val).replace(',', ''))
                                
                    elif banco == 'BBVA':
                        # BBVA: N°, F. Operación, F. Valor, Código, Nº. Doc., Concepto, Importe, Oficina
                        # Skip "Saldo Final" rows
                        concepto = row[5] if len(row) > 5 else None
                        if concepto and 'saldo final' in str(concepto).lower():
                            continue
                        
                        fecha = row[1] if len(row) > 1 else None
                        referencia = row[4] if len(row) > 4 else None
                        descripcion = row[5] if len(row) > 5 else None
                        importe = row[6] if len(row) > 6 else None
                        
                        if importe:
                            monto = float(importe) if not isinstance(importe, str) else float(str(importe).replace(',', ''))
                                
                    elif banco == 'IBK':
                        # IBK format: Nº, Fecha de operación, Fecha de proceso, Nro. de operación, 
                        #             Movimiento, Descripción, Canal, Cargo, Abono, Saldo contable
                        # Skip metadata rows (first 13 rows usually) and rows without valid number
                        if len(row) < 10:
                            continue
                        
                        # Check if it's a data row (should have a number in first column)
                        try:
                            nro = int(row[0]) if row[0] and str(row[0]).strip() else None
                            if not nro:
                                continue
                        except (ValueError, TypeError):
                            continue
                        
                        fecha = row[1] if len(row) > 1 else None  # Fecha de operación
                        referencia = row[3] if len(row) > 3 else None  # Nro. de operación
                        descripcion = row[5] if len(row) > 5 else None  # Descripción
                        cargo = row[7] if len(row) > 7 else None  # Cargo
                        abono = row[8] if len(row) > 8 else None  # Abono
                        
                        # Calculate monto (negative for cargo, positive for abono)
                        if cargo and cargo != '' and str(cargo).strip() != 'nan':
                            try:
                                cargo_val = float(cargo) if not isinstance(cargo, str) else float(str(cargo).replace(',', ''))
                                monto = -abs(cargo_val)  # Ensure it's negative
                            except (ValueError, TypeError):
                                monto = None
                        elif abono and abono != '' and str(abono).strip() != 'nan':
                            try:
                                abono_val = float(abono) if not isinstance(abono, str) else float(str(abono).replace(',', ''))
                                monto = abs(abono_val)  # Ensure it's positive
                            except (ValueError, TypeError):
                                monto = None
                        else:
                            monto = None
                            
                    else:
                        # PERSONALIZADO
                        fecha = row[0] if len(row) > 0 else None
                        descripcion = row[1] if len(row) > 1 else None
                        referencia = row[2] if len(row) > 2 else None
                        monto = row[3] if len(row) > 3 else None
                    
                    # Parse date
                    if fecha:
                        if isinstance(fecha, dt):
                            fecha = fecha.date()
                        elif hasattr(fecha, 'date'):
                            fecha = fecha.date()
                        elif isinstance(fecha, str):
                            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                                try:
                                    fecha = dt.strptime(fecha.strip(), fmt).date()
                                    break
                                except:
                                    continue
                    
                    # Skip rows without valid date or monto
                    if not fecha or monto is None:
                        continue
                    
                    # Clean fields
                    ref_clean = str(referencia).strip()[:200] if referencia else ''
                    desc_clean = str(descripcion).strip()[:500] if descripcion else ''
                    
                    # Check if exists and if it's already reconciled
                    existing = await conn.fetchrow("""
                        SELECT id, procesado FROM finanzas2.cont_banco_mov_raw 
                        WHERE cuenta_financiera_id = $1 
                          AND banco = $2
                          AND COALESCE(referencia, '') = $3
                          AND fecha = $4
                    """, cuenta_financiera_id, banco, ref_clean, fecha)
                    
                    if existing:
                        if existing['procesado']:
                            skipped += 1
                            continue
                        else:
                            # Update existing record
                            await conn.execute("""
                                UPDATE finanzas2.cont_banco_mov_raw 
                                SET descripcion = $1, monto = $2, banco_excel = $3
                                WHERE id = $4
                            """, desc_clean, monto, banco, existing['id'])
                            updated += 1
                    else:
                        # Insert new record
                        await conn.execute("""
                            INSERT INTO finanzas2.cont_banco_mov_raw 
                            (cuenta_financiera_id, banco, fecha, descripcion, referencia, monto, banco_excel, procesado, empresa_id)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, FALSE, $8)
                        """, cuenta_financiera_id, banco, fecha, desc_clean, 
                            ref_clean if ref_clean else None, monto, banco, empresa_id)
                        imported += 1
                    
                except Exception as row_error:
                    logger.warning(f"Error parsing row: {row_error}")
                    continue
        
        return {
            "message": f"Importados: {imported}, Actualizados: {updated}, Omitidos (ya conciliados): {skipped}",
            "imported": imported,
            "updated": updated,
            "skipped": skipped
        }
        
    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        raise HTTPException(500, f"Error al importar: {str(e)}")


@api_router.get("/conciliacion/historial")
async def get_historial_conciliaciones(empresa_id: int = Depends(get_empresa_id)):
    """Get all conciliation detail lines in flat format for the history view"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        rows = await conn.fetch("""
            SELECT 
                cl.id,
                cl.conciliacion_id,
                cl.banco_mov_id,
                cl.pago_id,
                cl.monto,
                cl.tipo,
                cl.conciliado,
                cl.created_at,
                -- banco mov data (from raw table)
                bm.fecha as fecha_banco,
                bm.descripcion as descripcion_banco,
                bm.referencia as ref_banco,
                bm.monto as monto_banco,
                -- pago data
                p.numero as numero_sistema,
                p.tipo as tipo_sistema,
                p.fecha as fecha_sistema,
                p.notas as descripcion_sistema,
                p.monto_total as monto_sistema,
                -- cuenta data
                cf.nombre as cuenta_nombre,
                cf.banco as banco
            FROM finanzas2.cont_conciliacion_linea cl
            LEFT JOIN finanzas2.cont_banco_mov_raw bm ON cl.banco_mov_id = bm.id
            LEFT JOIN finanzas2.cont_pago p ON cl.pago_id = p.id
            LEFT JOIN finanzas2.cont_conciliacion c ON cl.conciliacion_id = c.id
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
            WHERE cl.empresa_id = $1
            ORDER BY cl.created_at DESC
        """, empresa_id)
        
        result = []
        for r in rows:
            fecha_banco = r['fecha_banco']
            fecha_sistema = r['fecha_sistema']
            result.append({
                "id": r['id'],
                "conciliacion_id": r['conciliacion_id'],
                "banco_id": r['banco_mov_id'],
                "sistema_id": r['pago_id'],
                "banco_mov_id": r['banco_mov_id'],
                "pago_id": r['pago_id'],
                "fecha_banco": fecha_banco.isoformat() if fecha_banco else None,
                "fecha_sistema": fecha_sistema.isoformat() if fecha_sistema else None,
                "banco": r['banco'] or r['cuenta_nombre'] or '-',
                "ref_banco": r['ref_banco'] or '',
                "descripcion_banco": r['descripcion_banco'] or '',
                "monto": float(r['monto_banco'] or r['monto'] or 0),
                "numero_sistema": r['numero_sistema'] or '',
                "tipo_sistema": r['tipo_sistema'] or r['tipo'] or '',
                "descripcion_sistema": r['descripcion_sistema'] or '',
                "monto_sistema": float(r['monto_sistema'] or r['monto'] or 0),
                "conciliado": r['conciliado'],
            })
        
        return result

@api_router.post("/conciliacion/desconciliar")
async def desconciliar_movimientos(data: dict, empresa_id: int = Depends(get_empresa_id)):
    """Unreconcile movements"""
    banco_id = data.get('banco_id')
    pago_id = data.get('pago_id')
    
    if not banco_id and not pago_id:
        raise HTTPException(400, "Se requiere al menos banco_id o pago_id")
    
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Reset bank movement if provided
        if banco_id:
            await conn.execute("""
                UPDATE finanzas2.cont_banco_mov_raw 
                SET procesado = FALSE, conciliado = FALSE 
                WHERE id = $1
            """, banco_id)
        
        # Reset system payment if provided
        if pago_id:
            await conn.execute("""
                UPDATE finanzas2.cont_pago 
                SET conciliado = FALSE 
                WHERE id = $1
            """, pago_id)
        
        # Delete the conciliacion_linea records
        if banco_id and pago_id:
            await conn.execute("""
                DELETE FROM finanzas2.cont_conciliacion_linea 
                WHERE (banco_mov_id = $1 OR pago_id = $2) AND empresa_id = $3
            """, banco_id, pago_id, empresa_id)
        elif pago_id:
            await conn.execute("""
                DELETE FROM finanzas2.cont_conciliacion_linea 
                WHERE pago_id = $1 AND empresa_id = $2
            """, pago_id, empresa_id)
        elif banco_id:
            await conn.execute("""
                DELETE FROM finanzas2.cont_conciliacion_linea 
                WHERE banco_mov_id = $1 AND empresa_id = $2
            """, banco_id, empresa_id)
    
    return {"message": "Movimientos desconciliados exitosamente"}


@api_router.get("/conciliacion/movimientos-banco")
async def list_movimientos_banco(
    cuenta_financiera_id: Optional[int] = None,
    procesado: Optional[bool] = None,
    conciliado: Optional[bool] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        conditions = ["empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        
        if cuenta_financiera_id:
            conditions.append(f"cuenta_financiera_id = ${idx}")
            params.append(cuenta_financiera_id)
            idx += 1
        if procesado is not None:
            conditions.append(f"procesado = ${idx}")
            params.append(procesado)
            idx += 1
        # Support both procesado and conciliado (they should be the same after migration)
        if conciliado is not None:
            conditions.append(f"conciliado = ${idx}")
            params.append(conciliado)
            idx += 1
        
        query = f"""
            SELECT * FROM finanzas2.cont_banco_mov_raw
            WHERE {' AND '.join(conditions)}
            ORDER BY fecha DESC
        """
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]

@api_router.get("/conciliaciones", response_model=List[Conciliacion])
async def list_conciliaciones(cuenta_financiera_id: Optional[int] = None, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        if cuenta_financiera_id:
            rows = await conn.fetch("""
                SELECT c.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_conciliacion c
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
                WHERE c.cuenta_financiera_id = $1
                ORDER BY c.fecha_fin DESC
            """, cuenta_financiera_id)
        else:
            rows = await conn.fetch("""
                SELECT c.*, cf.nombre as cuenta_nombre
                FROM finanzas2.cont_conciliacion c
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
                ORDER BY c.fecha_fin DESC
            """)
        
        result = []
        for row in rows:
            conc_dict = dict(row)
            lineas = await conn.fetch("""
                SELECT * FROM finanzas2.cont_conciliacion_linea WHERE conciliacion_id = $1
            """, row['id'])
            conc_dict['lineas'] = [dict(l) for l in lineas]
            result.append(conc_dict)
        
        return result

@api_router.post("/conciliacion/conciliar")
async def conciliar_movimientos(
    banco_ids: List[int] = Query(...),
    pago_ids: List[int] = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    """Mark bank movements and system payments as reconciled and create historical record"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get account info from first bank movement
            cuenta_id = None
            if banco_ids:
                mov = await conn.fetchrow("""
                    SELECT cuenta_financiera_id, fecha 
                    FROM finanzas2.cont_banco_mov_raw 
                    WHERE id = $1
                """, banco_ids[0])
                if mov:
                    cuenta_id = mov['cuenta_financiera_id']
            
            # Mark bank movements as processed AND conciliado
            if banco_ids:
                await conn.execute("""
                    UPDATE finanzas2.cont_banco_mov_raw 
                    SET procesado = TRUE, conciliado = TRUE 
                    WHERE id = ANY($1::int[])
                """, banco_ids)
            
            # Mark payments as reconciled (add conciliado field if needed)
            if pago_ids:
                # First ensure the column exists
                try:
                    await conn.execute("""
                        ALTER TABLE finanzas2.cont_pago 
                        ADD COLUMN IF NOT EXISTS conciliado BOOLEAN DEFAULT FALSE
                    """)
                except:
                    pass
                
                await conn.execute("""
                    UPDATE finanzas2.cont_pago 
                    SET conciliado = TRUE 
                    WHERE id = ANY($1::int[])
                """, pago_ids)
            
            # Create historical conciliacion record
            if cuenta_id:
                from datetime import date
                today = date.today()
                
                # Get total amount
                total_banco = 0
                if banco_ids:
                    result = await conn.fetchrow("""
                        SELECT SUM(monto) as total FROM finanzas2.cont_banco_mov_raw 
                        WHERE id = ANY($1::int[])
                    """, banco_ids)
                    total_banco = float(result['total']) if result['total'] else 0
                
                # Create conciliacion record
                conciliacion = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_conciliacion 
                    (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_final, estado, notas, empresa_id)
                    VALUES ($1, $2, $2, $3, 'completado', $4, $5)
                    RETURNING id
                """, cuenta_id, today, total_banco,
                    f"Conciliación: {len(banco_ids)} mov. banco + {len(pago_ids)} mov. sistema", empresa_id)
                
                conciliacion_id = conciliacion['id']
                
                # Create detail records - pair bank movements with payments
                for i, pago_id in enumerate(pago_ids):
                    pago_info = await conn.fetchrow("""
                        SELECT monto_total FROM finanzas2.cont_pago WHERE id = $1
                    """, pago_id)
                    
                    # Link to corresponding bank movement if available
                    banco_mov_id = banco_ids[i] if i < len(banco_ids) else None
                    
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_conciliacion_linea 
                        (conciliacion_id, banco_mov_id, pago_id, tipo, monto, conciliado, empresa_id)
                        VALUES ($1, $2, $3, 'pago', $4, TRUE, $5)
                    """, conciliacion_id, banco_mov_id, pago_id, pago_info['monto_total'] if pago_info else 0, empresa_id)
                
                # Also create records for unmatched bank movements
                for i in range(len(pago_ids), len(banco_ids)):
                    banco_info = await conn.fetchrow(
                        "SELECT monto FROM finanzas2.cont_banco_mov_raw WHERE id = $1", banco_ids[i]
                    )
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_conciliacion_linea 
                        (conciliacion_id, banco_mov_id, tipo, monto, conciliado, empresa_id)
                        VALUES ($1, $2, 'banco', $3, TRUE, $4)
                    """, conciliacion_id, banco_ids[i], banco_info['monto'] if banco_info else 0, empresa_id)
        
        return {
            "message": f"Conciliados {len(banco_ids)} movimientos del banco y {len(pago_ids)} del sistema",
            "banco_conciliados": len(banco_ids),
            "sistema_conciliados": len(pago_ids)
        }

@api_router.post("/conciliacion/crear-gasto-bancario")
async def crear_gasto_desde_movimientos_bancarios(
    banco_ids: List[int] = Query(...),
    categoria_id: int = Query(...),
    descripcion: Optional[str] = Query("Gastos bancarios agrupados"),
    cuenta_financiera_id: int = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    """
    Create a gasto (expense) from multiple bank movements.
    Useful for grouping bank charges like ITF, commissions, etc.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        async with conn.transaction():
            # Get the bank movements to sum
            movimientos = await conn.fetch("""
                SELECT * FROM finanzas2.cont_banco_mov_raw 
                WHERE id = ANY($1::int[])
            """, banco_ids)
            
            if not movimientos:
                raise HTTPException(404, "No se encontraron movimientos bancarios")
            
            # Check if any are already reconciled
            already_conciliados = [m for m in movimientos if m['conciliado']]
            if already_conciliados:
                raise HTTPException(400, f"{len(already_conciliados)} movimientos ya están conciliados")
            
            # Calculate total (absolute value for negative amounts)
            total = sum(abs(float(m['monto'])) for m in movimientos)
            
            # Get next gasto number
            last_gasto = await conn.fetchrow("""
                SELECT numero FROM finanzas2.cont_gasto 
                ORDER BY id DESC LIMIT 1
            """)
            
            if last_gasto and last_gasto['numero']:
                try:
                    last_num = int(last_gasto['numero'].split('-')[1])
                    numero = f"GAS-{last_num + 1:06d}"
                except:
                    numero = f"GAS-{len(movimientos):06d}"
            else:
                numero = "GAS-000001"
            
            # Create the gasto
            gasto = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_gasto 
                (numero, fecha, beneficiario_nombre, moneda_id, subtotal, igv, total,
                 tipo_documento, numero_documento, notas, empresa_id)
                VALUES ($1, CURRENT_DATE, $2, 1, $3, 0, $3, 'gasto_bancario', $4, $5, $6)
                RETURNING id
            """, numero, 'Banco', total, numero, descripcion, empresa_id)
            
            gasto_id = gasto['id']
            
            # Insert gasto line
            await conn.execute("""
                INSERT INTO finanzas2.cont_gasto_linea 
                (gasto_id, categoria_id, descripcion, importe, igv_aplica, empresa_id)
                VALUES ($1, $2, $3, $4, FALSE, $5)
            """, gasto_id, categoria_id, f"{descripcion} ({len(movimientos)} movimientos)", total, empresa_id)
            
            # Create pago (automatic payment)
            pago_numero = f"PAG-E-{numero}"
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago 
                (numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, notas, empresa_id)
                VALUES ($1, 'egreso', CURRENT_DATE, $2, 1, $3, $4, $5)
                RETURNING id
            """, pago_numero, cuenta_financiera_id, total, f"Pago automático de {descripcion}", empresa_id)
            
            pago_id = pago['id']
            
            # Insert pago detalle
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_detalle 
                (pago_id, cuenta_financiera_id, medio_pago, monto, empresa_id)
                VALUES ($1, $2, 'cargo_bancario', $3, $4)
            """, pago_id, cuenta_financiera_id, total, empresa_id)
            
            # Link pago to gasto
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion 
                (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                VALUES ($1, 'gasto', $2, $3, $4)
            """, pago_id, gasto_id, total, empresa_id)
            
            # Mark bank movements as reconciled
            await conn.execute("""
                UPDATE finanzas2.cont_banco_mov_raw 
                SET procesado = TRUE, conciliado = TRUE 
                WHERE id = ANY($1::int[])
            """, banco_ids)
            
            # Mark pago as reconciled
            try:
                await conn.execute("""
                    ALTER TABLE finanzas2.cont_pago 
                    ADD COLUMN IF NOT EXISTS conciliado BOOLEAN DEFAULT FALSE
                """)
            except:
                pass
            
            await conn.execute("""
                UPDATE finanzas2.cont_pago 
                SET conciliado = TRUE 
                WHERE id = $1
            """, pago_id)
            
            # Create historical conciliacion record
            from datetime import date
            today = date.today()
            
            conciliacion = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_conciliacion 
                (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_final, estado, notas, empresa_id)
                VALUES ($1, $2, $2, $3, 'completado', $4, $5)
                RETURNING id
            """, cuenta_financiera_id, today, total,
                f"Gasto bancario automático: {descripcion} ({len(banco_ids)} movimientos)", empresa_id)
            
            conciliacion_id = conciliacion['id']
            
            # Create detail record for the pago
            await conn.execute("""
                INSERT INTO finanzas2.cont_conciliacion_linea 
                (conciliacion_id, pago_id, tipo, documento_id, monto, conciliado, empresa_id)
                VALUES ($1, $2, 'gasto', $3, $4, TRUE, $5)
            """, conciliacion_id, pago_id, gasto_id, total, empresa_id)
        
        return {
            "message": f"Gasto creado exitosamente con {len(banco_ids)} movimientos bancarios",
            "gasto_id": gasto_id,
            "gasto_numero": numero,
            "pago_id": pago_id,
            "total": total,
            "movimientos_conciliados": len(banco_ids)
        }

@api_router.post("/conciliaciones", response_model=Conciliacion)
async def create_conciliacion(data: ConciliacionCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_conciliacion 
            (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_inicial, saldo_final, notas, empresa_id)
            VALUES ($1, TO_DATE($2, 'YYYY-MM-DD'), TO_DATE($3, 'YYYY-MM-DD'), $4, $5, $6, $7)
            RETURNING *
        """, data.cuenta_financiera_id, safe_date_param(data.fecha_inicio), safe_date_param(data.fecha_fin),
            data.saldo_inicial, data.saldo_final, data.notas, empresa_id)
        
        result = dict(row)
        result['lineas'] = []
        return result

# =====================
# REPORTES
# =====================
@api_router.get("/reportes/flujo-caja")
async def reporte_flujo_caja(
    fecha_desde: date = Query(...),
    fecha_hasta: date = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        rows = await conn.fetch("""
            SELECT p.fecha, p.tipo, p.monto_total, p.notas,
                   cf.nombre as cuenta
            FROM finanzas2.cont_pago p
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON p.cuenta_financiera_id = cf.id
            WHERE p.fecha BETWEEN $1 AND $2 AND p.empresa_id = $3
            ORDER BY p.fecha ASC
        """, fecha_desde, fecha_hasta, empresa_id)
        
        resultado = []
        saldo_acumulado = 0
        
        for row in rows:
            if row['tipo'] == 'ingreso':
                saldo_acumulado += float(row['monto_total'])
            else:
                saldo_acumulado -= float(row['monto_total'])
            
            resultado.append({
                "fecha": row['fecha'].isoformat(),
                "concepto": row['notas'] or row['cuenta'],
                "tipo": row['tipo'],
                "monto": float(row['monto_total']),
                "saldo_acumulado": saldo_acumulado
            })
        
        return resultado

@api_router.get("/reportes/estado-resultados")
async def reporte_estado_resultados(
    fecha_desde: date = Query(...),
    fecha_hasta: date = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Ingresos from ventas POS
        ingresos = await conn.fetchval("""
            SELECT COALESCE(SUM(amount_total), 0) 
            FROM finanzas2.cont_venta_pos 
            WHERE date_order BETWEEN $1 AND $2 AND estado_local = 'confirmada' AND empresa_id = $3
        """, datetime.combine(fecha_desde, datetime.min.time()), 
            datetime.combine(fecha_hasta, datetime.max.time()), empresa_id) or 0
        
        # Egresos from pagos
        egresos_data = await conn.fetch("""
            SELECT c.nombre as categoria, COALESCE(SUM(gl.importe), 0) as monto
            FROM finanzas2.cont_gasto g
            JOIN finanzas2.cont_gasto_linea gl ON g.id = gl.gasto_id
            LEFT JOIN finanzas2.cont_categoria c ON gl.categoria_id = c.id
            WHERE g.fecha BETWEEN $1 AND $2 AND g.empresa_id = $3
            GROUP BY c.nombre
        """, fecha_desde, fecha_hasta, empresa_id)
        
        total_egresos = sum(float(e['monto']) for e in egresos_data)
        
        return {
            "ingresos": [
                {"categoria": "Ventas", "tipo": "ingreso", "monto": float(ingresos)}
            ],
            "egresos": [{"categoria": e['categoria'] or "Sin categoría", "tipo": "egreso", "monto": float(e['monto'])} for e in egresos_data],
            "total_ingresos": float(ingresos),
            "total_egresos": total_egresos,
            "resultado_neto": float(ingresos) - total_egresos
        }

@api_router.get("/reportes/balance-general")
async def reporte_balance_general(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        
        # Activos: Bancos y cajas
        activos_bancos = await conn.fetch("""
            SELECT nombre, saldo_actual FROM finanzas2.cont_cuenta_financiera WHERE activo = TRUE AND empresa_id = $1
        """, empresa_id)
        
        total_activos = sum(float(a['saldo_actual']) for a in activos_bancos)
        
        # Pasivos: CxP + Letras pendientes
        total_cxp = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) FROM finanzas2.cont_cxp WHERE estado NOT IN ('pagado', 'anulada') AND empresa_id = $1
        """, empresa_id) or 0
        
        total_letras = await conn.fetchval("""
            SELECT COALESCE(SUM(saldo_pendiente), 0) FROM finanzas2.cont_letra WHERE estado IN ('pendiente', 'parcial') AND empresa_id = $1
        """, empresa_id) or 0
        
        total_pasivos = float(total_cxp) + float(total_letras)
        
        patrimonio = total_activos - total_pasivos
        
        return {
            "activos": [{"cuenta": a['nombre'], "tipo": "activo", "monto": float(a['saldo_actual'])} for a in activos_bancos],
            "pasivos": [
                {"cuenta": "Cuentas por Pagar", "tipo": "pasivo", "monto": float(total_cxp)},
                {"cuenta": "Letras por Pagar", "tipo": "pasivo", "monto": float(total_letras)}
            ],
            "total_activos": total_activos,
            "total_pasivos": total_pasivos,
            "patrimonio": patrimonio
        }

# =============================================
# CUENTAS CONTABLES (Plan de Cuentas)
# =============================================
@api_router.get("/cuentas-contables", response_model=List[CuentaContable])
async def list_cuentas_contables(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM finanzas2.cont_cuenta WHERE empresa_id = $1 ORDER BY codigo",
            empresa_id
        )
        return [dict(r) for r in rows]

@api_router.post("/cuentas-contables", response_model=CuentaContable)
async def create_cuenta_contable(data: CuentaContableCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_cuenta (empresa_id, codigo, nombre, tipo, es_activa)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING *
        """, empresa_id, data.codigo, data.nombre, data.tipo, data.es_activa)
        return dict(row)

@api_router.put("/cuentas-contables/{id}", response_model=CuentaContable)
async def update_cuenta_contable(id: int, data: CuentaContableUpdate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        updates = []
        values = []
        idx = 1
        for field in ['codigo', 'nombre', 'tipo', 'es_activa']:
            val = getattr(data, field, None)
            if val is not None:
                updates.append(f"{field} = ${idx}")
                values.append(val)
                idx += 1
        if not updates:
            raise HTTPException(400, "No hay campos para actualizar")
        values.extend([empresa_id, id])
        row = await conn.fetchrow(
            f"UPDATE finanzas2.cont_cuenta SET {', '.join(updates)}, updated_at = NOW() WHERE empresa_id = ${idx} AND id = ${idx+1} RETURNING *",
            *values
        )
        if not row:
            raise HTTPException(404, "Cuenta no encontrada")
        return dict(row)

@api_router.delete("/cuentas-contables/{id}")
async def delete_cuenta_contable(id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM finanzas2.cont_cuenta WHERE id = $1 AND empresa_id = $2", id, empresa_id
        )
        if result == "DELETE 0":
            raise HTTPException(404, "Cuenta no encontrada")
        return {"ok": True}

@api_router.post("/cuentas-contables/seed-peru")
async def seed_cuentas_peru(empresa_id: int = Depends(get_empresa_id)):
    """Seed minimum chart of accounts for a Peruvian company. Idempotent (skips existing codes)."""
    CUENTAS_PERU = [
        ("101",  "Caja",                                "ACTIVO"),
        ("1041", "Banco BCP",                           "ACTIVO"),
        ("1042", "Banco BBVA",                          "ACTIVO"),
        ("1043", "Banco Interbank",                     "ACTIVO"),
        ("121",  "Cuentas por cobrar comerciales",      "ACTIVO"),
        ("4212", "Cuentas por pagar comerciales",       "PASIVO"),
        ("4011", "IGV por pagar",                       "PASIVO"),
        ("4012", "IGV crédito fiscal (compras)",        "IMPUESTO"),
        ("4099", "Otros tributos / tasas",              "IMPUESTO"),
        ("6311", "Alquileres",                          "GASTO"),
        ("6312", "Mantenimiento y reparaciones",        "GASTO"),
        ("6321", "Energía eléctrica / agua / servicios","GASTO"),
        ("6331", "Transporte / fletes",                 "GASTO"),
        ("6341", "Publicidad y marketing",              "GASTO"),
        ("6351", "Honorarios / servicios profesionales","GASTO"),
        ("6361", "Comisiones y gastos bancarios",       "GASTO"),
        ("6371", "Útiles / suministros",                "GASTO"),
        ("6399", "Otros servicios y gastos",            "GASTO"),
        ("201",  "Mercaderías / Inventario",            "ACTIVO"),
        ("691",  "Costo de ventas",                     "GASTO"),
    ]
    pool = await get_pool()
    async with pool.acquire() as conn:
        inserted = 0
        async with conn.transaction():
            for codigo, nombre, tipo in CUENTAS_PERU:
                result = await conn.execute("""
                    INSERT INTO finanzas2.cont_cuenta (empresa_id, codigo, nombre, tipo)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (empresa_id, codigo) DO NOTHING
                """, empresa_id, codigo, nombre, tipo)
                if result == "INSERT 0 1":
                    inserted += 1

            # Set defaults in cont_config_empresa
            cta_gastos = await conn.fetchval(
                "SELECT id FROM finanzas2.cont_cuenta WHERE empresa_id=$1 AND codigo='6399'", empresa_id)
            cta_igv = await conn.fetchval(
                "SELECT id FROM finanzas2.cont_cuenta WHERE empresa_id=$1 AND codigo='4012'", empresa_id)
            cta_xpagar = await conn.fetchval(
                "SELECT id FROM finanzas2.cont_cuenta WHERE empresa_id=$1 AND codigo='4212'", empresa_id)
            cta_otrib = await conn.fetchval(
                "SELECT id FROM finanzas2.cont_cuenta WHERE empresa_id=$1 AND codigo='4099'", empresa_id)

            await conn.execute("""
                INSERT INTO finanzas2.cont_config_empresa (empresa_id, cta_gastos_default_id, cta_igv_default_id, cta_xpagar_default_id, cta_otrib_default_id)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (empresa_id) DO UPDATE SET
                    cta_gastos_default_id = COALESCE(finanzas2.cont_config_empresa.cta_gastos_default_id, $2),
                    cta_igv_default_id = COALESCE(finanzas2.cont_config_empresa.cta_igv_default_id, $3),
                    cta_xpagar_default_id = COALESCE(finanzas2.cont_config_empresa.cta_xpagar_default_id, $4),
                    cta_otrib_default_id = COALESCE(finanzas2.cont_config_empresa.cta_otrib_default_id, $5)
            """, empresa_id, cta_gastos, cta_igv, cta_xpagar, cta_otrib)

        return {"inserted": inserted, "total": len(CUENTAS_PERU), "message": f"Seed completado: {inserted} cuentas nuevas insertadas"}

# =============================================
# CONFIG CONTABLE POR EMPRESA
# =============================================
@api_router.get("/config-contable", response_model=ConfigEmpresaContable)
async def get_config_contable(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_config_empresa WHERE empresa_id = $1", empresa_id
        )
        if not row:
            return {"empresa_id": empresa_id, "cta_gastos_default_id": None, "cta_igv_default_id": None, "cta_xpagar_default_id": None, "cta_otrib_default_id": None}
        return dict(row)

@api_router.put("/config-contable", response_model=ConfigEmpresaContable)
async def update_config_contable(data: ConfigEmpresaContable, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_config_empresa (empresa_id, cta_gastos_default_id, cta_igv_default_id, cta_xpagar_default_id, cta_otrib_default_id)
            VALUES ($1, $2, $3, $4, $5)
            ON CONFLICT (empresa_id) DO UPDATE SET
                cta_gastos_default_id = $2, cta_igv_default_id = $3, cta_xpagar_default_id = $4, cta_otrib_default_id = $5
            RETURNING *
        """, empresa_id, data.cta_gastos_default_id, data.cta_igv_default_id, data.cta_xpagar_default_id, data.cta_otrib_default_id)
        return dict(row)

# =============================================
# RETENCION / DETRACCION (manual fields per document)
# =============================================
@api_router.get("/retencion-detalle")
async def get_retencion_detalle(origen_tipo: str, origen_id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("""
            SELECT * FROM finanzas2.cont_retencion_detalle
            WHERE empresa_id = $1 AND origen_tipo = $2 AND origen_id = $3
        """, empresa_id, origen_tipo, origen_id)
        if not row:
            return None
        r = dict(row)
        r.pop('id', None)
        r.pop('created_at', None)
        r.pop('updated_at', None)
        return r

@api_router.put("/retencion-detalle")
async def upsert_retencion_detalle(origen_tipo: str, origen_id: int, data: RetencionDetalle, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_retencion_detalle
            (empresa_id, origen_tipo, origen_id, r_doc, r_numero, r_fecha, d_numero, d_fecha,
             retencion_01, pdb_ndes, codtasa, ind_ret, b_imp, igv_ret)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14)
            ON CONFLICT (empresa_id, origen_tipo, origen_id) DO UPDATE SET
                r_doc=$4, r_numero=$5, r_fecha=$6, d_numero=$7, d_fecha=$8,
                retencion_01=$9, pdb_ndes=$10, codtasa=$11, ind_ret=$12, b_imp=$13, igv_ret=$14,
                updated_at=NOW()
            RETURNING *
        """, empresa_id, origen_tipo, origen_id,
            data.r_doc, data.r_numero,
            data.r_fecha if data.r_fecha else None,
            data.d_numero,
            data.d_fecha if data.d_fecha else None,
            data.retencion_01, data.pdb_ndes, data.codtasa, data.ind_ret, data.b_imp, data.igv_ret)
        r = dict(row)
        r.pop('id', None)
        r.pop('created_at', None)
        r.pop('updated_at', None)
        return r

# =============================================
# CUENTAS FINANCIERAS - MAPEO CUENTAS CONTABLES
# =============================================
@api_router.post("/cuentas-financieras/mapear-cuentas-default")
async def mapear_cuentas_default(empresa_id: int = Depends(get_empresa_id)):
    """Auto-map financial accounts to accounting accounts based on name matching."""
    MAPPING = [
        ('BCP', '1041'), ('BBVA', '1042'), ('INTERBANK', '1043'), ('IBK', '1043'), ('CAJA', '101'),
    ]
    pool = await get_pool()
    async with pool.acquire() as conn:
        cuentas_fin = await conn.fetch(
            "SELECT id, nombre FROM finanzas2.cont_cuenta_financiera WHERE empresa_id = $1", empresa_id)
        cuentas_cont = await conn.fetch(
            "SELECT id, codigo FROM finanzas2.cont_cuenta WHERE empresa_id = $1", empresa_id)
        code_map = {r['codigo']: r['id'] for r in cuentas_cont}

        mapped = []
        faltantes = []
        for cf in cuentas_fin:
            nombre_upper = cf['nombre'].upper()
            for keyword, codigo in MAPPING:
                if keyword in nombre_upper:
                    if codigo in code_map:
                        await conn.execute(
                            "UPDATE finanzas2.cont_cuenta_financiera SET cuenta_contable_id = $1 WHERE id = $2",
                            code_map[codigo], cf['id'])
                        mapped.append(f"{cf['nombre']} -> {codigo}")
                    else:
                        faltantes.append(f"Cuenta contable {codigo} no existe para mapear {cf['nombre']}")
                    break
        return {"mapped": mapped, "faltantes": faltantes}

# =============================================
# ASIENTOS CONTABLES (Journal Entries)
# =============================================
@api_router.post("/asientos/generar")
async def generar_asiento(data: GenerarAsientoRequest, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            if data.origen_tipo == 'FPROV':
                asiento = await generar_asiento_fprov(conn, empresa_id, data.origen_id)
            elif data.origen_tipo == 'GASTO':
                asiento = await generar_asiento_gasto(conn, empresa_id, data.origen_id)
            elif data.origen_tipo == 'PAGO':
                asiento = await generar_asiento_pago(conn, empresa_id, data.origen_id)
            else:
                raise HTTPException(400, f"origen_tipo no soportado: {data.origen_tipo}")
            # Return with lines
            lineas = await conn.fetch("""
                SELECT al.*, cc.codigo as cuenta_codigo, cc.nombre as cuenta_nombre
                FROM finanzas2.cont_asiento_linea al
                JOIN finanzas2.cont_cuenta cc ON al.cuenta_id = cc.id
                WHERE al.asiento_id = $1 ORDER BY al.id
            """, asiento['id'])
            result = dict(asiento)
            result['lineas'] = [dict(l) for l in lineas]
            result['total_debe'] = round(sum(float(l['debe']) for l in lineas), 2)
            result['total_haber'] = round(sum(float(l['haber']) for l in lineas), 2)
            return result

@api_router.post("/asientos/{asiento_id}/postear")
async def postear_asiento(asiento_id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        asiento = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_asiento WHERE id = $1 AND empresa_id = $2",
            asiento_id, empresa_id)
        if not asiento:
            raise HTTPException(404, "Asiento no encontrado")
        if asiento['estado'] == 'posteado':
            raise HTTPException(400, "El asiento ya está posteado")
        if asiento['estado'] == 'anulado':
            raise HTTPException(400, "No se puede postear un asiento anulado")

        await check_periodo_cerrado(conn, empresa_id, asiento['fecha_contable'])

        # Validate balance
        sums = await conn.fetchrow("""
            SELECT COALESCE(SUM(debe), 0) as total_debe, COALESCE(SUM(haber), 0) as total_haber
            FROM finanzas2.cont_asiento_linea WHERE asiento_id = $1
        """, asiento_id)
        if abs(float(sums['total_debe']) - float(sums['total_haber'])) > 0.01:
            raise HTTPException(400, "Asiento descuadrado")

        await conn.execute("""
            UPDATE finanzas2.cont_asiento SET estado = 'posteado', updated_at = NOW()
            WHERE id = $1
        """, asiento_id)
        return {"ok": True, "message": "Asiento posteado"}

@api_router.post("/asientos/{asiento_id}/anular")
async def anular_asiento(asiento_id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        asiento = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_asiento WHERE id = $1 AND empresa_id = $2",
            asiento_id, empresa_id)
        if not asiento:
            raise HTTPException(404, "Asiento no encontrado")
        await conn.execute("""
            UPDATE finanzas2.cont_asiento SET estado = 'anulado', updated_at = NOW()
            WHERE id = $1
        """, asiento_id)
        return {"ok": True, "message": "Asiento anulado"}

@api_router.get("/asientos")
async def list_asientos(
    empresa_id: int = Depends(get_empresa_id),
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    cuenta_id: Optional[int] = None,
    tercero_id: Optional[int] = None,
    estado: Optional[str] = None,
    origen_tipo: Optional[str] = None,
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        conditions = ["a.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        if desde:
            conditions.append(f"a.fecha_contable >= ${idx}"); params.append(desde); idx += 1
        if hasta:
            conditions.append(f"a.fecha_contable <= ${idx}"); params.append(hasta); idx += 1
        if estado:
            conditions.append(f"a.estado = ${idx}"); params.append(estado); idx += 1
        if origen_tipo:
            conditions.append(f"a.origen_tipo = ${idx}"); params.append(origen_tipo); idx += 1
        if cuenta_id:
            conditions.append(f"a.id IN (SELECT asiento_id FROM finanzas2.cont_asiento_linea WHERE cuenta_id = ${idx})")
            params.append(cuenta_id); idx += 1
        if tercero_id:
            conditions.append(f"a.id IN (SELECT asiento_id FROM finanzas2.cont_asiento_linea WHERE tercero_id = ${idx})")
            params.append(tercero_id); idx += 1

        rows = await conn.fetch(f"""
            SELECT a.*,
                   (SELECT COALESCE(SUM(debe),0) FROM finanzas2.cont_asiento_linea WHERE asiento_id = a.id) as total_debe,
                   (SELECT COALESCE(SUM(haber),0) FROM finanzas2.cont_asiento_linea WHERE asiento_id = a.id) as total_haber
            FROM finanzas2.cont_asiento a
            WHERE {' AND '.join(conditions)}
            ORDER BY a.fecha_contable DESC, a.id DESC
        """, *params)
        return [dict(r) for r in rows]

@api_router.get("/asientos/{asiento_id}")
async def get_asiento(asiento_id: int, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        asiento = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_asiento WHERE id = $1 AND empresa_id = $2",
            asiento_id, empresa_id)
        if not asiento:
            raise HTTPException(404, "Asiento no encontrado")
        lineas = await conn.fetch("""
            SELECT al.*, cc.codigo as cuenta_codigo, cc.nombre as cuenta_nombre,
                   t.nombre as tercero_nombre
            FROM finanzas2.cont_asiento_linea al
            JOIN finanzas2.cont_cuenta cc ON al.cuenta_id = cc.id
            LEFT JOIN finanzas2.cont_tercero t ON al.tercero_id = t.id
            WHERE al.asiento_id = $1 ORDER BY al.id
        """, asiento_id)
        result = dict(asiento)
        result['lineas'] = [dict(l) for l in lineas]
        result['total_debe'] = round(sum(float(l['debe']) for l in lineas), 2)
        result['total_haber'] = round(sum(float(l['haber']) for l in lineas), 2)
        return result

# =============================================
# REPORTES CONTABLES
# =============================================
@api_router.get("/reportes/mayor")
async def api_reporte_mayor(
    empresa_id: int = Depends(get_empresa_id),
    cuenta_id: Optional[int] = None,
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        data = await reporte_mayor(conn, empresa_id, cuenta_id, desde, hasta)
        return data

@api_router.get("/reportes/balance")
async def api_reporte_balance(
    empresa_id: int = Depends(get_empresa_id),
    hasta: Optional[date] = None,
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        return await reporte_balance(conn, empresa_id, hasta)

@api_router.get("/reportes/pnl")
async def api_reporte_pnl(
    empresa_id: int = Depends(get_empresa_id),
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        return await reporte_pnl(conn, empresa_id, desde, hasta)

# =============================================
# PERIODOS CONTABLES
# =============================================
@api_router.get("/periodos-contables")
async def list_periodos(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT * FROM finanzas2.cont_periodo_cerrado
            WHERE empresa_id = $1 ORDER BY anio DESC, mes DESC
        """, empresa_id)
        return [dict(r) for r in rows]

@api_router.post("/periodos-contables/cerrar")
async def cerrar_periodo(anio: int = Query(...), mes: int = Query(...), empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Check all asientos in period are posteado
        pendientes = await conn.fetchval("""
            SELECT COUNT(*) FROM finanzas2.cont_asiento
            WHERE empresa_id = $1 AND estado = 'borrador'
              AND EXTRACT(YEAR FROM fecha_contable) = $2
              AND EXTRACT(MONTH FROM fecha_contable) = $3
        """, empresa_id, anio, mes)
        if pendientes > 0:
            raise HTTPException(400, f"Hay {pendientes} asiento(s) en borrador en el periodo {anio}-{mes:02d}")
        await conn.execute("""
            INSERT INTO finanzas2.cont_periodo_cerrado (empresa_id, anio, mes, cerrado, cerrado_at)
            VALUES ($1, $2, $3, true, NOW())
            ON CONFLICT (empresa_id, anio, mes) DO UPDATE SET cerrado = true, cerrado_at = NOW()
        """, empresa_id, anio, mes)
        return {"ok": True, "message": f"Periodo {anio}-{mes:02d} cerrado"}

@api_router.post("/periodos-contables/abrir")
async def abrir_periodo(anio: int = Query(...), mes: int = Query(...), empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("""
            UPDATE finanzas2.cont_periodo_cerrado SET cerrado = false
            WHERE empresa_id = $1 AND anio = $2 AND mes = $3
        """, empresa_id, anio, mes)
        return {"ok": True, "message": f"Periodo {anio}-{mes:02d} abierto"}

@api_router.get("/export/compraapp")
async def export_compraapp(
    empresa_id: int = Depends(get_empresa_id),
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
):
    """Export purchases (facturas proveedor + gastos) in compraAPP Excel format with voucher columns"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from re import sub as re_sub

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")

        # Build date filters
        fp_conditions = ["fp.empresa_id = $1"]
        g_conditions = ["g.empresa_id = $1"]
        params_fp = [empresa_id]
        params_g = [empresa_id]
        idx_fp = 2
        idx_g = 2

        if desde:
            fp_conditions.append(f"fp.fecha_factura >= ${idx_fp}")
            params_fp.append(desde)
            idx_fp += 1
            g_conditions.append(f"g.fecha >= ${idx_g}")
            params_g.append(desde)
            idx_g += 1
        if hasta:
            fp_conditions.append(f"fp.fecha_factura <= ${idx_fp}")
            params_fp.append(hasta)
            idx_fp += 1
            g_conditions.append(f"g.fecha <= ${idx_g}")
            params_g.append(hasta)
            idx_g += 1

        # Fetch facturas proveedor (include id, fecha_contable, vou_numero, category accounts)
        facturas = await conn.fetch(f"""
            SELECT fp.id, fp.numero, fp.fecha_factura, fp.fecha_contable, fp.fecha_vencimiento,
                   fp.tipo_comprobante_sunat, fp.base_gravada, fp.igv_sunat,
                   fp.base_no_gravada, fp.isc, fp.total, fp.vou_numero, fp.saldo_pendiente,
                   fp.tipo_cambio, fp.notas, fp.proveedor_id,
                   t.numero_documento as proveedor_doc, t.nombre as proveedor_nombre,
                   t.tipo_persona, t.tip_doc_iden, t.apellido1, t.apellido2, t.nombres as prov_nombres,
                   m.codigo as moneda_codigo
            FROM finanzas2.cont_factura_proveedor fp
            LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE {' AND '.join(fp_conditions)}
            ORDER BY COALESCE(fp.fecha_contable, fp.fecha_factura), fp.id
        """, *params_fp)

        # Fetch gastos (include id, fecha_contable, vou_numero)
        gastos = await conn.fetch(f"""
            SELECT g.id, g.numero_documento, g.fecha, g.fecha_contable,
                   g.tipo_comprobante_sunat, g.base_gravada, g.igv_sunat,
                   g.base_no_gravada, g.isc, g.total, g.vou_numero,
                   g.tipo_cambio, g.pago_id, g.notas, g.proveedor_id,
                   t.numero_documento as proveedor_doc, t.nombre as proveedor_nombre,
                   t.tipo_persona, t.tip_doc_iden, t.apellido1, t.apellido2, t.nombres as prov_nombres,
                   m.codigo as moneda_codigo
            FROM finanzas2.cont_gasto g
            LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
            WHERE {' AND '.join(g_conditions)}
            ORDER BY COALESCE(g.fecha_contable, g.fecha), g.id
        """, *params_g)

        # Fetch retencion details
        fp_ids_all = [f['id'] for f in facturas]
        g_ids_all = [g['id'] for g in gastos]
        ret_map = {}
        if fp_ids_all:
            ret_rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_retencion_detalle
                WHERE empresa_id = $1 AND origen_tipo = 'FPROV' AND origen_id = ANY($2)
            """, empresa_id, fp_ids_all)
            for r in ret_rows:
                ret_map[('FPROV', r['origen_id'])] = dict(r)
        if g_ids_all:
            ret_rows = await conn.fetch("""
                SELECT * FROM finanzas2.cont_retencion_detalle
                WHERE empresa_id = $1 AND origen_tipo = 'GASTO' AND origen_id = ANY($2)
            """, empresa_id, g_ids_all)
            for r in ret_rows:
                ret_map[('GASTO', r['origen_id'])] = dict(r)

        # Fetch pago aplicaciones for facturas
        fp_pagos_map = {}
        if fp_ids_all:
            pago_rows = await conn.fetch("""
                SELECT pa.documento_id, pa.monto_aplicado,
                       p.fecha as pago_fecha, p.numero as pago_numero, p.notas as pago_notas,
                       pd.medio_pago, pd.referencia,
                       cf.cuenta_contable_id, cc.codigo as cuenta_contable_codigo,
                       pm.codigo as pago_moneda_codigo
                FROM finanzas2.cont_pago_aplicacion pa
                JOIN finanzas2.cont_pago p ON p.id = pa.pago_id
                LEFT JOIN finanzas2.cont_pago_detalle pd ON pd.pago_id = p.id
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON cf.id = pd.cuenta_financiera_id
                LEFT JOIN finanzas2.cont_cuenta cc ON cc.id = cf.cuenta_contable_id
                LEFT JOIN finanzas2.cont_moneda pm ON pm.id = p.moneda_id
                WHERE pa.tipo_documento = 'factura' AND pa.documento_id = ANY($1) AND p.empresa_id = $2
            """, fp_ids_all, empresa_id)
            for pr in pago_rows:
                fp_pagos_map.setdefault(pr['documento_id'], []).append(dict(pr))

        g_pagos_map = {}
        if g_ids_all:
            gpago_rows = await conn.fetch("""
                SELECT g.id as gasto_id, p.fecha as pago_fecha, p.numero as pago_numero, p.notas as pago_notas,
                       pd.medio_pago, pd.referencia, pd.monto,
                       cf.cuenta_contable_id, cc.codigo as cuenta_contable_codigo,
                       pm.codigo as pago_moneda_codigo
                FROM finanzas2.cont_gasto g
                JOIN finanzas2.cont_pago p ON p.id = g.pago_id
                LEFT JOIN finanzas2.cont_pago_detalle pd ON pd.pago_id = p.id
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON cf.id = pd.cuenta_financiera_id
                LEFT JOIN finanzas2.cont_cuenta cc ON cc.id = cf.cuenta_contable_id
                LEFT JOIN finanzas2.cont_moneda pm ON pm.id = p.moneda_id
                WHERE g.id = ANY($1) AND g.empresa_id = $2 AND g.pago_id IS NOT NULL
            """, g_ids_all, empresa_id)
            for gpr in gpago_rows:
                g_pagos_map.setdefault(gpr['gasto_id'], []).append(dict(gpr))

        # Fetch config contable and build account lookup
        config_row = await conn.fetchrow(
            "SELECT * FROM finanzas2.cont_config_empresa WHERE empresa_id = $1", empresa_id
        )
        config = dict(config_row) if config_row else {}

        # Build account code lookup
        cta_ids = set(filter(None, [
            config.get('cta_gastos_default_id'),
            config.get('cta_igv_default_id'),
            config.get('cta_xpagar_default_id'),
            config.get('cta_otrib_default_id'),
        ]))
        # Also fetch category account ids for facturas
        fp_ids = [f['id'] for f in facturas]
        cat_account_map = {}  # factura_id -> cuenta_codigo
        fp_ccosto_map = {}    # factura_id -> centro_costo codigo (or 'MIX')
        fp_presup_map = {}    # factura_id -> presupuesto nombre (or 'MIX')
        if fp_ids:
            cat_rows = await conn.fetch("""
                SELECT DISTINCT fpl.factura_id, cc.codigo as cta_codigo, cc.id as cta_id
                FROM finanzas2.cont_factura_proveedor_linea fpl
                LEFT JOIN finanzas2.cont_categoria cat ON fpl.categoria_id = cat.id
                LEFT JOIN finanzas2.cont_cuenta cc ON cat.cuenta_gasto_id = cc.id
                WHERE fpl.factura_id = ANY($1) AND cc.id IS NOT NULL
            """, fp_ids)
            for cr in cat_rows:
                cat_account_map[cr['factura_id']] = cr['cta_codigo']

            # Centro costo lookup for facturas
            cc_rows = await conn.fetch("""
                SELECT fpl.factura_id, cco.codigo as cc_codigo
                FROM finanzas2.cont_factura_proveedor_linea fpl
                JOIN finanzas2.cont_centro_costo cco ON fpl.centro_costo_id = cco.id
                WHERE fpl.factura_id = ANY($1) AND fpl.centro_costo_id IS NOT NULL
            """, fp_ids)
            for r in cc_rows:
                fid = r['factura_id']
                if fid not in fp_ccosto_map:
                    fp_ccosto_map[fid] = r['cc_codigo']
                elif fp_ccosto_map[fid] != r['cc_codigo']:
                    fp_ccosto_map[fid] = 'MIX'

            # Presupuesto lookup for facturas
            pr_rows = await conn.fetch("""
                SELECT fpl.factura_id, p.nombre as pr_nombre
                FROM finanzas2.cont_factura_proveedor_linea fpl
                JOIN finanzas2.cont_presupuesto p ON fpl.presupuesto_id = p.id
                WHERE fpl.factura_id = ANY($1) AND fpl.presupuesto_id IS NOT NULL
            """, fp_ids)
            for r in pr_rows:
                fid = r['factura_id']
                if fid not in fp_presup_map:
                    fp_presup_map[fid] = r['pr_nombre']
                elif fp_presup_map[fid] != r['pr_nombre']:
                    fp_presup_map[fid] = 'MIX'

        # Same lookups for gastos
        g_ids = [g['id'] for g in gastos]
        gasto_cat_account_map = {}
        g_ccosto_map = {}
        g_presup_map = {}
        if g_ids:
            gcat_rows = await conn.fetch("""
                SELECT DISTINCT gl.gasto_id, cc.codigo as cta_codigo
                FROM finanzas2.cont_gasto_linea gl
                LEFT JOIN finanzas2.cont_categoria cat ON gl.categoria_id = cat.id
                LEFT JOIN finanzas2.cont_cuenta cc ON cat.cuenta_gasto_id = cc.id
                WHERE gl.gasto_id = ANY($1) AND cc.id IS NOT NULL
            """, g_ids)
            for gr in gcat_rows:
                gasto_cat_account_map[gr['gasto_id']] = gr['cta_codigo']

            # Centro costo lookup for gastos
            gcc_rows = await conn.fetch("""
                SELECT gl.gasto_id, cco.codigo as cc_codigo
                FROM finanzas2.cont_gasto_linea gl
                JOIN finanzas2.cont_centro_costo cco ON gl.centro_costo_id = cco.id
                WHERE gl.gasto_id = ANY($1) AND gl.centro_costo_id IS NOT NULL
            """, g_ids)
            for r in gcc_rows:
                gid = r['gasto_id']
                if gid not in g_ccosto_map:
                    g_ccosto_map[gid] = r['cc_codigo']
                elif g_ccosto_map[gid] != r['cc_codigo']:
                    g_ccosto_map[gid] = 'MIX'

            # Presupuesto lookup for gastos
            gpr_rows = await conn.fetch("""
                SELECT gl.gasto_id, p.nombre as pr_nombre
                FROM finanzas2.cont_gasto_linea gl
                JOIN finanzas2.cont_presupuesto p ON gl.presupuesto_id = p.id
                WHERE gl.gasto_id = ANY($1) AND gl.presupuesto_id IS NOT NULL
            """, g_ids)
            for r in gpr_rows:
                gid = r['gasto_id']
                if gid not in g_presup_map:
                    g_presup_map[gid] = r['pr_nombre']
                elif g_presup_map[gid] != r['pr_nombre']:
                    g_presup_map[gid] = 'MIX'

        # Build account id -> codigo map
        all_cta_ids = list(cta_ids)
        cta_code_map = {}
        if all_cta_ids:
            cta_rows = await conn.fetch(
                "SELECT id, codigo FROM finanzas2.cont_cuenta WHERE id = ANY($1)", all_cta_ids
            )
            cta_code_map = {r['id']: r['codigo'] for r in cta_rows}

        default_cta_gastos = cta_code_map.get(config.get('cta_gastos_default_id'), '')
        default_cta_igv = cta_code_map.get(config.get('cta_igv_default_id'), '')
        default_cta_xpagar = cta_code_map.get(config.get('cta_xpagar_default_id'), '')
        default_cta_otrib = cta_code_map.get(config.get('cta_otrib_default_id'), '')

        # Validate: check for missing required fields
        errors = []
        for i, f in enumerate(facturas):
            if not f['tipo_comprobante_sunat']:
                errors.append(f"Factura '{f['numero']}': falta Doc SUNAT")
            if not f['numero']:
                errors.append(f"Factura #{i+1}: falta Número")
            if not f['proveedor_doc']:
                errors.append(f"Factura '{f['numero']}': falta Código (RUC/DNI) del proveedor")
            if f['moneda_codigo'] == 'USD' and not f['tipo_cambio']:
                errors.append(f"Factura '{f['numero']}': moneda USD sin tipo de cambio")
        for i, g in enumerate(gastos):
            if not g['tipo_comprobante_sunat']:
                errors.append(f"Gasto '{g['numero_documento']}': falta Doc SUNAT")
            if not g['numero_documento']:
                errors.append(f"Gasto #{i+1}: falta Número documento")
            if not g['proveedor_doc']:
                errors.append(f"Gasto '{g['numero_documento']}': falta Código (RUC/DNI) del proveedor")
            if g['moneda_codigo'] == 'USD' and not g['tipo_cambio']:
                errors.append(f"Gasto '{g['numero_documento']}': moneda USD sin tipo de cambio")

        if errors:
            raise HTTPException(400, detail={"message": "Faltan datos obligatorios para la exportación", "errors": errors})

        # Convert to dicts for mutability
        facturas = [dict(r) for r in facturas]
        gastos = [dict(r) for r in gastos]

        # Assign voucher numbers atomically to docs that don't have one yet
        async with conn.transaction():
            for f in facturas:
                if not f['vou_numero']:
                    vou_fecha = f['fecha_contable'] or f['fecha_factura']
                    anio = vou_fecha.year if vou_fecha else date.today().year
                    row = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_correlativos (empresa_id, tipo_documento, prefijo, ultimo_numero, updated_at)
                        VALUES ($1, $2, $3, 1, NOW())
                        ON CONFLICT (empresa_id, tipo_documento, prefijo)
                        DO UPDATE SET ultimo_numero = finanzas2.cont_correlativos.ultimo_numero + 1, updated_at = NOW()
                        RETURNING ultimo_numero
                    """, empresa_id, f'VOU_COMPRAS_{anio}', '01')
                    f['vou_numero'] = f"{row['ultimo_numero']:06d}"
                    await conn.execute(
                        "UPDATE finanzas2.cont_factura_proveedor SET vou_numero = $1 WHERE id = $2",
                        f['vou_numero'], f['id']
                    )

            for g in gastos:
                if not g['vou_numero']:
                    vou_fecha = g['fecha_contable'] or g['fecha']
                    anio = vou_fecha.year if vou_fecha else date.today().year
                    row = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_correlativos (empresa_id, tipo_documento, prefijo, ultimo_numero, updated_at)
                        VALUES ($1, $2, $3, 1, NOW())
                        ON CONFLICT (empresa_id, tipo_documento, prefijo)
                        DO UPDATE SET ultimo_numero = finanzas2.cont_correlativos.ultimo_numero + 1, updated_at = NOW()
                        RETURNING ultimo_numero
                    """, empresa_id, f'VOU_COMPRAS_{anio}', '01')
                    g['vou_numero'] = f"{row['ultimo_numero']:06d}"
                    await conn.execute(
                        "UPDATE finanzas2.cont_gasto SET vou_numero = $1 WHERE id = $2",
                        g['vou_numero'], g['id']
                    )

        # Helper to clean doc number (digits only)
        def clean_doc(doc_str):
            if not doc_str:
                return ""
            return re_sub(r'[^0-9]', '', str(doc_str))

        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CompraAPP"

        # Fixed 61-column order (NEVER change order)
        COLUMNS_61 = [
            "Vou.Origen", "Vou.Numero", "Vou.Fecha", "Doc", "Numero",
            "Fec.Doc", "Fec.Venc.", "Codigo", "B.I.O.G y E. (A)",
            "B.I.O.G.y E. y NO GRA. (B)", "B.I.O.G.sin D.C.FIS(C)",
            "AD. NO GRAV.", "I.S.C.", "IGV (A)", "IGV (B)", "IGV (C)",
            "OTROS TRIB.", "IMP. BOLSA", "Moneda", "TC", "Glosa",
            "Cta Gastos", "Cta IGV", "Cta O. Trib.", "Cta x Pagar",
            "C.Costo", "Presupuesto", "R.Doc", "R.numero", "R.Fecha",
            "D.Numero", "D.Fecha", "RUC", "R.Social", "Tipo",
            "Tip.Doc.Iden", "Medio de Pago", "Apellido 1", "Apellido 2",
            "Nombre", "T.Bien", "P.origen", "P.vou", "P.fecha",
            "P.fecha D.", "P.fecha V.", "P.cta cob", "P.m.pago", "P.doc",
            "P.num doc", "P.moneda", "P.tc", "P.monto", "P.glosa",
            "P.fe", "Retencion 0/1", "PDB ndes", "CodTasa", "Ind.Ret",
            "B.Imp", "IGV",
        ]

        header_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(COLUMNS_61, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row_num = 2
        VOU_ORIGEN = "01"

        def fmt_date(d):
            if d is None:
                return None
            if isinstance(d, str):
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(d[:10], "%Y-%m-%d")
                    return parsed.strftime("%d/%m/%Y")
                except Exception:
                    return d[:10]
            return d.strftime("%d/%m/%Y")

        def fmt_num(val):
            """Return rounded float or None (truly empty cell)."""
            if val is None:
                return None
            v = round(float(val), 2)
            return v if v != 0 else None

        def moneda_tc(moneda_codigo, tipo_cambio):
            """Return (letra, tc) for Moneda/TC columns."""
            if moneda_codigo == 'USD':
                return 'D', round(float(tipo_cambio), 2)
            return 'S', 1.00

        def write_row(ws, row, data_dict):
            """Write a dict keyed by column name into the correct 61-col positions."""
            for col_idx, col_name in enumerate(COLUMNS_61, 1):
                val = data_dict.get(col_name)
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border

        def build_pago_cols(pago_info):
            """Build P.* columns from a pago info dict."""
            if not pago_info:
                return {}
            pm_code = pago_info.get('pago_moneda_codigo')
            pm_letra = 'D' if pm_code == 'USD' else ('S' if pm_code else None)
            pm_tc = 1.00 if pm_letra == 'S' else None
            return {
                "Medio de Pago": pago_info.get('medio_pago') or None,
                "P.fecha": fmt_date(pago_info.get('pago_fecha')),
                "P.fecha D.": fmt_date(pago_info.get('pago_fecha')),
                "P.cta cob": pago_info.get('cuenta_contable_codigo') or None,
                "P.m.pago": pago_info.get('medio_pago') or None,
                "P.num doc": pago_info.get('referencia') or None,
                "P.moneda": pm_letra,
                "P.tc": pm_tc,
                "P.monto": fmt_num(pago_info.get('monto_aplicado') or pago_info.get('monto')),
                "P.glosa": pago_info.get('pago_notas') or None,
            }

        def build_ret_cols(ret):
            """Build retention/deduction columns from ret dict."""
            if not ret:
                return {}
            return {
                "R.Doc": ret.get('r_doc') or None,
                "R.numero": ret.get('r_numero') or None,
                "R.Fecha": fmt_date(ret.get('r_fecha')),
                "D.Numero": ret.get('d_numero') or None,
                "D.Fecha": fmt_date(ret.get('d_fecha')),
                "Retencion 0/1": ret.get('retencion_01'),
                "PDB ndes": ret.get('pdb_ndes') or None,
                "CodTasa": ret.get('codtasa') or None,
                "Ind.Ret": ret.get('ind_ret') or None,
                "B.Imp": fmt_num(ret.get('b_imp')),
                "IGV": fmt_num(ret.get('igv_ret')),
            }

        def build_prov_cols(doc):
            """Build proveedor columns 33-40."""
            return {
                "RUC": clean_doc(doc.get('proveedor_doc')) or None,
                "R.Social": doc.get('proveedor_nombre') or None,
                "Tipo": doc.get('tipo_persona') or None,
                "Tip.Doc.Iden": doc.get('tip_doc_iden') or None,
                "Apellido 1": doc.get('apellido1') or None,
                "Apellido 2": doc.get('apellido2') or None,
                "Nombre": doc.get('prov_nombres') or None,
            }

        # Write facturas proveedor
        for f in facturas:
            f = dict(f) if not isinstance(f, dict) else f
            vou_fecha = f.get('fecha_contable') or f.get('fecha_factura')
            cta_gasto = cat_account_map.get(f['id'], default_cta_gastos) or None
            igv_val = fmt_num(f['igv_sunat'])
            cta_igv = default_cta_igv if igv_val else None
            saldo = float(f.get('saldo_pendiente') or 0)
            cta_xpagar = default_cta_xpagar if saldo > 0 else None
            m_letra, m_tc = moneda_tc(f.get('moneda_codigo'), f.get('tipo_cambio'))
            glosa = f.get('notas') or ''
            if not glosa.strip():
                doc_str = f.get('tipo_comprobante_sunat') or ''
                num_str = f.get('numero') or ''
                glosa = f"{f.get('proveedor_nombre', '')} {doc_str}-{num_str}".strip()
            isc_val = fmt_num(f['isc'])
            cta_otrib = default_cta_otrib if isc_val else None
            ccosto = fp_ccosto_map.get(f['id']) or None
            presupuesto = fp_presup_map.get(f['id']) or None

            base_row = {
                "Vou.Origen": VOU_ORIGEN,
                "Vou.Numero": f.get('vou_numero') or None,
                "Vou.Fecha": fmt_date(vou_fecha),
                "Doc": f['tipo_comprobante_sunat'] or None,
                "Numero": f['numero'] or None,
                "Fec.Doc": fmt_date(f['fecha_factura']),
                "Fec.Venc.": fmt_date(f['fecha_vencimiento']),
                "Codigo": clean_doc(f['proveedor_doc']) or None,
                "B.I.O.G y E. (A)": fmt_num(f['base_gravada']),
                "AD. NO GRAV.": fmt_num(f['base_no_gravada']),
                "I.S.C.": isc_val,
                "IGV (A)": igv_val,
                "Moneda": m_letra,
                "TC": m_tc,
                "Glosa": glosa or None,
                "Cta Gastos": cta_gasto,
                "Cta IGV": cta_igv,
                "Cta O. Trib.": cta_otrib,
                "Cta x Pagar": cta_xpagar,
                "C.Costo": ccosto,
                "Presupuesto": presupuesto,
            }
            base_row.update(build_prov_cols(f))
            base_row.update(build_ret_cols(ret_map.get(('FPROV', f['id']))))

            pagos = fp_pagos_map.get(f['id'], [])
            if pagos:
                for pago in pagos:
                    row_data = {**base_row, **build_pago_cols(pago)}
                    write_row(ws, row_num, row_data)
                    row_num += 1
            else:
                write_row(ws, row_num, base_row)
                row_num += 1

        # Write gastos
        for g in gastos:
            g = dict(g) if not isinstance(g, dict) else g
            vou_fecha = g.get('fecha_contable') or g.get('fecha')
            cta_gasto = gasto_cat_account_map.get(g['id'], default_cta_gastos) or None
            igv_val = fmt_num(g['igv_sunat'])
            cta_igv = default_cta_igv if igv_val else None
            has_pago = g.get('pago_id') is not None
            cta_xpagar = None if has_pago else (default_cta_xpagar if float(g.get('total') or 0) > 0 else None)
            m_letra, m_tc = moneda_tc(g.get('moneda_codigo'), g.get('tipo_cambio'))
            glosa = g.get('notas') or ''
            if not glosa.strip():
                doc_str = g.get('tipo_comprobante_sunat') or ''
                num_str = g.get('numero_documento') or ''
                glosa = f"{g.get('proveedor_nombre', '')} {doc_str}-{num_str}".strip()
            isc_val = fmt_num(g['isc'])
            cta_otrib = default_cta_otrib if isc_val else None
            ccosto = g_ccosto_map.get(g['id']) or None
            presupuesto = g_presup_map.get(g['id']) or None

            base_row = {
                "Vou.Origen": VOU_ORIGEN,
                "Vou.Numero": g.get('vou_numero') or None,
                "Vou.Fecha": fmt_date(vou_fecha),
                "Doc": g['tipo_comprobante_sunat'] or None,
                "Numero": g['numero_documento'] or None,
                "Fec.Doc": fmt_date(g['fecha']),
                "Codigo": clean_doc(g['proveedor_doc']) or None,
                "B.I.O.G y E. (A)": fmt_num(g['base_gravada']),
                "AD. NO GRAV.": fmt_num(g['base_no_gravada']),
                "I.S.C.": isc_val,
                "IGV (A)": igv_val,
                "Moneda": m_letra,
                "TC": m_tc,
                "Glosa": glosa or None,
                "Cta Gastos": cta_gasto,
                "Cta IGV": cta_igv,
                "Cta O. Trib.": cta_otrib,
                "Cta x Pagar": cta_xpagar,
                "C.Costo": ccosto,
                "Presupuesto": presupuesto,
            }
            base_row.update(build_prov_cols(g))
            base_row.update(build_ret_cols(ret_map.get(('GASTO', g['id']))))

            pagos = g_pagos_map.get(g['id'], [])
            if pagos:
                for pago in pagos:
                    row_data = {**base_row, **build_pago_cols(pago)}
                    write_row(ws, row_num, row_data)
                    row_num += 1
            else:
                write_row(ws, row_num, base_row)
                row_num += 1

        # Auto-adjust column widths for 61 columns
        for i in range(1, 62):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"CompraAPP_{empresa_id}"
        if desde:
            filename += f"_{desde}"
        if hasta:
            filename += f"_{hasta}"
        filename += ".xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

# Include router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
